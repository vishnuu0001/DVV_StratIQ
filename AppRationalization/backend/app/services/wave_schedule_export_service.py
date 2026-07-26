# ---------------------------------------------------------------------------
# Author: Vishnuu A
# Scope: Formatted Excel export of a calculated Harmonization Wave Schedule —
#        mirrors the sheet structure of the BASF_Harmonization_Wave_Gantt_
#        Schedule.xlsx reference workbook (Wave_Summary / Gantt_Schedule /
#        Gantt_View), plus the Wave Plan pipeline detail this app added.
# Date: 2026-07-21
# ---------------------------------------------------------------------------
"""Formatted Excel export of a calculated Harmonization Wave Schedule."""
import io
from datetime import date, datetime
from typing import Any, Dict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.wave_schedule import WaveSchedule

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEADER_FILL = PatternFill("solid", fgColor="C6D9EC")
SUBHEADER_FONT = Font(bold=True)
TOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")
TOTAL_FONT = Font(bold=True)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STAGE_FILLS = {
    "initiation": PatternFill("solid", fgColor="9DC3E6"),
    "assessment": PatternFill("solid", fgColor="2E75B6"),
    "migration": PatternFill("solid", fgColor="1F4E79"),
    "testing": PatternFill("solid", fgColor="548235"),
    "stabilisation": PatternFill("solid", fgColor="A9D18E"),
}
MILESTONE_FILL = PatternFill("solid", fgColor="FFD966")
WAVE_BAND_FILL = PatternFill("solid", fgColor="C6D9EC")


# Function: _autosize
def _autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# Function: _header_row
def _header_row(ws, row, headers):
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


# Function: _iso_to_date
def _iso_to_date(value):
    if not value:
        return None
    return datetime.fromisoformat(value).date()


# Function: _build_wave_summary_sheet
def _build_wave_summary_sheet(wb: Workbook, schedule: Dict[str, Any]) -> None:
    ws = wb.active
    ws.title = "Wave Summary"
    ws.cell(row=1, column=1, value=f"Harmonization Wave Summary — {schedule['topic']}").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=(
        f"Programme {schedule['program_start']} to {schedule['program_end']} · "
        f"AI review: {schedule.get('model_used') or 'n/a'} "
        f"({'succeeded' if schedule.get('llm_available') else 'unavailable — rule-based only'})"
    )).font = Font(italic=True, color="666666")

    headers = ["Wave", "Start", "Cutover", "Stabilisation Ends", "Gate Review", "Applications",
               "Effort (hrs)", "Quick Wins", "Topics", "Simple", "Medium", "Complex", "Very Complex",
               "Permitted Complexity", "AI Theme", "AI Rationale"]
    header_row = 4
    _header_row(ws, header_row, headers)

    row = header_row + 1
    total_apps = total_effort = total_qw = 0
    for w in schedule["waves"]:
        values = [
            f"Wave {w['wave_number']}", _iso_to_date(w["start_date"]), _iso_to_date(w["cutover_date"]),
            _iso_to_date(w["stabilisation_end_date"]), _iso_to_date(w["gate_review_date"]),
            w["application_count"], round(w["effort_hours"], 1), w["quick_win_count"], w["topic_count"],
            w["simple_count"], w["medium_count"], w["complex_count"], w["very_complex_count"],
            w["permitted_complexity"], w.get("theme") or "", w.get("rationale") or "",
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = BORDER
            if col in (2, 3, 4, 5) and value is not None:
                cell.number_format = "yyyy-mm-dd"
        total_apps += w["application_count"]
        total_effort += w["effort_hours"]
        total_qw += w["quick_win_count"]
        row += 1

    total_row = row
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=6, value=total_apps)
    ws.cell(row=total_row, column=7, value=round(total_effort, 1))
    ws.cell(row=total_row, column=8, value=total_qw)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        cell.border = BORDER

    _autosize(ws, [10, 12, 12, 16, 13, 12, 11, 10, 8, 8, 8, 8, 12, 24, 40, 40])


# Function: _write_wave_plan_app_row
def _write_wave_plan_app_row(ws, row: int, a: Dict[str, Any]) -> None:
    values = [
        a["topic"], f"Wave {a['wave_number']}", a["app_id"], a["application_name"], a["complexity"],
        a["tshirt_size"], a["migration_type"], "Yes" if a["quick_win"] else "No",
        round(a["effort_hours"], 1) if a["effort_hours"] is not None else None,
        a.get("assessment_sprint"),
        f"{a['migration_sprint_start']}-{a['migration_sprint_end']}" if a.get("migration_sprint_start") else None,
        a.get("qa_uat_sprint"), a.get("go_live_pi"), a.get("stabilization_pi"), a.get("decommissioning_pi"),
        "AI-reviewed" if a.get("source") == "llm" else "Rule-based", a.get("rationale") or "",
    ]
    for col, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = BORDER
        if col == 16 and value == "AI-reviewed":
            cell.font = Font(color="1F6B3A", bold=True)


# Function: _write_wave_plan_deferred_section
def _write_wave_plan_deferred_section(ws, row: int, deferred: list) -> int:
    """Write the DEFERRED section header + rows starting at `row`. Returns the next free row."""
    row += 1
    ws.cell(row=row, column=1, value=f"DEFERRED — {len(deferred)} application(s) exceeded the programme ceiling").font = Font(bold=True, color="B7472A")
    row += 1
    for a in deferred:
        ws.cell(row=row, column=1, value=a["topic"])
        ws.cell(row=row, column=3, value=a["app_id"])
        ws.cell(row=row, column=4, value=a["application_name"])
        ws.cell(row=row, column=17, value=a.get("deferred_reason") or "")
        row += 1
    return row


# Function: _build_wave_plan_sheet
def _build_wave_plan_sheet(wb: Workbook, schedule: Dict[str, Any]) -> None:
    ws = wb.create_sheet("Wave Plan")
    headers = ["Topic", "Wave", "App ID", "Application Name", "Complexity", "T-Shirt", "Migration Type",
               "Quick Win", "Effort (hrs)", "Assessment Sprint", "Migration Sprints", "QA/UAT Sprint",
               "Go-Live PI", "Stabilization", "Decommissioning", "Source", "Rationale"]
    _header_row(ws, 1, headers)

    apps = [a for a in schedule["apps"] if a["wave_number"] is not None]
    apps.sort(key=lambda a: ((a["topic"] or ""), a["wave_number"], a["app_id"]))

    row = 2
    for a in apps:
        _write_wave_plan_app_row(ws, row, a)
        row += 1

    deferred = [a for a in schedule["apps"] if a["wave_number"] is None]
    if deferred:
        row = _write_wave_plan_deferred_section(ws, row, deferred)

    _autosize(ws, [28, 9, 12, 30, 12, 8, 16, 9, 11, 12, 14, 11, 9, 20, 14, 13, 45])


# Function: _write_task_row
def _write_task_row(ws, row: int, t: Dict[str, Any]) -> None:
    is_wave_header = t["task_type"] == "wave_header"
    values = [
        t["wbs_code"], ("  " if not is_wave_header else "") + t["task_name"], t["wave_number"],
        _iso_to_date(t["start_date"]), _iso_to_date(t["end_date"]), t["duration_days"],
        t.get("predecessor_wbs") or "", "Yes" if t["is_milestone"] else "",
        t.get("applications") if t.get("applications") is not None else "",
        round(t["effort_hours"], 1) if t.get("effort_hours") is not None else "",
    ]
    for col, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = BORDER
        if col in (4, 5) and value is not None:
            cell.number_format = "yyyy-mm-dd"
        if is_wave_header:
            cell.fill = SUBHEADER_FILL
            cell.font = SUBHEADER_FONT


# Function: _build_task_list_sheet
def _build_task_list_sheet(wb: Workbook, schedule: Dict[str, Any]) -> None:
    ws = wb.create_sheet("Task List (WBS)")
    headers = ["WBS", "Task", "Wave", "Start", "Finish", "Duration (days)", "Predecessor",
               "Milestone", "Applications", "Effort (hrs)"]
    _header_row(ws, 1, headers)

    row = 2
    for t in schedule["tasks"]:
        _write_task_row(ws, row, t)
        row += 1

    _autosize(ws, [10, 34, 7, 13, 13, 15, 12, 11, 12, 12])


# Function: _add_months
def _add_months(d: date, months: int) -> date:
    total = d.month - 1 + months
    year = d.year + total // 12
    month = total % 12 + 1
    return date(year, month, 1)


# Function: _write_gantt_header
def _write_gantt_header(ws, header_row: int, months: list) -> None:
    ws.cell(row=header_row, column=1, value="Wave").fill = HEADER_FILL
    ws.cell(row=header_row, column=1).font = HEADER_FONT
    ws.cell(row=header_row, column=1).border = BORDER
    for col, m in enumerate(months, 2):
        cell = ws.cell(row=header_row, column=col, value=m.strftime("%b %y"))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER


# Function: _best_stage_fill_for_month
def _best_stage_fill_for_month(stage_tasks: list, m: date, month_end: date):
    best_fill, best_days = None, 0
    for t in stage_tasks:
        fill = STAGE_FILLS.get(t["task_type"])
        t_start, t_end = _iso_to_date(t["start_date"]), _iso_to_date(t["end_date"])
        if not fill or not t_start or not t_end:
            continue
        overlap_days = (min(t_end, month_end) - max(t_start, m)).days
        if overlap_days > best_days:
            best_days, best_fill = overlap_days, fill
    return best_fill


# Function: _paint_wave_stage_fills
def _paint_wave_stage_fills(ws, row: int, months: list, stage_tasks: list) -> None:
    """A stage's fill is assigned per-month by whichever stage occupies
    the most days in that month, not by "last stage painted wins" —
    otherwise a stage that only tips into a month for a few trailing
    days (e.g. Assessment starting on the 27th) blanks out the color
    of the stage that actually dominates it (e.g. Initiation, which
    owned the first 26 days)."""
    for col, m in enumerate(months, 2):
        month_end = _add_months(m, 1)
        best_fill = _best_stage_fill_for_month(stage_tasks, m, month_end)
        if best_fill:
            ws.cell(row=row, column=col).fill = best_fill


# Function: _paint_wave_milestones
def _paint_wave_milestones(ws, row: int, months: list, milestone_tasks: list) -> None:
    """Milestones are drawn after the stage fills, on top."""
    for t in milestone_tasks:
        t_start = _iso_to_date(t["start_date"])
        if not t_start:
            continue
        for col, m in enumerate(months, 2):
            if m <= t_start < _add_months(m, 1):
                mcell = ws.cell(row=row, column=col)
                mcell.fill = MILESTONE_FILL
                mcell.value = (mcell.value or "") + ("♦" if not mcell.value else " ♦")
                mcell.alignment = Alignment(horizontal="center")


# Function: _write_gantt_wave_row
def _write_gantt_wave_row(ws, row: int, w: Dict[str, Any], months: list, tasks_by_wave: Dict[int, list]) -> None:
    wave_number = w["wave_number"]
    label = f"Wave {wave_number} ({w['application_count']} apps)"
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = Font(bold=True)
    cell.border = BORDER
    for col in range(2, len(months) + 2):
        ws.cell(row=row, column=col).border = BORDER

    wave_tasks = tasks_by_wave.get(wave_number, [])
    stage_tasks = [t for t in wave_tasks if not t["is_milestone"]]
    milestone_tasks = [t for t in wave_tasks if t["is_milestone"]]

    _paint_wave_stage_fills(ws, row, months, stage_tasks)
    _paint_wave_milestones(ws, row, months, milestone_tasks)


# Function: _build_gantt_view_sheet
def _build_gantt_view_sheet(wb: Workbook, schedule: Dict[str, Any]) -> None:
    ws = wb.create_sheet("Gantt View")
    program_start = _iso_to_date(schedule["program_start"])
    program_end = _iso_to_date(schedule["program_end"])
    if not program_start or not program_end:
        return

    months = []
    cursor = date(program_start.year, program_start.month, 1)
    while cursor <= program_end:
        months.append(cursor)
        cursor = _add_months(cursor, 1)

    ws.cell(row=1, column=1, value="Wave Delivery Gantt — monthly view").font = Font(bold=True, size=13)
    header_row = 3
    _write_gantt_header(ws, header_row, months)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=2)

    tasks_by_wave: Dict[int, list] = {}
    for t in schedule["tasks"]:
        tasks_by_wave.setdefault(t["wave_number"], []).append(t)

    row = header_row + 1
    for w in schedule["waves"]:
        _write_gantt_wave_row(ws, row, w, months, tasks_by_wave)
        row += 1

    widths = [24] + [9] * len(months)
    _autosize(ws, widths)


# Function: build_wave_schedule_workbook
def build_wave_schedule_workbook(schedule: WaveSchedule) -> io.BytesIO:
    """Build a formatted, multi-sheet .xlsx workbook for a calculated wave
    schedule and return it as an in-memory buffer ready to stream."""
    data = schedule.to_dict(include_detail=True)
    wb = Workbook()
    _build_wave_summary_sheet(wb, data)
    _build_gantt_view_sheet(wb, data)
    _build_wave_plan_sheet(wb, data)
    _build_task_list_sheet(wb, data)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
