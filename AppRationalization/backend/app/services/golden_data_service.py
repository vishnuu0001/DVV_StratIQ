# ---------------------------------------------------------------------------
# Author: Vishnuu A
# Scope: Fetches CorentData and CASTData from the DB, maps them to the
# Date: 2025-11-23
# ---------------------------------------------------------------------------
"""
Golden Data Service
-------------------
Fetches CorentData and CASTData from the DB, maps them to the
'Inputs from Sources' sheet in APRAttributes.xlsx (starting at row 6),
and returns both the populated workbook bytes and a preview payload.

Column mapping (row 5 = header row, row 6+ = data):
  Col 1  APP ID                              corent.app_id
  Col 2  List of Applications                corent.app_name
  Col 3  Server Type                         corent.server_type
  Col 4  Operating System                    corent.operating_system
  Col 5  CPU Core                            corent.cpu_core
  Col 6  Memory                              corent.memory
  Col 7  Internal Storage                    corent.internal_storage
  Col 8  External Storage                    corent.external_storage
  Col 9  Storage Type                        corent.storage_type
  Col 10 DB Storage                          corent.db_storage
  Col 11 DB Engine                           corent.db_engine
  Col 12 Environment (INSTALL TYPE)          corent.environment + " / " + corent.install_type
  Col 13 Virtualization Attributes           corent.virtualization_attributes
  Col 14 Compute/Server HW Architecture      corent.compute_server_hardware_architecture
  Col 15 Application Stability               corent.application_stability
  Col 16 Virtualization State                corent.virtualization_state
  Col 17 Storage Decomposition               corent.storage_decomposition
  Col 18 FLASH Storage Used                  corent.flash_storage_used
  Col 19 CPU Requirement                     corent.cpu_requirement
  Col 20 Memory (RAM) Requirement            corent.memory_ram_requirement
  Col 21 Mainframe Dependency                corent.mainframe_dependency
  Col 22 Desktop Dependency                  corent.desktop_dependency
  Col 23 App OS/Platform Cloud Suitability   corent.app_os_platform_cloud_suitability
  Col 24 Database Cloud Readiness            corent.database_cloud_readiness
  Col 25 Integration Middleware Cloud Ready  corent.integration_middleware_cloud_readiness
  Col 26 Application Architecture            cast.application_architecture
  Col 27 Application Hardware Dependency     corent.application_hardware_dependency
  Col 28 App COTS vs. Non-COTS              corent.app_cots_vs_non_cots
  Col 29 Source Code Availability            cast.source_code_availability
  Col 30 Programming Language               cast.programming_language
  Col 31 Component Coupling                 cast.component_coupling
  Col 32 Cloud Suitability (CORENT+CAST)    corent.cloud_suitability or cast.cloud_suitability
  Col 33 Volume External Dependencies       corent.volume_external_dependencies or cast.volume_external_dependencies
  Col 34 App Service/API Readiness          cast.app_service_api_readiness
  Col 35 App Load Predictability/Elasticity corent.app_load_predictability_elasticity
  Col 36 Degree of Code Protocols           cast.degree_of_code_protocols
  Col 37 Code Design                        cast.code_design
  Col 38 App-Code Complexity/Volume         cast.application_code_complexity_volume
  Col 39 Financially Optimizable HW Usage   corent.financially_optimizable_hardware_usage
  Col 40 Distributed Architecture Design    corent.distributed_architecture_design or cast.distributed_architecture_design
  Col 41 Latency Requirements               corent.latency_requirements
  Col 42 Ubiquitous Access Requirements     corent.ubiquitous_access_requirements
  Col 43-52 (SURVEY)                        blank
  Col 53 No. Production Environments        corent.no_production_environments
  Col 54 No. Non-Production Environments    corent.no_non_production_environments
  Col 55 HA/DR Requirements                 corent.ha_dr_requirements
  Col 56 (SURVEY)                           blank
  Col 57 RTO Requirements                   corent.rto_requirements
  Col 58 RPO Requirements                   corent.rpo_requirements
  Col 59 Deployment Geography               corent.deployment_geography
  Col 60-64 (SURVEY/other)                  blank
"""

import shutil
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

import openpyxl
from openpyxl.styles import PatternFill, Alignment

from app import db
from app.models.corent_data import CorentData
from app.models.cast import CASTData
from app.models.industry_data import IndustryData
from app.models.golden_data import GoldenData

# ── Paths ─────────────────────────────────────────────────────────────────────
_TEMPLATE_PATH  = Path(__file__).resolve().parent.parent.parent / "data" / "APRAttributes.xlsx"
_OUTPUT_DIR     = Path(__file__).resolve().parent.parent.parent / "UpdatedData"
_OUTPUT_PATH    = _OUTPUT_DIR / "APRAttributes.xlsx"
_SHEET_NAME     = "Inputs from Sources"
_DATA_START_ROW = 6
_MAX_PREVIEW    = 500

# ── Colour fills ──────────────────────────────────────────────────────────────
_YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # AI-predicted
_AMBER_FILL  = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # empty survey
_NO_FILL     = PatternFill(fill_type=None)
_CENTER      = Alignment(wrap_text=False, vertical="center")

# ── Column map: (0-based idx, source, corent_attr, cast_attr, header) ────────
# source: "corent" | "cast" | "corent+cast" | "survey" | "corent_env"
_COL_MAP: List[Tuple[int, str, Optional[str], Optional[str], str]] = [
    (0,  "corent",      "app_id",                                   None,                               "APP ID"),
    (1,  "corent",      "app_name",                                 None,                               "List of Applications"),
    (2,  "corent",      "server_type",                              None,                               "Server Type"),
    (3,  "corent",      "operating_system",                         None,                               "Operating System"),
    (4,  "corent",      "cpu_core",                                 None,                               "CPU Core"),
    (5,  "corent",      "memory",                                   None,                               "Memory"),
    (6,  "corent",      "internal_storage",                         None,                               "Internal Storage"),
    (7,  "corent",      "external_storage",                         None,                               "External Storage"),
    (8,  "corent",      "storage_type",                             None,                               "Storage Type"),
    (9,  "corent",      "db_storage",                               None,                               "DB Storage"),
    (10, "corent",      "db_engine",                                None,                               "DB Engine"),
    (11, "corent_env",  None,                                       None,                               "Environment (INSTALL TYPE)"),
    (12, "corent",      "virtualization_attributes",                None,                               "Virtualization Attributes"),
    (13, "corent",      "compute_server_hardware_architecture",     None,                               "Compute / Server Hardware Architecture"),
    (14, "corent",      "application_stability",                    None,                               "Application Stability"),
    (15, "corent",      "virtualization_state",                     None,                               "Virtualization State"),
    (16, "corent",      "storage_decomposition",                    None,                               "Storage Decomposition"),
    (17, "corent",      "flash_storage_used",                       None,                               "FLASH Storage Used"),
    (18, "corent",      "cpu_requirement",                          None,                               "CPU Requirement"),
    (19, "corent",      "memory_ram_requirement",                   None,                               "Memory (RAM) Requirement"),
    (20, "corent",      "mainframe_dependency",                     None,                               "Mainframe Dependency"),
    (21, "corent",      "desktop_dependency",                       None,                               "Desktop Dependency"),
    (22, "corent",      "app_os_platform_cloud_suitability",        None,                               "App OS / Platform Cloud Suitability"),
    (23, "corent",      "database_cloud_readiness",                 None,                               "Database Cloud Readiness"),
    (24, "corent",      "integration_middleware_cloud_readiness",   None,                               "Integration Middleware Cloud Readiness"),
    (25, "cast",        None,                                       "application_architecture",          "Application Architecture"),
    (26, "corent",      "application_hardware_dependency",          None,                               "Application Hardware Dependency"),
    (27, "corent",      "app_cots_vs_non_cots",                     None,                               "App COTS vs. Non-COTS Only"),
    (28, "cast",        None,                                       "source_code_availability",          "Source Code Availability"),
    (29, "cast",        None,                                       "programming_language",              "Programming Language"),
    (30, "cast",        None,                                       "component_coupling",                "Component Coupling"),
    (31, "corent+cast", "cloud_suitability",                        "cloud_suitability",                "Cloud Suitability"),
    (32, "corent+cast", "volume_external_dependencies",             "volume_external_dependencies",     "Volume of External Dependencies"),
    (33, "cast",        None,                                       "app_service_api_readiness",         "App Service / API Readiness"),
    (34, "corent",      "app_load_predictability_elasticity",       None,                               "App Load Predictability / Elasticity"),
    (35, "cast",        None,                                       "degree_of_code_protocols",          "Degree of Code Protocols"),
    (36, "cast",        None,                                       "code_design",                      "Code Design"),
    (37, "cast",        None,                                       "application_code_complexity_volume","Application-Code Complexity / Volume"),
    (38, "corent",      "financially_optimizable_hardware_usage",   None,                               "Financially Optimizable Hardware Usage"),
    (39, "corent+cast", "distributed_architecture_design",          "distributed_architecture_design",  "Distributed Architecture Design or not"),
    (40, "corent",      "latency_requirements",                     None,                               "Latency Requirements"),
    (41, "corent",      "ubiquitous_access_requirements",           None,                               "Ubiquitous Access Requirements"),
    (42, "survey",      None,  None,  "Level of Data Residency Compliance"),
    (43, "survey",      None,  None,  "Data Classification"),
    (44, "survey",      None,  None,  "App Regulatory & Contractual Requirements"),
    (45, "survey",      None,  None,  "Impact Due to Data Loss"),
    (46, "survey",      None,  None,  "Financial Impact Due to Unavailability"),
    (47, "survey",      None,  None,  "Business Criticality"),
    (48, "survey",      None,  None,  "Customer Facing"),
    (49, "survey",      None,  None,  "Application Status & Lifecycle State"),
    (50, "survey",      None,  None,  "Availability Requirements"),
    (51, "survey",      None,  None,  "Support Level"),
    (52, "corent",      "no_production_environments",               None,                               "No. of Production Environments"),
    (53, "corent",      "no_non_production_environments",           None,                               "No. of Non-Production Environments"),
    (54, "corent",      "ha_dr_requirements",                       None,                               "HA/DR Requirements"),
    (55, "survey",      None,  None,  "Business Function Readiness"),
    (56, "corent",      "rto_requirements",                         None,                               "RTO Requirements"),
    (57, "corent",      "rpo_requirements",                         None,                               "RPO Requirements"),
    (58, "corent",      "deployment_geography",                     None,                               "Deployment Geography"),
    (59, "survey",      None,  None,  "Level of Internal Governance"),
    (60, "survey",      None,  None,  "No. of Internal Users"),
    (61, "survey",      None,  None,  "No. of External Users"),
    (62, "survey",      None,  None,  "Estimated App Growth"),
    (63, "survey",      None,  None,  "Impact to Users"),
]

# Maps Excel column index → GoldenData survey field name
_SURVEY_DB: Dict[int, str] = {
    42: "level_of_data_residency_compliance",
    43: "data_classification",
    44: "app_regulatory_contractual_requirements",
    45: "impact_due_to_data_loss",
    46: "financial_impact_due_to_unavailability",
    47: "business_criticality",
    48: "customer_facing",
    49: "application_status_lifecycle_state",
    50: "availability_requirements",
    51: "support_level",
    55: "business_function_readiness",
    59: "level_of_internal_governance",
    60: "no_of_internal_users",
    61: "no_of_external_users",
    62: "estimated_app_growth",
    63: "impact_to_users",
}

# Attribute name differences: CorentData attr → WorkspaceCorentRow attr
_CORENT_WS_ALIAS: Dict[str, str] = {
    "financially_optimizable_hardware_usage": "financially_optimizable_hardware",
}
# Attribute name differences: CASTData attr → WorkspaceCastRow attr
_CAST_WS_ALIAS: Dict[str, str] = {
    "application_code_complexity_volume": "app_code_complexity_volume",
}


# Function: _coalesce
def _coalesce(*values):
    """Return first non-None, non-empty value (as str, preserving ints)."""
    for v in values:
        if v is not None and str(v).strip() not in ("", "None", "null"):
            return v
    return None


# Function: _env_install
def _env_install(row) -> Optional[str]:
    parts = [p for p in (getattr(row, "environment", None), getattr(row, "install_type", None)) if p]
    return " / ".join(parts) if parts else None


# Function: _collect_ai_fill_cols
def _collect_ai_fill_cols(rows) -> Dict[str, Set[str]]:
    """Build {app_id: set(ai-filled column names)} from workspace rows."""
    fills: Dict[str, Set[str]] = {}
    for row in rows:
        if not row.app_id:
            continue
        cols: Set[str] = set()
        if row.updated_rows:
            cols.update(row.updated_rows.get("updated_columns", []))
        if row.ai_predicted_columns:
            cols.update(row.ai_predicted_columns)
        if cols:
            fills[row.app_id] = cols
    return fills


# Function: _get_workspace_ai_fills
def _get_workspace_ai_fills() -> Tuple[Dict[str, Set[str]], Dict[str, Set[str]]]:
    """Query latest workspace run and return AI-filled column sets per app_id."""
    corent_fills: Dict[str, Set[str]] = {}
    cast_fills:   Dict[str, Set[str]] = {}
    try:
        from app.models.correlation_workspace import WorkspaceCorentRow, WorkspaceCastRow, WorkspaceRun
        # status='done' is required here — this must reference the SAME run that
        # _load_corent_rows_or_workspace()/_build_data_indexes() actually display data
        # from (which does filter status='done'). Without this filter, if a newer run
        # exists that is still 'running' or ended 'failed', this AI-fill lookup silently
        # keys off the wrong run_id (one with no/incomplete data), so ai_filled_cols
        # comes back empty for cells that really were AI-populated in the real
        # (older, done) run — the cells then render with no highlight at all. This is
        # the root cause of "AI-populated data not getting highlighted" in Golden Data.
        latest = WorkspaceRun.query.filter_by(status='done').order_by(WorkspaceRun.id.desc()).first()
        if latest is None:
            return corent_fills, cast_fills
        corent_fills = _collect_ai_fill_cols(WorkspaceCorentRow.query.filter_by(run_id=latest.id).all())
        cast_fills   = _collect_ai_fill_cols(WorkspaceCastRow.query.filter_by(run_id=latest.id).all())
    except Exception:
        pass
    return corent_fills, cast_fills


# Function: _is_ai_marker
def _is_ai_marker(val) -> bool:
    return val is not None and str(val).lower().replace("-", " ") == "ai populated value"


# Function: _resolve_corent_env_col
def _resolve_corent_env_col(corent, corent_ai: Set[str], corent_ws) -> Tuple[Any, bool]:
    """Resolve corent_env source column value and AI-fill flag."""
    val = _env_install(corent)
    ai_filled = "environment" in corent_ai or "install_type" in corent_ai
    if not ai_filled and corent_ws is not None:
        if val is None and _env_install(corent_ws) is not None:
            val = _env_install(corent_ws)
            ai_filled = True
    return val, ai_filled


# Function: _resolve_corent_col
def _resolve_corent_col(corent, c_attr: str, corent_ai: Set[str], corent_ws) -> Tuple[Any, bool]:
    """Resolve corent source column value and AI-fill flag."""
    raw = getattr(corent, c_attr, None)
    val = _coalesce(raw)
    ws_col = _CORENT_WS_ALIAS.get(c_attr, c_attr)
    ai_filled = c_attr in corent_ai or ws_col in corent_ai or _is_ai_marker(val)
    if not ai_filled and corent_ws is not None and val is None:
        ws_raw = getattr(corent_ws, ws_col, None) or getattr(corent_ws, c_attr, None)
        ws_val = _coalesce(ws_raw)
        if ws_val is not None:
            val = ws_val
            ai_filled = True
    return val, ai_filled


# Function: _resolve_cast_col
def _resolve_cast_col(cast, k_attr: str, industry, cast_ai: Set[str], cast_ws) -> Tuple[Any, bool]:
    """Resolve cast source column value and AI-fill flag."""
    raw_cast = getattr(cast, k_attr, None) if cast else None
    val = _coalesce(raw_cast)
    if val is None and k_attr == "application_architecture" and industry:
        val = _coalesce(industry.architecture_type)
    ws_col = _CAST_WS_ALIAS.get(k_attr, k_attr) if k_attr else None
    ai_filled = bool(ws_col and (k_attr in cast_ai or ws_col in cast_ai)) or _is_ai_marker(val)
    if not ai_filled and cast_ws is not None and val is None and ws_col:
        ws_raw = getattr(cast_ws, ws_col, None) or (getattr(cast_ws, k_attr, None) if k_attr else None)
        ws_val = _coalesce(ws_raw)
        if ws_val is not None:
            val = ws_val
            ai_filled = True
    return val, ai_filled


# Function: _is_corent_cast_ai_filled
def _is_corent_cast_ai_filled(c_attr, k_attr, ws_c, ws_k, corent_ai: Set[str], cast_ai: Set[str], val) -> bool:
    if c_attr in corent_ai or ws_c in corent_ai:
        return True
    if ws_k and (k_attr in cast_ai or ws_k in cast_ai):
        return True
    return _is_ai_marker(val)


# Function: _lookup_corent_cast_ws_value
def _lookup_corent_cast_ws_value(corent_ws, cast_ws, c_attr, ws_c, k_attr, ws_k):
    if corent_ws is not None:
        ws_val = _coalesce(getattr(corent_ws, ws_c, None) or getattr(corent_ws, c_attr, None))
        if ws_val is not None:
            return ws_val
    if cast_ws is not None and ws_k:
        return _coalesce(getattr(cast_ws, ws_k, None) or (getattr(cast_ws, k_attr, None) if k_attr else None))
    return None


# Function: _resolve_corent_cast_col
def _resolve_corent_cast_col(
    corent, cast, c_attr: str, k_attr: Optional[str],
    corent_ai: Set[str], cast_ai: Set[str], corent_ws, cast_ws,
) -> Tuple[Any, bool]:
    """Resolve corent+cast source column value and AI-fill flag."""
    cv = _coalesce(getattr(corent, c_attr, None))
    kv = _coalesce(getattr(cast, k_attr, None) if cast else None)
    val = _coalesce(cv, kv)
    ws_c = _CORENT_WS_ALIAS.get(c_attr, c_attr)
    ws_k = _CAST_WS_ALIAS.get(k_attr, k_attr) if k_attr else None
    ai_filled = _is_corent_cast_ai_filled(c_attr, k_attr, ws_c, ws_k, corent_ai, cast_ai, val)

    if not ai_filled and val is None:
        ws_val = _lookup_corent_cast_ws_value(corent_ws, cast_ws, c_attr, ws_c, k_attr, ws_k)
        if ws_val is not None:
            val = ws_val
            ai_filled = True

    return val, ai_filled


# Function: _resolve_survey_col
def _resolve_survey_col(idx, golden_rec) -> Tuple[Any, bool]:
    db_field = _SURVEY_DB.get(idx)
    val = getattr(golden_rec, db_field, None) if (db_field and golden_rec) else None
    return val, False


# Function: _resolve_col_value
def _resolve_col_value(
    idx, src, c_attr, k_attr, corent, cast, industry, golden_rec,
    corent_ai: Set[str], cast_ai: Set[str], corent_ws, cast_ws,
) -> Tuple[Any, bool]:
    """Dispatch to the resolver for this column's source type."""
    if src == "corent_env":
        return _resolve_corent_env_col(corent, corent_ai, corent_ws)
    if src == "corent":
        return _resolve_corent_col(corent, c_attr, corent_ai, corent_ws)
    if src == "cast":
        return _resolve_cast_col(cast, k_attr, industry, cast_ai, cast_ws)
    if src == "corent+cast":
        return _resolve_corent_cast_col(corent, cast, c_attr, k_attr, corent_ai, cast_ai, corent_ws, cast_ws)
    if src == "survey":
        return _resolve_survey_col(idx, golden_rec)
    return None, False


# Function: _build_row
def _build_row(
    corent,
    cast,
    industry,
    golden_rec,
    corent_ai: Set[str],
    cast_ai: Set[str],
    corent_ws=None,
    cast_ws=None,
) -> Tuple[List[Any], List[Optional[str]]]:
    """
    Build two parallel 64-element lists: values and colors.
    colors: "yellow" = AI-predicted, "amber" = empty survey col, None = normal.
    """
    values: List[Any] = [None] * 64
    colors: List[Optional[str]] = [None] * 64

    for idx, src, c_attr, k_attr, _header in _COL_MAP:
        val, ai_filled = _resolve_col_value(
            idx, src, c_attr, k_attr, corent, cast, industry, golden_rec,
            corent_ai, cast_ai, corent_ws, cast_ws,
        )

        values[idx] = val
        is_empty = (val is None or str(val).strip() == "")

        if ai_filled:
            colors[idx] = "yellow"
        elif is_empty and src == "survey":
            colors[idx] = "amber"

    return values, colors


# Function: _build_row_values
def _build_row_values(corent_row, cast_row):
    """
    Return a list of 64 values in column order (index 0 = col 1).
    cast_row may be None when no CAST data exists for that app_id.
    """
    c = corent_row
    k = cast_row  # may be None

    # Function: cast
    def cast(attr):
        return getattr(k, attr, None) if k else None

    row = [None] * 64

    row[0]  = str(c.id)
    row[1]  = None
    row[2]  = c.server_type
    row[3]  = c.operating_system
    row[4]  = c.cpu_core
    row[5]  = c.memory
    row[6]  = c.internal_storage
    row[7]  = c.external_storage
    row[8]  = c.storage_type
    row[9]  = c.db_storage
    row[10] = c.db_engine
    row[11] = _env_install(c)
    row[12] = c.virtualization_attributes
    row[13] = c.compute_server_hardware_architecture
    row[14] = c.application_stability
    row[15] = c.virtualization_state
    row[16] = c.storage_decomposition
    row[17] = c.flash_storage_used
    row[18] = c.cpu_requirement
    row[19] = c.memory_ram_requirement
    row[20] = c.mainframe_dependency
    row[21] = c.desktop_dependency
    row[22] = c.app_os_platform_cloud_suitability
    row[23] = c.database_cloud_readiness
    row[24] = c.integration_middleware_cloud_readiness
    row[25] = cast('application_architecture')           # CAST
    row[26] = c.application_hardware_dependency
    row[27] = c.app_cots_vs_non_cots
    row[28] = cast('source_code_availability')           # CAST
    row[29] = cast('programming_language')               # CAST
    row[30] = cast('component_coupling')                 # CAST
    row[31] = _coalesce(c.cloud_suitability, cast('cloud_suitability'))           # CORENT+CAST
    row[32] = _coalesce(c.volume_external_dependencies, cast('volume_external_dependencies'))  # CORENT+CAST
    row[33] = cast('app_service_api_readiness')          # CAST
    row[34] = c.app_load_predictability_elasticity
    row[35] = cast('degree_of_code_protocols')           # CAST
    row[36] = cast('code_design')                        # CAST
    row[37] = cast('application_code_complexity_volume') # CAST
    row[38] = c.financially_optimizable_hardware_usage
    row[39] = _coalesce(c.distributed_architecture_design, cast('distributed_architecture_design'))  # CORENT+CAST
    row[40] = c.latency_requirements
    row[41] = c.ubiquitous_access_requirements
    # 42-51 → SURVEY (blank)
    row[52] = c.no_production_environments
    row[53] = c.no_non_production_environments
    row[54] = c.ha_dr_requirements
    # 55 → SURVEY (blank)
    row[56] = c.rto_requirements
    row[57] = c.rpo_requirements
    row[58] = c.deployment_geography
    # 59-63 → SURVEY/other (blank)

    return row


# Function: _get_column_headers
def _get_column_headers() -> List[str]:
    return [c[4] for c in _COL_MAP]


# Function: _load_corent_rows_or_workspace
def _load_corent_rows_or_workspace():
    """
    Return (corent_rows, using_workspace_source).
    Falls back to the latest WorkspaceRun when legacy CorentData is empty.
    """
    corent_rows = CorentData.query.order_by(CorentData.app_id).all()
    if corent_rows:
        return corent_rows, False
    try:
        from app.models.correlation_workspace import WorkspaceCorentRow, WorkspaceRun
        latest_run = WorkspaceRun.query.filter_by(status='done').order_by(WorkspaceRun.id.desc()).first()
        if latest_run:
            ws_rows = WorkspaceCorentRow.query.filter_by(run_id=latest_run.id).all()
            if ws_rows:
                return ws_rows, True
    except Exception:
        pass
    return [], False


# Function: _build_data_indexes
def _build_data_indexes(corent_rows, _using_workspace_source):
    """
    Build cast_index, industry_index, corent_ws_index, cast_ws_index from DB.
    When using workspace source, workspace tables also serve as the cast index.
    """
    cast_index     = {r.app_id: r for r in CASTData.query.all()     if r.app_id}
    industry_index = {r.app_id: r for r in IndustryData.query.all() if r.app_id}

    try:
        from app.models.correlation_workspace import WorkspaceCorentRow, WorkspaceCastRow, WorkspaceRun
        # status='done' — must reference the same completed run as _load_corent_rows_or_workspace().
        latest_run = WorkspaceRun.query.filter_by(status='done').order_by(WorkspaceRun.id.desc()).first()
        if latest_run:
            corent_ws_index = {r.app_id: r for r in WorkspaceCorentRow.query.filter_by(run_id=latest_run.id).all() if r.app_id}
            cast_ws_index   = {r.app_id: r for r in WorkspaceCastRow.query.filter_by(run_id=latest_run.id).all()   if r.app_id}
        else:
            corent_ws_index, cast_ws_index = {}, {}
    except Exception:
        corent_ws_index, cast_ws_index = {}, {}

    if _using_workspace_source and not cast_index:
        cast_index = cast_ws_index
    if _using_workspace_source and not corent_ws_index:
        corent_ws_index = {r.app_id: r for r in corent_rows if r.app_id}

    return cast_index, industry_index, corent_ws_index, cast_ws_index


# Function: _prepare_output_workbook
def _prepare_output_workbook(missing_template_msg: str) -> None:
    """Validate the template exists, then copy it to the output path."""
    if not _TEMPLATE_PATH.exists():
        raise FileNotFoundError(missing_template_msg)
    _OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    shutil.copy2(str(_TEMPLATE_PATH), str(_OUTPUT_PATH))


# Function: _clear_data_rows
def _clear_data_rows(ws) -> None:
    """Blank out all data cells (cols 1-64) from _DATA_START_ROW to the sheet's last row."""
    if ws.max_row >= _DATA_START_ROW:
        for r in range(_DATA_START_ROW, ws.max_row + 1):
            for c in range(1, 65):
                ws.cell(row=r, column=c).value = None


# Function: _write_row_to_sheet
def _write_row_to_sheet(ws, excel_row, values, colors) -> None:
    for col_i, (val, color) in enumerate(zip(values, colors), start=1):
        cell = ws.cell(row=excel_row, column=col_i)
        cell.value = val if (val is not None and str(val).strip() != "") else None
        if color == "yellow":
            cell.fill = _YELLOW_FILL
        elif color == "amber":
            cell.fill = _AMBER_FILL
        else:
            cell.fill = _NO_FILL
        cell.alignment = _CENTER


# Function: _write_golden_rows
def _write_golden_rows(
    ws, corent_rows, cast_index, industry_index,
    corent_ai_fills, cast_ai_fills, corent_ws_index, cast_ws_index,
) -> Tuple[List[str], List[List], List[List]]:
    """Write each corent row to the sheet, upsert GoldenData, and collect preview data."""
    missing_cast:   List[str]  = []
    preview_rows:   List[List] = []
    preview_colors: List[List] = []

    for row_idx, corent_row in enumerate(corent_rows):
        app_id       = corent_row.app_id or str(corent_row.id)
        cast_row     = cast_index.get(app_id)
        industry_row = industry_index.get(app_id)

        if cast_row is None:
            missing_cast.append(app_id)

        c_ai     = corent_ai_fills.get(app_id, set())
        k_ai     = cast_ai_fills.get(app_id, set())
        c_ws_row = corent_ws_index.get(app_id)
        k_ws_row = cast_ws_index.get(app_id)

        values, colors = _build_row(corent_row, cast_row, industry_row, None, c_ai, k_ai,
                                    corent_ws=c_ws_row, cast_ws=k_ws_row)

        excel_row = _DATA_START_ROW + row_idx
        _write_row_to_sheet(ws, excel_row, values, colors)

        if row_idx < _MAX_PREVIEW:
            preview_rows.append(values)
            preview_colors.append(colors)

        _upsert_golden_record(app_id, None, values, colors)

    return missing_cast, preview_rows, preview_colors


# Function: generate_golden_data
def generate_golden_data() -> Dict[str, Any]:
    """
    1. Fetch CorentData, CASTData, IndustryData from DB.
    2. Track AI-filled columns from latest workspace run.
    3. Populate APRAttributes.xlsx (UpdatedData/) with correct colour coding:
         Yellow = AI-predicted cell
         Amber  = empty SURVEY-source cell
         No fill = real DB value
    4. Upsert each row into GoldenData DB table.
    5. Return JSON preview with cell_colors list.
    """
    # ── 1. Fetch data (falls back to workspace tables when legacy CorentData is empty) ──
    corent_rows, _using_workspace_source = _load_corent_rows_or_workspace()

    if not corent_rows:
        return {
            "row_count": 0,
            "preview_headers": _get_column_headers(),
            "preview_rows": [],
            "cell_colors": [],
            "missing_cast": [],
            "error": "No CORENT data in the database. Upload a CORENT Excel file first.",
        }

    GoldenData.query.delete()
    db.session.flush()

    cast_index, industry_index, corent_ws_index, cast_ws_index = _build_data_indexes(
        corent_rows, _using_workspace_source
    )
    corent_ai_fills, cast_ai_fills = _get_workspace_ai_fills()

    # ── 2. Copy template ──────────────────────────────────────────────────────
    _prepare_output_workbook(f"APRAttributes template not found: {_TEMPLATE_PATH}")

    # ── 3. Open workbook ──────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(str(_OUTPUT_PATH))
    ws = wb[_SHEET_NAME]
    _clear_data_rows(ws)

    # ── 4. Write rows ──────────────────────────────────────────────────────────
    missing_cast, preview_rows, preview_colors = _write_golden_rows(
        ws, corent_rows, cast_index, industry_index,
        corent_ai_fills, cast_ai_fills, corent_ws_index, cast_ws_index,
    )

    db.session.commit()
    wb.save(str(_OUTPUT_PATH))

    return {
        "row_count":       len(corent_rows),
        "preview_headers": _get_column_headers(),
        "preview_rows":    preview_rows,
        "cell_colors":     preview_colors,
        "missing_cast":    missing_cast,
        "output_path":     str(_OUTPUT_PATH),
        "error":           None,
    }


# Function: _upsert_golden_record
def _upsert_golden_record(
    app_id: str,
    existing: Optional[GoldenData],
    values: List[Any],
    colors: List[Optional[str]],
) -> None:
    record = GoldenData(app_id=app_id)
    db.session.add(record)

    # Function: v
    def v(i):
        x = values[i]
        return str(x) if x is not None else None

    # Function: vi
    def vi(i):
        x = values[i]
        try:
            return int(x) if x is not None else None
        except (ValueError, TypeError):
            return None

    record.app_name                               = v(1)
    record.server_type                            = v(2)
    record.operating_system                       = v(3)
    record.cpu_core                               = v(4)
    record.memory                                 = v(5)
    record.internal_storage                       = v(6)
    record.external_storage                       = v(7)
    record.storage_type                           = v(8)
    record.db_storage                             = v(9)
    record.db_engine                              = v(10)
    record.environment_install_type               = v(11)
    record.virtualization_attributes              = v(12)
    record.compute_server_hardware_architecture   = v(13)
    record.application_stability                  = v(14)
    record.virtualization_state                   = v(15)
    record.storage_decomposition                  = v(16)
    record.flash_storage_used                     = v(17)
    record.cpu_requirement                        = v(18)
    record.memory_ram_requirement                 = v(19)
    record.mainframe_dependency                   = v(20)
    record.desktop_dependency                     = v(21)
    record.app_os_platform_cloud_suitability      = v(22)
    record.database_cloud_readiness               = v(23)
    record.integration_middleware_cloud_readiness = v(24)
    record.application_architecture               = v(25)
    record.application_hardware_dependency        = v(26)
    record.app_cots_vs_non_cots                   = v(27)
    record.source_code_availability               = v(28)
    record.programming_language                   = v(29)
    record.component_coupling                     = v(30)
    record.cloud_suitability                      = v(31)
    record.volume_external_dependencies           = v(32)
    record.app_service_api_readiness              = v(33)
    record.app_load_predictability_elasticity     = v(34)
    record.degree_of_code_protocols               = v(35)
    record.code_design                            = v(36)
    record.application_code_complexity_volume     = v(37)
    record.financially_optimizable_hardware_usage = v(38)
    record.distributed_architecture_design        = v(39)
    record.latency_requirements                   = v(40)
    record.ubiquitous_access_requirements         = v(41)
    # Survey cols (42-63): preserve existing values, overwrite only if user supplied
    for ci, db_field in _SURVEY_DB.items():
        existing_val = getattr(record, db_field, None)
        new_val = v(ci)
        if new_val is not None:
            setattr(record, db_field, new_val)
        elif existing_val is not None:
            pass  # preserve existing user-entered value
    record.no_of_production_environments          = vi(52)
    record.no_of_non_production_environments      = vi(53)
    record.ha_dr_requirements                     = v(54)
    record.rto_requirements                       = v(56)
    record.rpo_requirements                       = v(57)
    record.deployment_geography                   = v(58)
    record.ai_filled_cols = [i for i, c in enumerate(colors) if c == "yellow"] or None


# Function: _build_preview_colors_for_record
def _build_preview_colors_for_record(rec, row) -> List[Optional[str]]:
    colors = [None] * 64
    ai_set = set(rec.ai_filled_cols or [])
    for ci in ai_set:
        if 0 <= ci < 64:
            colors[ci] = "yellow"
    for idx, src, *_ in _COL_MAP:
        if src == "survey" and colors[idx] is None:
            val = row[idx]
            if val is None or str(val).strip() == "":
                colors[idx] = "amber"
    return colors


# Function: _preview_from_db_records
def _preview_from_db_records(records, headers) -> Dict[str, Any]:
    preview_rows   = []
    preview_colors = []
    missing_cast   = []
    cast_index = {r.app_id: r for r in CASTData.query.all() if r.app_id}

    for rec in records[:_MAX_PREVIEW]:
        if rec.app_id and rec.app_id not in cast_index:
            missing_cast.append(rec.app_id)
        row = _db_record_to_row(rec)
        preview_rows.append(row)
        preview_colors.append(_build_preview_colors_for_record(rec, row))

    return {
        "row_count":       len(records),
        "preview_headers": headers,
        "preview_rows":    preview_rows,
        "cell_colors":     preview_colors,
        "missing_cast":    missing_cast,
        "source":          "db",
    }


# Function: _load_latest_workspace_indexes
def _load_latest_workspace_indexes() -> Tuple[Dict[str, Any], Dict[str, Any]]:
    """Return (corent_ws_index, cast_ws_index) for the latest done WorkspaceRun, or ({}, {})."""
    try:
        from app.models.correlation_workspace import WorkspaceCorentRow, WorkspaceCastRow, WorkspaceRun
        # status='done' — see _get_workspace_ai_fills for why this filter is required.
        lrun = WorkspaceRun.query.filter_by(status='done').order_by(WorkspaceRun.id.desc()).first()
        c_ws_idx = {r.app_id: r for r in WorkspaceCorentRow.query.filter_by(run_id=lrun.id).all() if r.app_id} if lrun else {}
        k_ws_idx = {r.app_id: r for r in WorkspaceCastRow.query.filter_by(run_id=lrun.id).all()   if r.app_id} if lrun else {}
    except Exception:
        c_ws_idx, k_ws_idx = {}, {}
    return c_ws_idx, k_ws_idx


# Function: _preview_from_live_query
def _preview_from_live_query(headers) -> Dict[str, Any]:
    corent_rows  = CorentData.query.order_by(CorentData.app_id).all()
    cast_index   = {r.app_id: r for r in CASTData.query.all()     if r.app_id}
    industry_idx = {r.app_id: r for r in IndustryData.query.all() if r.app_id}
    missing_cast, preview_rows, preview_colors = [], [], []
    c_ws_idx, k_ws_idx = _load_latest_workspace_indexes()

    for cr in corent_rows[:_MAX_PREVIEW]:
        app_id = cr.app_id or str(cr.id)
        cast_r = cast_index.get(app_id)
        ind_r  = industry_idx.get(app_id)
        if not cast_r:
            missing_cast.append(app_id)
        vals, cols = _build_row(cr, cast_r, ind_r, None, set(), set(),
                                corent_ws=c_ws_idx.get(app_id),
                                cast_ws=k_ws_idx.get(app_id))
        preview_rows.append(vals)
        preview_colors.append(cols)

    return {
        "row_count":       len(corent_rows),
        "preview_headers": headers,
        "preview_rows":    preview_rows,
        "cell_colors":     preview_colors,
        "missing_cast":    missing_cast,
        "source":          "live",
    }


# Function: get_preview_data
def get_preview_data() -> Dict[str, Any]:
    """Return preview from GoldenData DB table (no Excel rebuild)."""
    records = GoldenData.query.order_by(GoldenData.app_id).all()
    headers = _get_column_headers()

    if records:
        return _preview_from_db_records(records, headers)

    return _preview_from_live_query(headers)


# Function: _write_db_record_row_to_sheet
def _write_db_record_row_to_sheet(ws, excel_r, values, ai_set) -> None:
    for col_i, val in enumerate(values, start=1):
        cell = ws.cell(row=excel_r, column=col_i)
        cell.value = val if (val is not None and str(val).strip() != "") else None
        ci  = col_i - 1
        src = _COL_MAP[ci][1] if ci < len(_COL_MAP) else "unknown"
        is_empty = (val is None or str(val).strip() == "")

        if ci in ai_set:
            cell.fill = _YELLOW_FILL
        elif is_empty and src == "survey":
            cell.fill = _AMBER_FILL
        else:
            cell.fill = _NO_FILL
        cell.alignment = _CENTER


# Function: regenerate_excel_from_db
def regenerate_excel_from_db() -> Dict[str, Any]:
    """Rebuild APRAttributes.xlsx from GoldenData DB table (post-edit rebuild)."""
    records = GoldenData.query.order_by(GoldenData.app_id).all()
    if not records:
        return {"error": "No GoldenData records. Run Generate first.", "row_count": 0}

    _prepare_output_workbook(f"Template not found: {_TEMPLATE_PATH}")

    wb = openpyxl.load_workbook(str(_OUTPUT_PATH))
    ws = wb[_SHEET_NAME]
    _clear_data_rows(ws)

    for idx, rec in enumerate(records):
        values  = _db_record_to_row(rec)
        ai_set  = set(rec.ai_filled_cols or [])
        excel_r = _DATA_START_ROW + idx
        _write_db_record_row_to_sheet(ws, excel_r, values, ai_set)

    wb.save(str(_OUTPUT_PATH))
    return {"row_count": len(records), "output_path": str(_OUTPUT_PATH), "error": None}


# Function: _db_record_to_row
def _db_record_to_row(rec: GoldenData) -> List[Any]:
    """Map a GoldenData DB record to the 64-column preview list."""
    row = [None] * 64
    row[0]  = rec.app_id
    row[1]  = rec.app_name
    row[2]  = rec.server_type
    row[3]  = rec.operating_system
    row[4]  = rec.cpu_core
    row[5]  = rec.memory
    row[6]  = rec.internal_storage
    row[7]  = rec.external_storage
    row[8]  = rec.storage_type
    row[9]  = rec.db_storage
    row[10] = rec.db_engine
    row[11] = rec.environment_install_type
    row[12] = rec.virtualization_attributes
    row[13] = rec.compute_server_hardware_architecture
    row[14] = rec.application_stability
    row[15] = rec.virtualization_state
    row[16] = rec.storage_decomposition
    row[17] = rec.flash_storage_used
    row[18] = rec.cpu_requirement
    row[19] = rec.memory_ram_requirement
    row[20] = rec.mainframe_dependency
    row[21] = rec.desktop_dependency
    row[22] = rec.app_os_platform_cloud_suitability
    row[23] = rec.database_cloud_readiness
    row[24] = rec.integration_middleware_cloud_readiness
    row[25] = rec.application_architecture
    row[26] = rec.application_hardware_dependency
    row[27] = rec.app_cots_vs_non_cots
    row[28] = rec.source_code_availability
    row[29] = rec.programming_language
    row[30] = rec.component_coupling
    row[31] = getattr(rec, "cloud_suitability",             None)
    row[32] = getattr(rec, "volume_external_dependencies",  None)
    row[33] = rec.app_service_api_readiness
    row[34] = rec.app_load_predictability_elasticity
    row[35] = rec.degree_of_code_protocols
    row[36] = rec.code_design
    row[37] = rec.application_code_complexity_volume
    row[38] = rec.financially_optimizable_hardware_usage
    row[39] = getattr(rec, "distributed_architecture_design", None)
    row[40] = rec.latency_requirements
    row[41] = rec.ubiquitous_access_requirements
    for ci, db_field in _SURVEY_DB.items():
        row[ci] = getattr(rec, db_field, None)
    row[52] = rec.no_of_production_environments
    row[53] = rec.no_of_non_production_environments
    row[54] = rec.ha_dr_requirements
    row[56] = rec.rto_requirements
    row[57] = rec.rpo_requirements
    row[58] = rec.deployment_geography
    return row

