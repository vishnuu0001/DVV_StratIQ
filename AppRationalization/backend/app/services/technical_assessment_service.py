# ---------------------------------------------------------------------------
# Author: Vishnuu A
# Scope: Validated, transactional imports for Technical Assessment workbooks.
# Date: 2025-08-07
# ---------------------------------------------------------------------------
"""Validated, transactional imports for Technical Assessment workbooks."""
import hashlib
import json
import math
import re
from datetime import date, datetime
from pathlib import Path
from openpyxl import load_workbook
from sqlalchemy import desc
from app import db
from app.models.technical_assessment import (
    BusinessValidation,
    TechnicalAssessmentImport,
    TechnicalEvaluationCategorizeMeta,
    TechnicalEvaluationCategorizeRow,
    WaveInput,
)
from app.services.ollama_service import OllamaService

BUSINESS_SHEET = "Business_Applications"
WAVE_SHEET = "Wave_Plan_Input"
TECHNICAL_EVALUATION_TOPIC = "Harmonize Maintenance Management Systems"
BUSINESS_HEADERS = [
    "Number", "Name", "Categorization", "Application family", "Business owner", "Department",
    "OLB Level 2", "IT Application owner", "GD Segments", "Department2", "OLB Level 23",
    "Architecture type", "Platform Host", "Application type", "Install type",
    "Capabilities", "Inconsistency", "Rationale",
]
WAVE_HEADERS = [
    "App ID", "Application Name", "Topic (Categorization)", "Business Capability (from source col P)",
    "Migration Type",  # first of two same-labelled columns in the real template — see _dedupe_headers
    "Capability Confirmed?", "Business Rationale", "Disposition", "Target Platform", "Migration Type",
    "T-Shirt Size", "Complexity", "# Tables", "Data Volume (records)", "Change Impact", "Risk",
    "Quick Win", "Business Criticality", "Department", "Business Owner", "Install Type", "Site / Region",
    "Dependencies (App IDs)", "Dependency Readiness", "Earliest Start", "Latest Finish", "Regulatory Hold?",
    "Assessment Effort (hrs)", "Migration Effort (hrs)", "Total Effort (hrs)", "Wave Eligibility Score",
    "Proposed Wave", "SME / Owner", "Workshop Date", "Sign-off Status", "Comments",
]


# Function: _clean
def _clean(value):
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    return value.strip() if isinstance(value, str) else value


# Function: _is_technical_evaluation_topic
def _is_technical_evaluation_topic(value):
    """Match the single topic approved for Technical Evaluation imports."""
    return str(_clean(value) or "").casefold() == TECHNICAL_EVALUATION_TOPIC.casefold()


# Function: _bool
def _bool(value):
    value = _clean(value)
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"yes", "true", "1", "x"}


# Function: _int
def _int(value):
    value = _clean(value)
    if value in (None, "") or (isinstance(value, str) and value.startswith("=")):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


# Function: _number
def _number(value):
    value = _clean(value)
    if value in (None, "") or (isinstance(value, str) and value.startswith("=")):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


# Function: _date
def _date(value):
    value = _clean(value)
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str) and value:
        try:
            return date.fromisoformat(value[:10])
        except ValueError:
            return None
    return None


# Function: _checksum
def _checksum(path):
    digest = hashlib.sha256()
    with open(path, "rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


# Function: _dedupe_headers
def _dedupe_headers(headers):
    """Disambiguate repeated header names (e.g. the real Wave_Plan_Input
    template has two columns both literally titled "Migration Type") into
    unique dict keys — "Name", "Name #2", "Name #3", ... — so a naive
    ``dict(zip(headers, values))`` can't silently drop a duplicate column's
    values by overwriting them with the next same-named column's values.
    The *equality* check against expected_headers still uses the raw,
    undeduplicated list, since that's what's literally in the file.
    """
    seen = {}
    keys = []
    for header in headers:
        seen[header] = seen.get(header, 0) + 1
        keys.append(header if seen[header] == 1 else f"{header} #{seen[header]}")
    return keys


# Function: _is_yellow_cell
def _is_yellow_cell(cell):
    fill = getattr(cell, "fill", None)
    if not fill or getattr(fill, "fill_type", None) in (None, "none"):
        return False
    color = getattr(fill, "fgColor", None) or getattr(fill, "start_color", None)
    if not color:
        return False
    rgb = (getattr(color, "rgb", None) or "").upper()
    if rgb and any(rgb.endswith(token) for token in ("FFFF00", "FFF200", "FFEB3B", "FFEA00")):
        return True
    index = (getattr(color, "index", None) or "").upper()
    return index in {"FFFFFF00", "FFFF00"}


# Function: _detect_header_row
def _detect_header_row(sheet):
    max_scan = min(sheet.max_row, 20)
    best_row = 1
    best_score = -1
    for row_idx in range(1, max_scan + 1):
        values = [
            _norm_header_key(_clean(sheet.cell(row=row_idx, column=col).value))
            for col in range(1, sheet.max_column + 1)
        ]
        has_topic = any(value in {"topic", "categorization", "category"} for value in values)
        has_product = any(
            value in {"product", "applicationname", "name"}
            or value.startswith("product")
            for value in values
        )
        if not (has_topic and has_product):
            continue

        has_size = any(value in {"size", "tshirtsize", "complexity"} for value in values)
        keyword_hits = sum(
            1
            for value in values
            if any(token in value for token in ("capabilit", "producttype", "cots", "compliance", "dynamic"))
        )

        yellow_here = any(
            _is_yellow_cell(sheet.cell(row=row_idx, column=col))
            for col in range(1, sheet.max_column + 1)
        )
        next_row_idx = min(row_idx + 1, sheet.max_row)
        next_values = [
            _clean(sheet.cell(row=next_row_idx, column=col).value)
            for col in range(1, sheet.max_column + 1)
        ]
        next_non_empty = sum(1 for value in next_values if value not in (None, ""))
        yellow_next = any(
            _is_yellow_cell(sheet.cell(row=next_row_idx, column=col))
            for col in range(1, sheet.max_column + 1)
        )

        score = 10
        if has_size:
            score += 3
        score += min(keyword_hits, 5)
        score += min(next_non_empty, 6)
        if yellow_here:
            score += 4
        if yellow_next:
            score += 4

        if score > best_score:
            best_score = score
            best_row = row_idx

    return best_row


# Function: _market_enrichment_default
def _market_enrichment_default(highlighted_headers):
    return {header: "Unknown" for header in highlighted_headers}


# Function: _sanitize_market_payload
def _sanitize_market_payload(payload, highlighted_headers):
    def _normalize_matrix_value(header, value):
        text = str(value or "").strip()
        if not text:
            return "Unknown"
        lower = text.casefold()

        is_product_type = any(token in header.casefold() for token in ("product type", "cots", "custom"))
        if is_product_type:
            if "hybrid" in lower:
                return "Hybrid"
            if "custom" in lower and "cots" in lower:
                return "Hybrid"
            if "cots" in lower or "available in market" in lower:
                return "COTS"
            if "custom" in lower:
                return "Custom"
            return "Unknown"

        if lower in {"yes", "y", "x", "true", "supported", "provides", "available"}:
            return "Yes"
        if lower in {"no", "n", "false", "not supported", "unavailable"}:
            return "No"
        if "partial" in lower or "limited" in lower:
            return "Partial"
        if "unknown" in lower or "uncertain" in lower or "n/a" in lower:
            return "Unknown"

        # Fallback for verbose model answers: infer coarse matrix value.
        if any(token in lower for token in ("does not", "not provide", "unsupported")):
            return "No"
        if any(token in lower for token in ("provide", "supports", "compliant", "compliance", "flood")):
            return "Yes"
        return "Unknown"

    result = _market_enrichment_default(highlighted_headers)
    if not isinstance(payload, dict):
        return result
    for header in highlighted_headers:
        value = payload.get(header)
        if value is None:
            continue
        text = _normalize_matrix_value(header, value)
        if text:
            result[header] = text[:50]
    return result


def _is_product_type_header(header):
    key = str(header or "").casefold()
    return any(token in key for token in ("product type", "cots", "custom product", "available in market"))


def _dynamic_matrix_headers(rows):
    """Return the stable, topic-specific schema stored by the enrichment run."""
    headers = []
    seen = set()
    for row in rows:
        payload = row.to_dict().get("enrichment_payload", {})
        if not isinstance(payload, dict) or payload.get("_matrix_schema_version") != 2:
            continue
        for header in payload:
            name = str(header or "").strip()
            key = name.casefold()
            if name and not name.startswith("_") and key not in seen:
                seen.add(key)
                headers.append(name)
    capability_headers = [header for header in headers if not _is_product_type_header(header)]
    product_type_headers = [header for header in headers if _is_product_type_header(header)]
    return capability_headers, product_type_headers


# Function: _norm_key
def _norm_key(value):
    if value is None:
        return ""
    text = str(value).strip().casefold()
    return " ".join(text.split())


# Function: _norm_header_key
def _norm_header_key(value):
    if value is None:
        return ""
    text = str(value).strip().casefold()
    text = re.sub(r"[^a-z0-9]+", "", text)
    return text


# Function: _pick_header_by_alias
def _pick_header_by_alias(headers, aliases, startswith=False):
    normalized = [(_norm_header_key(header), header) for header in headers]
    for key, header in normalized:
        if key in aliases:
            return header
    if startswith:
        for key, header in normalized:
            if any(key.startswith(alias) for alias in aliases):
                return header
    return None


# Function: _is_generic_capability_header
def _is_generic_capability_header(value):
    key = _norm_header_key(value)
    return key in {"capabilities", "capability", "producttype", "type", "column", ""} or key.startswith("capabilities")


# Function: _parse_categorize_sheet
def _parse_categorize_sheet(sheet):
    topic_aliases = {
        "topic",
        "categorization",
        "category",
        "topiccategorization",
    }
    product_aliases = {
        "product",
        "applicationname",
        "application",
        "businessapplication",
        "businessapplications",
        "appname",
        "name",
    }
    size_aliases = {
        "size",
        "tshirtsize",
        "complexity",
    }

    header_row_idx = _detect_header_row(sheet)
    secondary_row_idx = min(header_row_idx + 1, sheet.max_row)

    raw_headers = []
    sub_headers = []
    raw_highlight_cols = set()
    capability_span_cols = set()
    for col_idx in range(1, sheet.max_column + 1):
        header_cell = sheet.cell(row=header_row_idx, column=col_idx)
        sub_cell = sheet.cell(row=secondary_row_idx, column=col_idx)
        header = str(_clean(header_cell.value) or "").strip()
        sub_header = str(_clean(sub_cell.value) or "").strip()
        if not header:
            header = f"Column {col_idx}"
        raw_headers.append(header)
        sub_headers.append(sub_header)
        if _is_yellow_cell(header_cell) or _is_yellow_cell(sub_cell):
            raw_highlight_cols.add(col_idx - 1)
        if "capabilit" in _norm_header_key(header):
            capability_span_cols.add(col_idx - 1)

    # If a yellow capability parent header is merged across multiple columns,
    # include the full merged span as highlighted capability candidates.
    for merged in sheet.merged_cells.ranges:
        if merged.max_row < header_row_idx or merged.min_row > secondary_row_idx:
            continue
        top_left = sheet.cell(row=merged.min_row, column=merged.min_col)
        merged_header = str(_clean(top_left.value) or "").strip()
        merged_header_key = _norm_header_key(merged_header)
        is_capability_group = "capabilit" in merged_header_key
        is_yellow_group = _is_yellow_cell(top_left)
        if not (is_capability_group or is_yellow_group):
            continue
        for col_idx in range(merged.min_col, merged.max_col + 1):
            zero_idx = col_idx - 1
            if is_yellow_group:
                raw_highlight_cols.add(zero_idx)
            if is_capability_group:
                capability_span_cols.add(zero_idx)

    key_map = []
    for idx, header in enumerate(raw_headers):
        sub_header = sub_headers[idx]
        if sub_header and _is_generic_capability_header(header):
            key_map.append(sub_header)
        else:
            key_map.append(header)
    key_map = _dedupe_headers(key_map)

    topic_col = _pick_header_by_alias(
        key_map,
        topic_aliases,
        startswith=True,
    ) or (key_map[0] if key_map else "Topic")

    product_col = _pick_header_by_alias(
        key_map,
        product_aliases,
        startswith=True,
    )

    size_col = _pick_header_by_alias(
        key_map,
        size_aliases,
        startswith=True,
    )

    dynamic_cols = set(raw_highlight_cols)
    dynamic_cols.update(capability_span_cols)
    for idx, sub_header in enumerate(sub_headers):
        if not sub_header:
            continue
        header_key = _norm_header_key(raw_headers[idx])
        if header_key.startswith("column") and idx in capability_span_cols:
            dynamic_cols.add(idx)

    highlighted_headers = [
        key_map[idx]
        for idx in sorted(dynamic_cols)
        if idx < len(key_map) and key_map[idx] not in {topic_col, product_col, size_col}
    ]
    if not highlighted_headers:
        highlighted_headers = [
            header for header in key_map
            if header not in {topic_col, product_col, size_col}
            and any(
                token in header.casefold()
                for token in (
                    "compliance",
                    "dynamic",
                    "alarm",
                    "market",
                    "cots",
                    "capabilities",
                    "product type",
                    "custom",
                )
            )
        ]

    # Ensure Product Type is always visible/populated when present in the workbook.
    product_type_header = _pick_header_by_alias(
        key_map,
        {"producttype", "type", "cots"},
        startswith=True,
    )
    if product_type_header and product_type_header not in highlighted_headers:
        highlighted_headers.append(product_type_header)

    rows = []
    last_topic = ""
    data_start_row = header_row_idx + 1
    if any(sub_headers):
        data_start_row = min(header_row_idx + 2, sheet.max_row + 1)

    for row_idx in range(data_start_row, sheet.max_row + 1):
        values = [_clean(sheet.cell(row=row_idx, column=col).value) for col in range(1, sheet.max_column + 1)]
        if not any(value not in (None, "") for value in values):
            continue
        payload = {key: values[i] for i, key in enumerate(key_map)}
        raw_topic = str(payload.get(topic_col) or "").strip()
        topic = raw_topic or last_topic
        product = str(payload.get(product_col) or "").strip() if product_col else ""
        size = str(payload.get(size_col) or "").strip() if size_col else ""
        if raw_topic:
            last_topic = raw_topic
        if not topic or not product:
            continue
        rows.append((row_idx, topic, product, size or None, payload))

    return {
        "sheet": sheet.title,
        "header_row_idx": header_row_idx,
        "key_map": key_map,
        "highlighted_headers": highlighted_headers,
        "highlighted_count": len(highlighted_headers),
        "topic_col": topic_col,
        "product_col": product_col,
        "size_col": size_col,
        "rows": rows,
    }


# Function: _wave_size_lookup
def _wave_size_lookup():
    """Build product/application-name -> size lookup from latest Wave Inputs."""
    wave_import = latest_import("wave_inputs")
    if not wave_import:
        return {}

    lookup = {}
    rows = WaveInput.query.filter_by(import_id=wave_import.id).all()
    for row in rows:
        size_value = row.tshirt_size or row.complexity
        if not size_value:
            continue
        app_name_key = _norm_key(row.application_name)
        app_id_key = _norm_key(row.app_id)
        if app_name_key and app_name_key not in lookup:
            lookup[app_name_key] = str(size_value)
        if app_id_key and app_id_key not in lookup:
            lookup[app_id_key] = str(size_value)
    return lookup


def _calculated_tshirt_size(row):
    """Estimate size from portfolio attributes when no Wave Input exists.

    The score is intentionally deterministic and conservative: deployment,
    customization, architecture, hosting, and flagged inconsistency each add
    delivery complexity. Wave Inputs always take precedence over this estimate.
    """
    payload = row.to_dict().get("row_payload", {}) if hasattr(row, "to_dict") else {}
    normalized = {str(key).casefold(): str(value or "").strip().casefold() for key, value in payload.items()}
    application_type = normalized.get("application type", "")
    architecture = normalized.get("architecture type", "")
    install_type = normalized.get("install type", "")
    platform_host = normalized.get("platform host", "")
    inconsistency = normalized.get("inconsistency", "")

    score = 0
    if any(token in application_type for token in ("major modification", "modified")):
        score += 1
    if any(token in application_type for token in ("custom", "in-house", "in house", "bespoke")):
        score += 2
    if any(token in architecture for token in ("distributed", "client server", "mainframe", "multi-tier")):
        score += 1
    if "mainframe" in architecture:
        score += 1
    if "on premise" in install_type or "on-premise" in install_type:
        score += 1
    if platform_host and platform_host not in {"none", "n/a", "unknown"}:
        score += 1
    if inconsistency in {"true", "yes", "1"}:
        score += 1

    if score <= 0:
        return "XXS"
    if score == 1:
        return "XS"
    if score == 2:
        return "S"
    if score == 3:
        return "M"
    if score == 4:
        return "L"
    return "XL"


def _resolve_categorize_size(row, size_lookup):
    """Resolve size by Wave Input ID/name, workbook value, then calculation."""
    payload = row.to_dict().get("row_payload", {})
    validated_override = str(payload.get("_validated_size_override") or "").strip().upper()
    if validated_override:
        return validated_override, "validated"
    candidate_keys = [
        payload.get("App ID"),
        payload.get("Number"),
        payload.get("Application Number"),
        row.product,
    ]
    for candidate in candidate_keys:
        matched = size_lookup.get(_norm_key(candidate))
        if matched:
            return str(matched).strip().upper(), "wave_inputs"
    if row.size:
        return str(row.size).strip().upper(), "categorize_workbook"
    return _calculated_tshirt_size(row), "calculated"


def update_technical_evaluation_validation(row_id, updates, updated_by):
    """Persist editable Validate-menu decisions for one matrix row."""
    if not isinstance(updates, dict):
        raise ValueError("Validation payload must be an object")
    import_record = latest_import("technical_evaluation_categorize")
    if not import_record:
        raise ValueError("No categorized workbook import found")
    row = TechnicalEvaluationCategorizeRow.query.filter_by(
        id=row_id,
        import_id=import_record.id,
    ).first()
    if not row:
        raise ValueError("Technical Evaluation row was not found in the latest import")

    changed_fields = []
    if "size" in updates:
        size = str(updates.get("size") or "").strip().upper()
        allowed_sizes = {"", "XXS", "XS", "S", "M", "L", "XL", "XXL"}
        if size not in allowed_sizes:
            raise ValueError(f"Unsupported T-shirt size: {size}")
        row_payload = row.to_dict().get("row_payload", {})
        if size:
            row_payload["_validated_size_override"] = size
            row_payload["_validated_size_at"] = datetime.utcnow().isoformat()
            row_payload["_validated_size_by"] = str(updated_by or "system")
        else:
            row_payload.pop("_validated_size_override", None)
            row_payload.pop("_validated_size_at", None)
            row_payload.pop("_validated_size_by", None)
        row.row_payload_json = json.dumps(row_payload, ensure_ascii=False, default=str)
        changed_fields.append("Size")

    values = updates.get("values")
    if values is not None:
        if not isinstance(values, dict):
            raise ValueError("Capability values must be an object")
        all_rows = TechnicalEvaluationCategorizeRow.query.filter_by(import_id=import_record.id).all()
        capability_headers, product_type_headers = _dynamic_matrix_headers(all_rows)
        allowed_headers = set(capability_headers + product_type_headers)
        enrichment = row.to_dict().get("enrichment_payload", {})
        for header, raw_value in values.items():
            if header not in allowed_headers:
                raise ValueError(f"Unknown matrix column: {header}")
            value = str(raw_value or "").strip()
            allowed = (
                {"COTS", "Custom", "Hybrid", "Unknown"}
                if _is_product_type_header(header)
                else {"Yes", "No", "Partial", "Unknown"}
            )
            if value not in allowed:
                raise ValueError(f"Unsupported value '{value}' for {header}")
            enrichment[header] = value
            changed_fields.append(header)
        enrichment["_matrix_schema_version"] = 2
        enrichment["_validated_at"] = datetime.utcnow().isoformat()
        enrichment["_validated_by"] = str(updated_by or "system")
        enrichment["_validated_fields"] = sorted(set(
            list(enrichment.get("_validated_fields") or []) + changed_fields
        ))
        row.enrichment_payload_json = json.dumps(enrichment, ensure_ascii=False)

    if not changed_fields:
        raise ValueError("No editable validation fields were supplied")
    db.session.commit()
    dashboard = get_technical_evaluation_categorize_dashboard(topic=row.topic)
    item = next((item for item in dashboard["items"] if item["id"] == row.id), None)
    return {
        "item": item,
        "updated_fields": sorted(set(changed_fields)),
        "updated_by": str(updated_by or "system"),
        "updated_at": datetime.utcnow().isoformat(),
    }


# Function: _read
def _read(path, sheet, expected_headers):
    # data_only=True: several columns (effort hours, wave eligibility score)
    # are Excel formulas in the real template — read the cached calculated
    # value, not the formula text, or every numeric converter downstream
    # would discard them as unparseable and silently null the column.
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet not in workbook.sheetnames:
            raise ValueError(f"Required worksheet '{sheet}' was not found")
        worksheet = workbook[sheet]
        headers = [_clean(cell.value) for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
        if headers != expected_headers:
            missing = [header for header in expected_headers if header not in headers]
            raise ValueError(f"Unexpected {sheet} schema. Missing columns: {missing}")
        dict_keys = _dedupe_headers(headers)
        rows = []
        for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 2):
            if any(_clean(value) is not None for value in values):
                rows.append((row_number, dict(zip(dict_keys, values))))
        return rows
    finally:
        workbook.close()


# Function: _replace_dataset
def _replace_dataset(dataset_type, checksum, child_model):
    """Reject an exact re-upload of a file already loaded for this dataset;
    otherwise clear every previous import of this dataset type so only the
    newly uploaded file's data exists afterward — imports never accumulate
    across re-uploads, keeping the table's contents unique to the current file.
    """
    existing_imports = TechnicalAssessmentImport.query.filter_by(dataset_type=dataset_type).all()
    if any(imp.checksum_sha256 == checksum for imp in existing_imports):
        raise ValueError("This exact file has already been imported — no changes detected.")
    for imp in existing_imports:
        child_model.query.filter_by(import_id=imp.id).delete(synchronize_session=False)
        if dataset_type == "technical_evaluation_categorize":
            TechnicalEvaluationCategorizeMeta.query.filter_by(import_id=imp.id).delete(synchronize_session=False)
        db.session.delete(imp)
    if existing_imports:
        db.session.flush()


# Function: clear_dataset
def clear_dataset(dataset_type):
    """Delete all imported data (and import log rows) for one dataset type."""
    child_model = {
        "business_validations": BusinessValidation,
        "wave_inputs": WaveInput,
        "technical_evaluation_categorize": TechnicalEvaluationCategorizeRow,
    }.get(dataset_type)
    if child_model is None:
        raise ValueError(f"Unsupported dataset type: {dataset_type}")
    imports = TechnicalAssessmentImport.query.filter_by(dataset_type=dataset_type).all()
    cleared_rows = 0
    for imp in imports:
        cleared_rows += child_model.query.filter_by(import_id=imp.id).delete(synchronize_session=False)
        if dataset_type == "technical_evaluation_categorize":
            TechnicalEvaluationCategorizeMeta.query.filter_by(import_id=imp.id).delete(synchronize_session=False)
        db.session.delete(imp)
    if dataset_type == "wave_inputs":
        # Derived schedules can't outlive their source Wave Inputs.
        from app.models.wave_schedule import WaveSchedule, WaveScheduleWave, WaveScheduleTask, WaveScheduleApp
        WaveScheduleApp.query.delete(synchronize_session=False)
        WaveScheduleTask.query.delete(synchronize_session=False)
        WaveScheduleWave.query.delete(synchronize_session=False)
        WaveSchedule.query.delete(synchronize_session=False)
    db.session.commit()
    return cleared_rows


# Function: import_business_validations
def import_business_validations(path, source_filename, imported_by):
    rows = _read(path, BUSINESS_SHEET, BUSINESS_HEADERS)
    errors, seen = [], set()
    for row_number, row in rows:
        app_id = str(_clean(row["Number"]) or "")
        if not app_id:
            errors.append({"row": row_number, "message": "Number is required"})
        elif app_id in seen:
            errors.append({"row": row_number, "message": f"Duplicate application Number: {app_id}"})
        seen.add(app_id)
    if errors:
        raise ValueError(f"Workbook validation failed: {errors[:10]}")
    checksum = _checksum(path)
    _replace_dataset("business_validations", checksum, BusinessValidation)
    record = TechnicalAssessmentImport(
        dataset_type="business_validations", source_filename=source_filename,
        source_sheet=BUSINESS_SHEET, checksum_sha256=checksum,
        row_count=len(rows), imported_by=imported_by,
    )
    db.session.add(record)
    db.session.flush()
    for row_number, row in rows:
        db.session.add(BusinessValidation(
            import_id=record.id, row_number=row_number, application_number=str(_clean(row["Number"])),
            name=str(_clean(row["Name"]) or ""), categorization=_clean(row["Categorization"]),
            application_family=_clean(row["Application family"]), business_owner=_clean(row["Business owner"]),
            department=_clean(row["Department"]), olb_level_2=_clean(row["OLB Level 2"]),
            it_application_owner=_clean(row["IT Application owner"]), gd_segments=_clean(row["GD Segments"]),
            department_2=_clean(row["Department2"]), olb_level_23=_clean(row["OLB Level 23"]),
            architecture_type=_clean(row["Architecture type"]), platform_host=_clean(row["Platform Host"]),
            application_type=_clean(row["Application type"]), install_type=_clean(row["Install type"]),
            capabilities=_clean(row["Capabilities"]), inconsistency=bool(_bool(row["Inconsistency"])),
            rationale=_clean(row["Rationale"]),
        ))
    db.session.commit()
    return record


# Function: import_wave_inputs
def import_wave_inputs(path, source_filename, imported_by):
    rows = _read(path, WAVE_SHEET, WAVE_HEADERS)
    # The workbook's documented example is never operational data.
    rows = [(number, row) for number, row in rows if str(_clean(row["App ID"]) or "") != "APM0000000"]
    rows = [
        (number, row)
        for number, row in rows
        if _is_technical_evaluation_topic(row["Topic (Categorization)"])
    ]
    if not rows:
        raise ValueError(
            "Workbook validation failed: no rows found for required topic "
            f"'{TECHNICAL_EVALUATION_TOPIC}'"
        )
    errors, seen = [], set()
    for row_number, row in rows:
        app_id = str(_clean(row["App ID"]) or "")
        if not app_id:
            errors.append({"row": row_number, "message": "App ID is required"})
        elif app_id in seen:
            errors.append({"row": row_number, "message": f"Duplicate App ID: {app_id}"})
        seen.add(app_id)
    if errors:
        raise ValueError(f"Workbook validation failed: {errors[:10]}")
    checksum = _checksum(path)
    _replace_dataset("wave_inputs", checksum, WaveInput)
    record = TechnicalAssessmentImport(
        dataset_type="wave_inputs", source_filename=source_filename, source_sheet=WAVE_SHEET,
        checksum_sha256=checksum, row_count=len(rows), imported_by=imported_by,
    )
    db.session.add(record)
    db.session.flush()
    fields = [
        ("topic", "Topic (Categorization)", _clean), ("business_capability", "Business Capability (from source col P)", _clean),
        ("capability_confirmed", "Capability Confirmed?", _bool), ("business_rationale", "Business Rationale", _clean),
        ("disposition", "Disposition", _clean), ("target_platform", "Target Platform", _clean),
        # Two columns are both literally titled "Migration Type" in the real template (col E, col J).
        # _dedupe_headers renames the first occurrence's dict key unchanged and the second to "... #2".
        ("migration_approach", "Migration Type", _clean), ("migration_type", "Migration Type #2", _clean),
        ("tshirt_size", "T-Shirt Size", _clean),
        ("complexity", "Complexity", _clean), ("table_count", "# Tables", _int),
        ("data_volume_records", "Data Volume (records)", _int), ("change_impact", "Change Impact", _clean),
        ("risk", "Risk", _clean), ("quick_win", "Quick Win", _bool),
        ("business_criticality", "Business Criticality", _clean), ("department", "Department", _clean),
        ("business_owner", "Business Owner", _clean), ("install_type", "Install Type", _clean),
        ("site_region", "Site / Region", _clean), ("dependencies", "Dependencies (App IDs)", _clean),
        ("dependency_readiness", "Dependency Readiness", _clean), ("earliest_start", "Earliest Start", _date),
        ("latest_finish", "Latest Finish", _date), ("regulatory_hold", "Regulatory Hold?", _bool),
        ("assessment_effort_hours", "Assessment Effort (hrs)", _number),
        ("migration_effort_hours", "Migration Effort (hrs)", _number), ("total_effort_hours", "Total Effort (hrs)", _number),
        ("wave_eligibility_score", "Wave Eligibility Score", _number), ("proposed_wave", "Proposed Wave", _clean),
        ("sme_owner", "SME / Owner", _clean), ("workshop_date", "Workshop Date", _date),
        ("signoff_status", "Sign-off Status", _clean), ("comments", "Comments", _clean),
    ]
    for row_number, row in rows:
        values = {field: converter(row[header]) for field, header, converter in fields}
        db.session.add(WaveInput(import_id=record.id, row_number=row_number,
                                 app_id=str(_clean(row["App ID"])), application_name=str(_clean(row["Application Name"]) or ""),
                                 **values))
    db.session.commit()

    # Wave Planning is calculated on demand (the "Predict Wave Planning"
    # button), not eagerly here — every calculation now mandatorily calls
    # Ollama, so eagerly recomputing all ~20 topics on every import would
    # block the upload response for many minutes. The next "Predict Wave
    # Planning" click always recalculates fresh against this import anyway.

    return record


# Function: import_technical_evaluation_categorize
def import_technical_evaluation_categorize(path, source_filename, imported_by):
    workbook = load_workbook(path, read_only=False, data_only=True)
    try:
        candidates = []
        missing_product = []
        for sheet in workbook.worksheets:
            parsed = _parse_categorize_sheet(sheet)
            if parsed["product_col"]:
                candidates.append(parsed)
            else:
                missing_product.append(parsed)

        if not candidates:
            diagnostics = []
            for parsed in missing_product[:4]:
                sample_headers = ", ".join(parsed["key_map"][:8])
                diagnostics.append(f"{parsed['sheet']}: {sample_headers}")
            detail = "; ".join(diagnostics) if diagnostics else "no readable worksheets"
            raise ValueError(
                "Workbook validation failed: Product column is required "
                f"(checked sheets: {detail})"
            )

        best = max(
            candidates,
            key=lambda item: (
                item.get("highlighted_count", 0),
                len(item["rows"]),
            ),
        )
        if not best["rows"]:
            raise ValueError(
                "Workbook validation failed: no Topic/Product rows found "
                f"(detected sheet: {best['sheet']})"
            )

        key_map = best["key_map"]
        highlighted_headers = best["highlighted_headers"]
        topic_col = best["topic_col"]
        product_col = best["product_col"]
        size_col = best["size_col"]
        rows = best["rows"]
        rows = [row for row in rows if _is_technical_evaluation_topic(row[1])]
        if not rows:
            raise ValueError(
                "Workbook validation failed: no rows found for required topic "
                f"'{TECHNICAL_EVALUATION_TOPIC}'"
            )
        # Persist one canonical spelling so dashboard filters, automatic
        # enrichment and validation edits cannot diverge on case/whitespace.
        rows = [
            (row_number, TECHNICAL_EVALUATION_TOPIC, product, size, payload)
            for row_number, _topic, product, size, payload in rows
        ]

        checksum = _checksum(path)
        _replace_dataset("technical_evaluation_categorize", checksum, TechnicalEvaluationCategorizeRow)

        record = TechnicalAssessmentImport(
            dataset_type="technical_evaluation_categorize",
            source_filename=source_filename,
            source_sheet=best["sheet"],
            checksum_sha256=checksum,
            row_count=len(rows),
            imported_by=imported_by,
        )
        db.session.add(record)
        db.session.flush()

        db.session.add(TechnicalEvaluationCategorizeMeta(
            import_id=record.id,
            headers_json=json.dumps(key_map, ensure_ascii=False),
            highlighted_headers_json=json.dumps(highlighted_headers, ensure_ascii=False),
            topic_column=topic_col,
            product_column=product_col,
            size_column=size_col,
        ))

        for row_number, topic, product, size, payload in rows:
            db.session.add(TechnicalEvaluationCategorizeRow(
                import_id=record.id,
                row_number=row_number,
                topic=topic,
                product=product,
                size=size,
                row_payload_json=json.dumps(payload, ensure_ascii=False, default=str),
                enrichment_payload_json="{}",
            ))

        db.session.commit()
        return record
    finally:
        workbook.close()


# Function: get_technical_evaluation_categorize_dashboard
def get_technical_evaluation_categorize_dashboard(topic=None, search=""):
    selected_topic = (
        str(topic or "").strip()
        if _is_technical_evaluation_topic(topic)
        else TECHNICAL_EVALUATION_TOPIC
    )
    import_record = latest_import("technical_evaluation_categorize")
    if not import_record:
        return {
            "items": [],
            "topics": [],
            "total": 0,
            "import": None,
            "highlighted_headers": [],
            "capability_headers": [],
            "product_type_headers": [],
            "headers": [],
            "selected_topic": selected_topic,
            "market_search": OllamaService.market_search_status(),
        }

    meta = TechnicalEvaluationCategorizeMeta.query.filter_by(import_id=import_record.id).first()
    meta_data = meta.to_dict() if meta else {"headers": [], "highlighted_headers": []}

    query = TechnicalEvaluationCategorizeRow.query.filter_by(
        import_id=import_record.id,
        topic=selected_topic,
    )
    if search:
        pattern = f"%{search.strip()}%"
        query = query.filter(db.or_(
            TechnicalEvaluationCategorizeRow.topic.ilike(pattern),
            TechnicalEvaluationCategorizeRow.product.ilike(pattern),
        ))
    rows = query.order_by(TechnicalEvaluationCategorizeRow.row_number).all()
    size_lookup = _wave_size_lookup()
    capability_headers, product_type_headers = _dynamic_matrix_headers(rows)
    dynamic_headers = capability_headers + product_type_headers

    topics = [
        value[0]
        for value in TechnicalEvaluationCategorizeRow.query
        .with_entities(TechnicalEvaluationCategorizeRow.topic)
        .filter_by(import_id=import_record.id, topic=TECHNICAL_EVALUATION_TOPIC)
        .distinct()
        .order_by(TechnicalEvaluationCategorizeRow.topic)
        .all()
    ]
    items = []
    for row in rows:
        resolved_size, size_source = _resolve_categorize_size(row, size_lookup)
        items.append({
            **row.to_dict(),
            "size": resolved_size,
            "size_source": size_source,
        })

    return {
        "items": items,
        "topics": topics,
        "total": len(rows),
        "import": import_record.to_dict(),
        "headers": meta_data.get("headers", []),
        "highlighted_headers": dynamic_headers,
        "capability_headers": capability_headers,
        "product_type_headers": product_type_headers,
        "selected_topic": selected_topic,
        "meta": meta_data,
        "market_search": OllamaService.market_search_status(),
    }


# Function: enrich_technical_evaluation_categorize_topic
def enrich_technical_evaluation_categorize_topic(topic):
    selected_topic = (topic or "").strip()
    if not selected_topic:
        raise ValueError("Topic is required")
    if not _is_technical_evaluation_topic(selected_topic):
        raise ValueError(
            "Technical Evaluation enrichment is restricted to the approved topic "
            f"'{TECHNICAL_EVALUATION_TOPIC}'"
        )

    import_record = latest_import("technical_evaluation_categorize")
    if not import_record:
        raise ValueError("No categorized workbook import found")

    meta = TechnicalEvaluationCategorizeMeta.query.filter_by(import_id=import_record.id).first()
    if not meta:
        raise ValueError("Categorize metadata is unavailable for the latest import")
    rows = TechnicalEvaluationCategorizeRow.query.filter_by(
        import_id=import_record.id,
        topic=selected_topic,
    ).order_by(TechnicalEvaluationCategorizeRow.row_number).all()
    if not rows:
        raise ValueError(f"No products found for topic '{selected_topic}'")
    previous_capabilities, previous_product_types = _dynamic_matrix_headers(rows)
    previous_headers = previous_capabilities + previous_product_types

    size_lookup = _wave_size_lookup()
    products = []
    for row in rows:
        data = row.to_dict()
        resolved_size, size_source = _resolve_categorize_size(row, size_lookup)
        products.append({
            "id": row.id,
            "product": row.product,
            "size": resolved_size,
            "size_source": size_source,
            "context": data.get("row_payload", {}),
        })

    matrix = OllamaService.discover_market_capability_matrix(
        selected_topic,
        products,
    )
    highlighted_headers = matrix["headers"]
    enrichment = matrix["values"]

    now = datetime.utcnow()
    for row in rows:
        payload = _sanitize_market_payload(enrichment.get(str(row.id)) or {}, highlighted_headers)
        payload["_matrix_schema_version"] = 2
        row.enrichment_payload_json = json.dumps(payload, ensure_ascii=False)
        row.market_checked_at = now
    db.session.commit()

    dashboard = get_technical_evaluation_categorize_dashboard(topic=selected_topic)
    current_headers = matrix["headers"]
    dashboard["enrichment_run"] = {
        "topic": selected_topic,
        "rows_updated": len(rows),
        "evidence_count": matrix.get("evidence_count", 0),
        "capabilities": matrix["capabilities"],
        "added_headers": [header for header in current_headers if header not in previous_headers],
        "removed_headers": [header for header in previous_headers if header not in current_headers],
        "completed_at": now.isoformat(),
    }
    return dashboard


# Function: latest_import
def latest_import(dataset_type):
    return TechnicalAssessmentImport.query.filter_by(dataset_type=dataset_type).order_by(
        desc(TechnicalAssessmentImport.imported_at), desc(TechnicalAssessmentImport.id)).first()
