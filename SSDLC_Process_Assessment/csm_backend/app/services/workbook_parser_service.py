# ---------------------------------------------------------------------------
# Author: Vishnuu A
# Scope: SSDLC_Process_Assessment — csm_backend/app/services (workbook_parser_service.py)
# Date: 2025-08-10
# ---------------------------------------------------------------------------
from __future__ import annotations

import io
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.models.inputs import InputAssumptions
from app.models.vendor_spend import VendorSpendRecord
from app.models.tower_model import TowerParam, DEFAULT_TOWER_PARAMS


INPUTS_LABEL_MAP: Dict[str, tuple[str, str]] = {
    "Total Technology Spend":                          ("total_technology_spend",                          "B"),
    "Direct Tech OpEx":                                ("direct_tech_opex",                                "B"),
    "Tech CapEx":                                      ("tech_capex",                                      "B"),
    "Internal Labor %":                                ("internal_labor_pct",                              "B"),
    "External Talent / Labor Spend":                   ("external_talent_labor_spend",                     "B"),
    "Total Third Party Spend":                         ("total_third_party_spend",                         "B"),
    "Vendor Management Overhead % of External Spend":  ("vendor_management_overhead_pct",                  "B"),
    "Target Vendor Management Overhead Reduction %":   ("target_vendor_management_overhead_reduction_pct", "B"),
    "Default Rate Compression %":                      ("default_rate_compression_pct",                    "B"),
    "Default Productivity Improvement %":              ("default_productivity_improvement_pct",            "B"),
    "Default Transition Duration Months":              ("default_transition_duration_months",               "B"),
    "Scenario":                                        ("scenario_name",                                   "E"),
    "Conservative Rate Compression":                   ("conservative_rate_compression_pct",               "E"),
    "Base Rate Compression":                           ("base_rate_compression_pct",                       "E"),
    "Aggressive Rate Compression":                     ("aggressive_rate_compression_pct",                 "E"),
    "Conservative Productivity":                       ("conservative_productivity_pct",                   "E"),
    "Base Productivity":                               ("base_productivity_pct",                           "E"),
    "Aggressive Productivity":                         ("aggressive_productivity_pct",                     "E"),
    "One-Time Transition Cost %":                      ("one_time_transition_cost_pct",                    "E"),
}

COLUMN_MAP: Dict[str, str] = {
    "vendor":              "vendor",
    "spend category":      "spend_category",
    "tower":               "tower",
    "service / scope":     "service_scope",
    "role type":           "role_type",
    "criticality":         "criticality",
    "pricing basis":       "pricing_basis",
    "fte count":           "fte_count",
    "avg rate / hr":       "avg_rate_per_hr",
    "annual fixed spend":  "annual_fixed_spend",
    "annual spend":        "annual_spend",
    "source / notes":      "source_notes",
}

# Percentage fields that should be stored as decimals (0-1 range)
PCT_FIELDS = {
    "internal_labor_pct",
    "vendor_management_overhead_pct",
    "target_vendor_management_overhead_reduction_pct",
    "default_rate_compression_pct",
    "default_productivity_improvement_pct",
    "conservative_rate_compression_pct",
    "base_rate_compression_pct",
    "aggressive_rate_compression_pct",
    "conservative_productivity_pct",
    "base_productivity_pct",
    "aggressive_productivity_pct",
    "one_time_transition_cost_pct",
}


@dataclass
class ParsedWorkbook:
    inputs: InputAssumptions
    vendor_records: List[VendorSpendRecord]
    tower_params: List[TowerParam]
    raw_sheets: Dict[str, List[List[Any]]]
    validation: List[str] = field(default_factory=list)
    sheet_names: List[str] = field(default_factory=list)


# Function: default_inputs
def default_inputs() -> InputAssumptions:
    """Return InputAssumptions populated with workbook default values."""
    return InputAssumptions(
        total_technology_spend=Decimal("142750000"),
        direct_tech_opex=Decimal("114100000"),
        tech_capex=Decimal("31500000"),
        internal_labor_pct=Decimal("0.25"),
        external_talent_labor_spend=Decimal("40800000"),
        total_third_party_spend=Decimal("75400000"),
        vendor_management_overhead_pct=Decimal("0.03"),
        target_vendor_management_overhead_reduction_pct=Decimal("0.40"),
        default_rate_compression_pct=Decimal("0.08"),
        default_productivity_improvement_pct=Decimal("0.10"),
        default_transition_duration_months=Decimal("6"),
        scenario_name="Base",
        conservative_rate_compression_pct=Decimal("0.05"),
        base_rate_compression_pct=Decimal("0.08"),
        aggressive_rate_compression_pct=Decimal("0.12"),
        conservative_productivity_pct=Decimal("0.05"),
        base_productivity_pct=Decimal("0.10"),
        aggressive_productivity_pct=Decimal("0.15"),
        one_time_transition_cost_pct=Decimal("0.05"),
    )


# Function: _safe_decimal
def _safe_decimal(value: Any, field_name: str = "") -> Optional[Decimal]:
    """Convert a cell value to Decimal, returning None on failure."""
    if value is None:
        return None
    try:
        d = Decimal(str(value))
        # If field is a pct field and value looks like a percentage (> 1), normalise it
        if field_name in PCT_FIELDS and d > Decimal("1"):
            d = d / Decimal("100")
        return d
    except (InvalidOperation, ValueError):
        return None


# Function: _safe_str
def _safe_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


# Function: _safe_int
def _safe_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None


# Function: _find_header_row
def _find_header_row(ws: Worksheet, key: str) -> Optional[int]:
    """Find the 1-based row index of the first row containing `key` in any cell."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and key.lower() in str(cell.value).lower():
                return cell.row
    return None


# Function: _col_letter_to_idx
def _col_letter_to_idx(letter: str) -> int:
    """Convert column letter (A, B, ...) to 0-based index."""
    letter = letter.upper()
    result = 0
    for ch in letter:
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result - 1


# Function: _build_label_to_row
def _build_label_to_row(ws: Worksheet) -> Dict[str, int]:
    """Build a lookup of stripped label text → row number from column A."""
    label_to_row: Dict[str, int] = {}
    for row in ws.iter_rows():
        col_a = row[0] if row else None
        if col_a and col_a.value:
            label_to_row[str(col_a.value).strip()] = col_a.row
    return label_to_row


# Function: _fuzzy_match_row
def _fuzzy_match_row(label: str, label_to_row: Dict[str, int]) -> Optional[int]:
    """Find a row number whose label loosely matches `label` (substring either way)."""
    for k, v in label_to_row.items():
        if label.lower() in k.lower() or k.lower() in label.lower():
            return v
    return None


# Function: _extract_input_field
def _extract_input_field(field_name: str, label: str, raw_val: Any, validation: List[str]) -> Optional[Any]:
    """Convert a raw cell value into the typed value for a single input field."""
    if field_name == "scenario_name":
        return _safe_str(raw_val) or "Base"
    if field_name == "default_transition_duration_months":
        d = _safe_decimal(raw_val, field_name)
        return d if d is not None else Decimal("6")

    d = _safe_decimal(raw_val, field_name)
    if d is None:
        validation.append(f"Inputs: field '{field_name}' (label='{label}') has no numeric value.")
        return None
    return d


# Function: _parse_inputs_sheet
def _parse_inputs_sheet(ws: Worksheet, validation: List[str]) -> InputAssumptions:
    """Parse the Inputs sheet and return InputAssumptions."""
    parsed: Dict[str, Any] = {}
    label_to_row = _build_label_to_row(ws)
    col_idx_by_letter = {"B": _col_letter_to_idx("B"), "E": _col_letter_to_idx("E")}

    for label, (field_name, col_letter) in INPUTS_LABEL_MAP.items():
        row_num = label_to_row.get(label)
        if row_num is None:
            row_num = _fuzzy_match_row(label, label_to_row)
        if row_num is None:
            validation.append(f"Inputs: label '{label}' not found in sheet.")
            continue

        # openpyxl rows/cells are 1-indexed
        cell = ws.cell(row=row_num, column=col_idx_by_letter[col_letter] + 1)
        value = _extract_input_field(field_name, label, cell.value, validation)
        if value is not None:
            parsed[field_name] = value

    defaults = default_inputs()
    defaults_dict = defaults.model_dump()
    defaults_dict.update(parsed)

    try:
        return InputAssumptions(**defaults_dict)
    except Exception as exc:
        validation.append(f"Inputs: model construction failed: {exc}. Using defaults.")
        return defaults


# Function: _build_vendor_col_index
def _build_vendor_col_index(ws: Worksheet, header_row_num: int) -> Dict[str, int]:
    """Map known column names to their index using COLUMN_MAP, from a header row."""
    header_row = list(ws.iter_rows(min_row=header_row_num, max_row=header_row_num, values_only=True))[0]
    col_index: Dict[str, int] = {}
    for idx, cell_val in enumerate(header_row):
        if cell_val is not None:
            mapped = COLUMN_MAP.get(str(cell_val).strip().lower())
            if mapped:
                col_index[mapped] = idx
    return col_index


# Function: _blank_vendor_record_data
def _blank_vendor_record_data(vendor_val: Any, annual_spend: Decimal) -> Dict[str, Any]:
    return {
        "vendor":        _safe_str(vendor_val),
        "annual_spend":  annual_spend,
        "spend_category": "",
        "tower":          "",
        "service_scope":  "",
        "role_type":      "",
        "criticality":    "",
        "pricing_basis":  "",
        "fte_count":      None,
        "avg_rate_per_hr":    None,
        "annual_fixed_spend": None,
        "source_notes":   "",
    }


# Function: _fill_vendor_record_fields
def _fill_vendor_record_fields(record_data: Dict[str, Any], row_vals: tuple, col_index: Dict[str, int]) -> None:
    for field_name, idx in col_index.items():
        if field_name in ("vendor", "annual_spend"):
            continue
        raw = row_vals[idx] if idx < len(row_vals) else None
        if field_name == "fte_count":
            record_data[field_name] = _safe_int(raw)
        elif field_name in ("avg_rate_per_hr", "annual_fixed_spend"):
            record_data[field_name] = _safe_decimal(raw)
        else:
            record_data[field_name] = _safe_str(raw)


# Function: _parse_vendor_row
def _parse_vendor_row(row_vals: tuple, col_index: Dict[str, int], validation: List[str]) -> Optional[VendorSpendRecord]:
    """Build a VendorSpendRecord from one data row, or None if the row should be skipped."""
    vendor_val = row_vals[col_index["vendor"]] if "vendor" in col_index else None
    spend_val = row_vals[col_index["annual_spend"]] if "annual_spend" in col_index else None

    if not vendor_val or str(vendor_val).strip() == "":
        return None
    annual_spend = _safe_decimal(spend_val)
    if annual_spend is None or annual_spend == Decimal("0"):
        return None

    record_data = _blank_vendor_record_data(vendor_val, annual_spend)
    _fill_vendor_record_fields(record_data, row_vals, col_index)

    try:
        return VendorSpendRecord(**record_data)
    except Exception as exc:
        validation.append(f"Current_Vendor_Spend: Skipped row for vendor '{vendor_val}': {exc}")
        return None


# Function: _parse_vendor_spend_sheet
def _parse_vendor_spend_sheet(ws: Worksheet, validation: List[str]) -> List[VendorSpendRecord]:
    """Parse the Current_Vendor_Spend sheet."""
    records: List[VendorSpendRecord] = []

    header_row_num = _find_header_row(ws, "Vendor") or _find_header_row(ws, "Annual Spend")
    if header_row_num is None:
        validation.append("Current_Vendor_Spend: No header row found.")
        return records

    col_index = _build_vendor_col_index(ws, header_row_num)
    if "annual_spend" not in col_index:
        validation.append("Current_Vendor_Spend: 'Annual Spend' column not found.")
        return records
    if "vendor" not in col_index:
        validation.append("Current_Vendor_Spend: 'Vendor' column not found.")
        return records

    for row_vals in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
        record = _parse_vendor_row(row_vals, col_index, validation)
        if record is not None:
            records.append(record)

    return records


# Function: _classify_tower_header_cell
def _classify_tower_header_cell(k: str) -> Optional[str]:
    """Map a lower-cased header cell's text to its tower-model field name."""
    if "tower" in k:
        return "tower"
    if "scope" in k or "consolidation" in k:
        return "consolidation_scope_pct"
    if "action" in k or "recommendation" in k:
        return "recommended_action"
    if "note" in k:
        return "notes"
    return None


# Function: _build_tower_col_index
def _build_tower_col_index(ws: Worksheet, header_row_num: int) -> Dict[str, int]:
    header_row = list(ws.iter_rows(min_row=header_row_num, max_row=header_row_num, values_only=True))[0]
    col_index: Dict[str, int] = {}
    for idx, val in enumerate(header_row):
        if val:
            field_name = _classify_tower_header_cell(str(val).strip().lower())
            if field_name:
                col_index[field_name] = idx
    return col_index


# Function: _parse_tower_row
def _parse_tower_row(row_vals: tuple, col_index: Dict[str, int], default_map: Dict[str, TowerParam]) -> Optional[TowerParam]:
    """Build a TowerParam from one data row, or None if the row should be skipped."""
    tower_val = row_vals[col_index["tower"]] if "tower" in col_index else None
    if not tower_val or str(tower_val).strip() == "":
        return None

    tower_name = _safe_str(tower_val)
    fallback = default_map.get(tower_name, DEFAULT_TOWER_PARAMS[0])

    scope_val = row_vals[col_index["consolidation_scope_pct"]] if "consolidation_scope_pct" in col_index else None
    scope = _safe_decimal(scope_val, "consolidation_scope_pct")
    if scope is None:
        scope = fallback.consolidation_scope_pct

    action_val = row_vals[col_index["recommended_action"]] if "recommended_action" in col_index else None
    action = _safe_str(action_val) or fallback.recommended_action

    notes_val = row_vals[col_index["notes"]] if "notes" in col_index else None
    notes = _safe_str(notes_val)

    return TowerParam(tower=tower_name, consolidation_scope_pct=scope, recommended_action=action, notes=notes)


# Function: _parse_tower_model_sheet
def _parse_tower_model_sheet(ws: Worksheet, validation: List[str]) -> List[TowerParam]:
    """Parse Option1_Tower_Model sheet for tower consolidation scope params."""
    header_row_num = _find_header_row(ws, "Tower")
    if header_row_num is None:
        validation.append("Option1_Tower_Model: No header row found. Using defaults.")
        return list(DEFAULT_TOWER_PARAMS)

    col_index = _build_tower_col_index(ws, header_row_num)
    if "tower" not in col_index:
        validation.append("Option1_Tower_Model: 'Tower' column not found. Using defaults.")
        return list(DEFAULT_TOWER_PARAMS)

    default_map = {p.tower: p for p in DEFAULT_TOWER_PARAMS}
    params: List[TowerParam] = []
    for row_vals in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
        param = _parse_tower_row(row_vals, col_index, default_map)
        if param is not None:
            params.append(param)

    return params if params else list(DEFAULT_TOWER_PARAMS)


# Function: _raw_sheet_data
def _raw_sheet_data(ws: Worksheet) -> List[List[Any]]:
    """Capture raw sheet data as list of lists."""
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


class WorkbookParserService:
    """Parse an uploaded .xlsx workbook into a ParsedWorkbook dataclass."""

    # Function: _capture_raw_sheets
    @staticmethod
    def _capture_raw_sheets(wb, sheet_names: List[str], validation: List[str]) -> Dict[str, List[List[Any]]]:
        raw_sheets: Dict[str, List[List[Any]]] = {}
        for name in sheet_names:
            try:
                raw_sheets[name] = _raw_sheet_data(wb[name])
            except Exception as exc:
                validation.append(f"Sheet '{name}': failed to read raw data: {exc}")
        return raw_sheets

    # Function: _parse_inputs_or_default
    def _parse_inputs_or_default(self, wb, sheet_names: List[str], validation: List[str]) -> InputAssumptions:
        inputs_sheet_name = self._find_sheet(sheet_names, ["Inputs", "Input", "Assumptions"])
        if not inputs_sheet_name:
            validation.append("No 'Inputs' sheet found. Using default assumptions.")
            return default_inputs()
        try:
            return _parse_inputs_sheet(wb[inputs_sheet_name], validation)
        except Exception as exc:
            validation.append(f"Inputs sheet parse error: {exc}. Using defaults.")
            return default_inputs()

    # Function: _parse_vendors_or_empty
    def _parse_vendors_or_empty(self, wb, sheet_names: List[str], validation: List[str]) -> List[VendorSpendRecord]:
        vendor_sheet_name = self._find_sheet(sheet_names, [
            "Current_Vendor_Spend", "Vendor_Spend", "VendorSpend",
            "Current Vendor Spend", "Vendor Spend", "Vendors",
        ])
        if not vendor_sheet_name:
            validation.append("No 'Current_Vendor_Spend' sheet found. Vendor records will be empty.")
            return []
        try:
            return _parse_vendor_spend_sheet(wb[vendor_sheet_name], validation)
        except Exception as exc:
            validation.append(f"Vendor spend sheet parse error: {exc}.")
            return []

    # Function: _parse_towers_or_default
    def _parse_towers_or_default(self, wb, sheet_names: List[str], validation: List[str]) -> List[TowerParam]:
        tower_sheet_name = self._find_sheet(sheet_names, [
            "Option1_Tower_Model", "Tower_Model", "Tower Model",
            "Option1 Tower Model", "TowerModel",
        ])
        if not tower_sheet_name:
            return list(DEFAULT_TOWER_PARAMS)
        try:
            return _parse_tower_model_sheet(wb[tower_sheet_name], validation)
        except Exception as exc:
            validation.append(f"Tower model sheet parse error: {exc}. Using defaults.")
            return list(DEFAULT_TOWER_PARAMS)

    # Function: parse
    def parse(self, file_bytes: bytes) -> ParsedWorkbook:
        validation: List[str] = []

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        except Exception as exc:
            validation.append(f"Failed to open workbook: {exc}")
            return ParsedWorkbook(
                inputs=default_inputs(),
                vendor_records=[],
                tower_params=list(DEFAULT_TOWER_PARAMS),
                raw_sheets={},
                validation=validation,
                sheet_names=[],
            )

        sheet_names = wb.sheetnames
        raw_sheets = self._capture_raw_sheets(wb, sheet_names, validation)
        inputs = self._parse_inputs_or_default(wb, sheet_names, validation)
        vendor_records = self._parse_vendors_or_empty(wb, sheet_names, validation)

        if not vendor_records:
            validation.append("No vendor records parsed. Tower calculations will show zero spend.")

        tower_params = self._parse_towers_or_default(wb, sheet_names, validation)

        return ParsedWorkbook(
            inputs=inputs,
            vendor_records=vendor_records,
            tower_params=tower_params,
            raw_sheets=raw_sheets,
            validation=validation,
            sheet_names=sheet_names,
        )

    # Function: _find_sheet
    @staticmethod
    def _find_sheet(sheet_names: List[str], candidates: List[str]) -> Optional[str]:
        """Find the first matching sheet name (case-insensitive)."""
        lower_sheets = {s.lower(): s for s in sheet_names}
        for candidate in candidates:
            match = lower_sheets.get(candidate.lower())
            if match:
                return match
        # Partial match fallback
        for candidate in candidates:
            for lower, actual in lower_sheets.items():
                if candidate.lower() in lower or lower in candidate.lower():
                    return actual
        return None
