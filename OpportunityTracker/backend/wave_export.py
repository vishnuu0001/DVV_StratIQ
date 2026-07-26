# ---------------------------------------------------------------------------
# Author: Vishnuu A
# Scope: Regenerates the 6-sheet Wave Plan workbook (matching the structure of the real
# Date: 2026-04-16
# ---------------------------------------------------------------------------
"""
Regenerates the 6-sheet Wave Plan workbook (matching the structure of the real
BASF_LMP_Fasihi_Wave_Plan.xlsx worked example) fresh from the DB on every export
— never cached, so it can never drift from what the dashboard shows.
"""
from __future__ import annotations

import io
import re

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from models_wave import WaveApp, WaveCategorySummary

_TITLE_FONT = Font(bold=True, size=13)
_NOTE_FONT = Font(italic=True, size=9)
_HEADER_FONT = Font(bold=True)


# Function: _wave_sort_key
def _wave_sort_key(w: str) -> int:
    m = re.search(r"(\d+)", w or "")
    return int(m.group(1)) if m else 999


# Function: _autosize
def _autosize(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# Function: build_workbook
def build_workbook(db: Session, job_id: int, category: str) -> io.BytesIO:
    summary = db.query(WaveCategorySummary).filter_by(job_id=job_id, category=category).first()
    apps = db.query(WaveApp).filter_by(job_id=job_id, category=category).order_by(WaveApp.app_id).all()

    wb = Workbook()
    _build_wave_assignment_detail(wb, category, apps)
    _build_classification_logic(wb, category, summary, apps)
    _build_timeline_gantt(wb, apps)
    _build_remediation_detail(wb, apps)
    _build_wave_summary(wb, apps)
    _build_leadership_summary(wb, category, summary, apps)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# Function: _build_wave_assignment_detail
def _build_wave_assignment_detail(wb: Workbook, category: str, apps: list[WaveApp]) -> None:
    ws = wb.active
    ws.title = "Wave Assignment (2)"
    ws.append([f"StratIQ Wave Plan - {category} - Application Wave Assignment"])
    ws["A1"].font = _TITLE_FONT
    ws.append(["Migration Type classified via capability-count heuristic (LLM-assisted); Wave assigned deterministically by department clustering. See 'Wave Assignment' sheet for full logic."])
    ws["A2"].font = _NOTE_FONT
    ws.append([])
    headers = [
        "App ID", "Application name", "Dept (OLB L2)", "Business owner", "IT app owner",
        "Capabilities", "Capability tier", "Data quality flag", "Remediation required",
        "Disposition", "Service module", "Modernization trigger", "Migration type",
        "Change management", "Decommissioning", "Wave",
        "Stream 2 assessment", "Stream 3 execution", "AMS transition start",
        "Disposition rationale", "Notes / assumptions", "Wave Assignment Rationale",
    ]
    ws.append(headers)
    for cell in ws[4]:
        cell.font = _HEADER_FONT

    for a in sorted(apps, key=lambda a: (_wave_sort_key(a.wave), a.app_id)):
        ws.append([
            a.app_id, a.app_name, a.dept, a.business_owner, a.it_owner,
            a.capabilities_raw, a.capability_tier, "Yes" if a.data_quality_flag else "No",
            "Yes" if a.remediation_required else "No",
            a.disposition, a.service_module or "", a.modernization_trigger_category or "", a.migration_type,
            a.change_management, a.decommissioning, a.wave,
            a.stream2_assessment, a.stream3_execution, a.ams_transition_start,
            a.disposition_rationale, a.notes_assumptions, a.wave_assignment_rationale,
        ])
    _autosize(ws, [14, 30, 12, 18, 18, 40, 12, 12, 16, 16, 16, 20, 22, 30, 30, 14, 16, 16, 16, 50, 50, 50])


# Function: _build_classification_logic
def _build_classification_logic(wb: Workbook, category: str, summary: WaveCategorySummary, apps: list[WaveApp]) -> None:
    ws = wb.create_sheet("Wave Assignment")
    ws.append(["Classification Logic and Open Items"])
    ws["A1"].font = _TITLE_FONT
    ws.append([summary.classification_logic_notes or "No narrative available."])
    ws.append([])
    ws.append(["Disposition logic"])
    ws["A4"].font = _HEADER_FONT
    ws.append([
        "Harmonization is the default treatment. Modernization is applied only as an exception, where "
        "harmonization onto a strategic target platform is infeasible and at least one approved trigger "
        "category applies (IT Security Risk, Legal/Compliance Requirement, Technology End-of-Life, Structural Necessity)."
    ])
    ws.append([])
    total = len(apps)
    harmonization = sum(1 for a in apps if a.disposition == "Harmonization")
    modernization = sum(1 for a in apps if a.disposition == "Modernization")
    ws.append([f"Harmonization: {harmonization} of {total} apps ({round(100*harmonization/total) if total else 0}%)"])
    ws.append([f"Modernization: {modernization} of {total} apps ({round(100*modernization/total) if total else 0}%)"])
    ws.append([])
    ws.append(["Migration Type classification logic (Harmonization track)"])
    ws[f"A{ws.max_row}"].font = _HEADER_FONT
    ws.append([
        "Functional Migration Only: single, simple communication-style capability, minimal data transformation. "
        "Full Consolidation: 3+ capabilities, or a specialized/structured-data capability. "
        "Data Migration Only: pure archival candidate. TBD (data gap): capabilities field empty/unparseable."
    ])
    ws.append([])
    _append_migration_type_counts(ws, apps, total)
    _append_service_module_breakdown(ws, apps, modernization)
    _autosize(ws, [110])


# Function: _append_migration_type_counts
def _append_migration_type_counts(ws, apps: list[WaveApp], total: int) -> None:
    func_only = sum(1 for a in apps if a.migration_type == "Functional Migration Only")
    full_con = sum(1 for a in apps if a.migration_type == "Full Consolidation")
    data_only = sum(1 for a in apps if a.migration_type == "Data Migration Only")
    tbd = sum(1 for a in apps if a.migration_type == "TBD (data gap)")
    ws.append([f"Functional Migration Only: {func_only} of {total} apps ({round(100*func_only/total) if total else 0}%)"])
    ws.append([f"Full Consolidation: {full_con} of {total} apps ({round(100*full_con/total) if total else 0}%)"])
    if data_only:
        ws.append([f"Data Migration Only: {data_only} of {total} apps ({round(100*data_only/total) if total else 0}%)"])
    if tbd:
        ws.append([f"TBD (data gap): {tbd} of {total} apps ({round(100*tbd/total) if total else 0}%)"])


# Function: _append_service_module_breakdown
def _append_service_module_breakdown(ws, apps: list[WaveApp], modernization: int) -> None:
    if not modernization:
        return
    ws.append([])
    ws.append(["Service Module classification logic (Modernization track)"])
    ws[f"A{ws.max_row}"].font = _HEADER_FONT
    for module in ("Rehost", "Replatform", "Refactor", "Replace"):
        count = sum(1 for a in apps if a.service_module == module)
        if count:
            ws.append([f"{module}: {count} of {modernization} Modernization apps"])


# Function: _build_timeline_gantt
def _build_timeline_gantt(wb: Workbook, apps: list[WaveApp]) -> None:
    ws = wb.create_sheet("Timeline Gantt")
    ws.append(["Wave Timeline"])
    ws["A1"].font = _TITLE_FONT
    ws.append([])

    quarters = _quarter_sequence(apps)
    ws.append(["Wave / Phase", "App count"] + quarters)
    for cell in ws[3]:
        cell.font = _HEADER_FONT

    by_wave: dict[str, list[WaveApp]] = {}
    for a in apps:
        by_wave.setdefault(a.wave, []).append(a)

    phase_labels = ["Assess.", "Migr.", "Test/CO", "AMS", "Steady"]
    _append_remediation_row(ws, apps, quarters)

    for wave in sorted(by_wave.keys(), key=_wave_sort_key):
        ws.append(_wave_timeline_row(wave, by_wave[wave], quarters, phase_labels))
    _autosize(ws, [40, 12] + [10] * len(quarters))


# Function: _append_remediation_row
def _append_remediation_row(ws, apps: list[WaveApp], quarters: list[str]) -> None:
    flagged = [a for a in apps if a.data_quality_flag]
    if not flagged:
        return
    row = ["Data remediation (%d flagged apps)" % len(flagged), len(flagged)] + [""] * len(quarters)
    if quarters:
        row[2] = "Remed."
    ws.append(row)


# Function: _top_depts_label
def _top_depts_label(wave_apps: list[WaveApp], limit: int) -> str:
    depts: dict[str, int] = {}
    for a in wave_apps:
        depts[a.dept] = depts.get(a.dept, 0) + 1
    return ", ".join(f"{d}" for d, _ in sorted(depts.items(), key=lambda kv: -kv[1])[:limit])


# Function: _wave_timeline_row
def _wave_timeline_row(wave: str, wave_apps: list[WaveApp], quarters: list[str], phase_labels: list[str]) -> list:
    start_q = wave_apps[0].stream2_assessment or (quarters[0] if quarters else "")
    primary = _top_depts_label(wave_apps, 2)
    label = f"{wave} - {primary}, {len(wave_apps)} apps"
    row = [label, len(wave_apps)] + [""] * len(quarters)
    if start_q not in quarters:
        return row
    start_idx = quarters.index(start_q)
    for j, phase in enumerate(phase_labels):
        if start_idx + j < len(quarters):
            row[2 + start_idx + j] = phase
    return row


# Function: _collect_seen_quarters
def _collect_seen_quarters(apps: list[WaveApp]) -> list[str]:
    seen: list[str] = []
    for a in apps:
        for q in (a.stream2_assessment, a.stream3_execution, a.ams_transition_start):
            if q and q not in seen:
                seen.append(q)
    return seen


# Function: _quarter_sort_key
def _quarter_sort_key(q: str):
    m = re.match(r"Q(\d) (\d+)", q or "")
    return (int(m.group(2)), int(m.group(1))) if m else (9999, 9)


# Function: _extend_trailing_quarters
def _extend_trailing_quarters(ordered: list[str]) -> None:
    # Extend a couple quarters past the last one so "Steady" phase has somewhere to land.
    if not ordered:
        return
    m = re.match(r"Q(\d) (\d+)", ordered[-1])
    if not m:
        return
    q, y = int(m.group(1)), int(m.group(2))
    for _ in range(2):
        q += 1
        if q > 4:
            q = 1
            y += 1
        ordered.append(f"Q{q} {y}")


# Function: _quarter_sequence
def _quarter_sequence(apps: list[WaveApp]) -> list[str]:
    seen = _collect_seen_quarters(apps)
    ordered = sorted(seen, key=_quarter_sort_key)
    _extend_trailing_quarters(ordered)
    return ordered


# Function: _build_remediation_detail
def _build_remediation_detail(wb: Workbook, apps: list[WaveApp]) -> None:
    from routers.wave_analysis import REMEDIATION_PROCESS_TEXT

    ws = wb.create_sheet("Data Remediation Detail")
    flagged = [a for a in apps if a.data_quality_flag]
    ws.append([f"Data Remediation ({len(flagged)} flagged apps) - App Detail"])
    ws["A1"].font = _TITLE_FONT
    ws.append([REMEDIATION_PROCESS_TEXT])
    ws.append([])
    ws.append(["App ID", "Application name", "Dept", "Business owner", "Migration type", "Rejoins wave", "Rejoin rationale"])
    for cell in ws[4]:
        cell.font = _HEADER_FONT
    for a in flagged:
        ws.append([a.app_id, a.app_name, a.dept, a.business_owner, a.migration_type, a.rejoins_wave, a.rejoin_rationale])
    _autosize(ws, [14, 30, 10, 18, 22, 14, 50])


# Function: _build_wave_summary
def _build_wave_summary(wb: Workbook, apps: list[WaveApp]) -> None:
    ws = wb.create_sheet("Wave Summary")
    ws.append(["Wave Summary - Effort Mix by Migration Type"])
    ws["A1"].font = _TITLE_FONT
    ws.append(["Counts computed live from Wave Assignment detail; recomputed fresh on every export."])
    ws.append([])
    ws.append([
        "Wave", "App count", "Harmonization", "Modernization", "Data quality flags",
        "Functional Migr. Only", "Full Consolidation",
        "Data Migration Only", "TBD (data gap)", "Stream 2 assessment", "Stream 3 execution",
        "AMS transition start", "Primary departments",
    ])
    for cell in ws[4]:
        cell.font = _HEADER_FONT

    by_wave: dict[str, list[WaveApp]] = {}
    for a in apps:
        by_wave.setdefault(a.wave, []).append(a)

    for wave in sorted(by_wave.keys(), key=_wave_sort_key):
        wa = by_wave[wave]
        depts = {}
        for a in wa:
            depts[a.dept] = depts.get(a.dept, 0) + 1
        primary = ", ".join(f"{d} ({c})" for d, c in sorted(depts.items(), key=lambda kv: -kv[1])[:3])
        ws.append([
            wave, len(wa),
            sum(1 for a in wa if a.disposition == "Harmonization"),
            sum(1 for a in wa if a.disposition == "Modernization"),
            sum(1 for a in wa if a.data_quality_flag),
            sum(1 for a in wa if a.migration_type == "Functional Migration Only"),
            sum(1 for a in wa if a.migration_type == "Full Consolidation"),
            sum(1 for a in wa if a.migration_type == "Data Migration Only"),
            sum(1 for a in wa if a.migration_type == "TBD (data gap)"),
            wa[0].stream2_assessment, wa[0].stream3_execution, wa[0].ams_transition_start,
            primary,
        ])
    _autosize(ws, [16, 10, 14, 14, 16, 20, 18, 18, 14, 16, 16, 16, 40])


# Function: _build_leadership_summary
def _build_leadership_summary(wb: Workbook, category: str, summary: WaveCategorySummary, apps: list[WaveApp]) -> None:
    ws = wb.create_sheet("Leadership Summary")
    ws.append([f"StratIQ Wave Plan - {category} - Leadership Summary"])
    ws["A1"].font = _TITLE_FONT
    ws.append(["Executive Dashboard | Data recomputed live from Wave Assignment detail on every export."])
    ws.append([])
    ws.append(["PORTFOLIO OVERVIEW"])
    ws["A4"].font = _HEADER_FONT
    ws.append([
        "Total Apps", "Migration Waves", "Harmonization", "Modernization", "Data Quality Issues",
        "Func. Migration Only", "Full Consolidation", "Data Migration Only", "TBD (Data Gap)",
        "Timeline Start", "Timeline End",
    ])
    for cell in ws[5]:
        cell.font = _HEADER_FONT

    by_wave: dict[str, list[WaveApp]] = {}
    for a in apps:
        by_wave.setdefault(a.wave, []).append(a)

    ws.append([
        len(apps), len(by_wave),
        sum(1 for a in apps if a.disposition == "Harmonization"),
        sum(1 for a in apps if a.disposition == "Modernization"),
        sum(1 for a in apps if a.data_quality_flag),
        sum(1 for a in apps if a.migration_type == "Functional Migration Only"),
        sum(1 for a in apps if a.migration_type == "Full Consolidation"),
        sum(1 for a in apps if a.migration_type == "Data Migration Only"),
        sum(1 for a in apps if a.migration_type == "TBD (data gap)"),
        summary.timeline_start, summary.timeline_end,
    ])
    ws.append([])
    ws.append(["WAVE-LEVEL EXECUTIVE SUMMARY"])
    ws[f"A{ws.max_row}"].font = _HEADER_FONT
    ws.append(["Wave", "App count", "Functional Migr. Only", "Full Consolidation", "Data quality flags"])
    for cell in ws[ws.max_row]:
        cell.font = _HEADER_FONT
    for wave in sorted(by_wave.keys(), key=_wave_sort_key):
        wa = by_wave[wave]
        ws.append([
            wave, len(wa),
            sum(1 for a in wa if a.migration_type == "Functional Migration Only"),
            sum(1 for a in wa if a.migration_type == "Full Consolidation"),
            sum(1 for a in wa if a.data_quality_flag),
        ])
    _autosize(ws, [18, 14, 20, 18, 16, 14, 14])
