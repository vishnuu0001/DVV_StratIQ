# ---------------------------------------------------------------------------
# Author: TraceForge Team
# Scope: Enhanced test case Excel export with comprehensive sheets and metadata
# Date: 2026-02-20
# ---------------------------------------------------------------------------
"""Enhanced test case Excel export matching reference quality standards.

Generates comprehensive test case workbooks with:
- Test Cases sheet: Full metadata, requirement mapping, test data, steps
- Requirements Traceability Matrix: Coverage and mapping
- Test Data sheet: Example data sets for testing
- Defect Log: Template for defect tracking during execution
- Lists: Value lists and enumerations
"""
from __future__ import annotations

import io
import json
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Font, PatternFill, Border, Side, Protection,
)
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from traceforge.agents.coverage_policy import minimum_scenarios_for_requirement
from traceforge.db.models import Chunk, Requirement, SourceCitation, TestCase, TestPlan


_HEADER_FILL = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# Function: _format_cell
def _format_cell(cell, value: Any, header: bool = False, wrap: bool = True) -> None:
    """Format a cell with consistent styling."""
    cell.value = value
    if header:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    else:
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=wrap)
    cell.border = _BORDER


# Function: _set_column_width
def _set_column_width(sheet, col_num: int, width: float) -> None:
    """Set column width."""
    sheet.column_dimensions[get_column_letter(col_num)].width = width


# Function: _parse_tc_metadata
def _parse_tc_metadata(test_case) -> dict:
    """Decode rich metadata stored in the gherkin column."""
    raw = getattr(test_case, "gherkin", None) or ""
    if raw and raw.strip().startswith("{"):
        try:
            import json
            return json.loads(raw)
        except (ValueError, Exception):
            pass
    return {}


# Function: _create_test_cases_sheet
async def _create_test_cases_sheet(
    workbook: Workbook,
    session: AsyncSession,
    project_id: str,
) -> None:
    """Create detailed Test Cases sheet with enterprise-quality columns."""
    sheet = workbook.create_sheet("Test Cases")

    headers = [
        "Test Case ID",
        "Requirement ID",
        "Process Area",
        "Test Case Title",
        "Objective",
        "Test Type",
        "Test Level",
        "Priority",
        "Risk Rating",
        "Automation Status",
        "Systems Involved",
        "Required Roles",
        "Preconditions",
        "Test Data",
        "Test Steps",
        "Expected Result",
        "Cleanup / Reversal",
        "Ambiguities",
        "Assumptions",
        "Automation Blockers",
        "Status",
        "Execution Cycle",
        "Actual Result",
        "Defect ID",
    ]

    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        _format_cell(cell, header, header=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    result = await session.execute(
        select(TestCase, Requirement)
        .join(Requirement, Requirement.id == TestCase.requirement_id)
        .where(TestCase.project_id == project_id)
        .order_by(TestCase.tc_id)
    )
    rows = sorted(
        result.all(),
        key=lambda row: (
            _parse_tc_metadata(row[0]).get("process_area") or row[1].level or "General",
            row[0].tc_id,
        ),
    )

    for row_num, (test_case, requirement) in enumerate(rows, 2):
        meta = _parse_tc_metadata(test_case)
        process_area = meta.get("process_area") or requirement.level or "General"
        objective = meta.get("objective") or ""
        risk_rating = meta.get("risk_rating") or "MEDIUM"
        automation_status = meta.get("automation_status") or "AUTOMATION_BLOCKED"
        systems_involved = "; ".join(meta.get("systems_involved") or [])
        required_roles = "; ".join(meta.get("required_roles") or [])
        cleanup_instructions = "\n".join(meta.get("cleanup_instructions") or [])
        ambiguities = "\n".join(meta.get("ambiguities") or [])
        assumptions = "\n".join(meta.get("assumptions") or [])
        automation_blockers = "\n".join(meta.get("automation_blockers") or [])

        # Format steps with action and expected result per step
        steps_text = ""
        expected_result_text = ""
        if test_case.steps:
            for step in test_case.steps:
                step_no = step.get("step_no", "")
                action = step.get("action", "")
                expected = step.get("expected_result", "")
                steps_text += f"{step_no}. {action}\n"
                expected_result_text += f"{step_no}. {expected}\n"

        preconditions_text = "\n".join(test_case.preconditions or [])

        test_data_items: list[str] = []
        for step in (test_case.steps or []):
            td = step.get("test_data", "")
            if td and td not in test_data_items:
                test_data_items.append(td)
        test_data_text = "\n".join(test_data_items)

        data = [
            test_case.tc_id,
            requirement.req_id,
            process_area,
            test_case.title,
            objective,
            test_case.test_type,
            test_case.test_level or "INTEGRATION",
            test_case.priority or "P2",
            risk_rating,
            automation_status,
            systems_involved,
            required_roles,
            preconditions_text,
            test_data_text,
            steps_text,
            expected_result_text,
            cleanup_instructions,
            ambiguities,
            assumptions,
            automation_blockers,
            test_case.status or "DRAFT",
            "",  # Execution Cycle
            "",  # Actual Result
            "",  # Defect ID
        ]

        for col_num, value in enumerate(data, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            _format_cell(cell, value, header=False, wrap=True)

    column_widths = [12, 12, 15, 30, 25, 15, 12, 8, 10, 22, 20, 20, 22, 22, 30, 30, 22, 25, 25, 25, 10, 15, 20, 12]
    for col_num, width in enumerate(column_widths, 1):
        _set_column_width(sheet, col_num, width)


# Function: _create_requirements_traceability_sheet
async def _create_requirements_traceability_sheet(
    workbook: Workbook,
    session: AsyncSession,
    project_id: str,
) -> None:
    """Create Requirements Traceability Matrix sheet."""
    sheet = workbook.create_sheet("Requirements Traceability")
    
    headers = [
        "Requirement ID",
        "Requirement Description",
        "Level",
        "Priority",
        "Mapped Test Cases",
        "Test Case Count",
        "Coverage Status",
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        _format_cell(cell, header, header=True)
    
    sheet.freeze_panes = "A2"
    
    # Fetch requirements and count mapped test cases
    result = await session.execute(
        select(Requirement, func.count(TestCase.id))
        .outerjoin(TestCase, TestCase.requirement_id == Requirement.id)
        .where(Requirement.project_id == project_id)
        .group_by(Requirement.id)
        .order_by(Requirement.req_id)
    )
    rows = list(result.all())
    
    for row_num, (requirement, tc_count) in enumerate(rows, 2):
        # Get test case IDs for this requirement
        tc_result = await session.execute(
            select(TestCase.tc_id)
            .where(TestCase.requirement_id == requirement.id)
            .order_by(TestCase.tc_id)
        )
        tc_ids = [tc_id for (tc_id,) in tc_result.all()]
        mapped_tcs = ", ".join(tc_ids) if tc_ids else "No test cases"
        coverage_status = "Covered" if tc_count > 0 else "NOT COVERED"
        
        data = [
            requirement.req_id,
            requirement.statement[:200] + ("..." if len(requirement.statement) > 200 else ""),
            requirement.level,
            requirement.priority or "P2",
            mapped_tcs,
            tc_count or 0,
            coverage_status,
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            bg_color = "FFE6E6" if coverage_status == "NOT COVERED" else None
            if bg_color:
                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            _format_cell(cell, value, header=False, wrap=True)
    
    column_widths = [12, 40, 12, 8, 30, 15, 15]
    for col_num, width in enumerate(column_widths, 1):
        _set_column_width(sheet, col_num, width)


# Function: _create_test_data_sheet
async def _create_test_data_sheet(
    workbook: Workbook,
    session: AsyncSession,
    project_id: str,
) -> None:
    """Create Test Data sheet with example data sets."""
    sheet = workbook.create_sheet("Test Data")
    
    headers = [
        "Test Data Set ID",
        "Description",
        "Usage / Requirement",
        "Data Type",
        "Sample Value",
        "Constraints",
        "Owner",
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        _format_cell(cell, header, header=True)
    
    sheet.freeze_panes = "A2"
    
    # Fetch unique test data from test cases
    result = await session.execute(
        select(TestCase).where(TestCase.project_id == project_id).limit(100)
    )
    test_cases = list(result.scalars().all())
    
    # Extract test data patterns from steps
    test_data_sets = {}
    for tc in test_cases:
        if tc.steps:
            for step in tc.steps:
                test_data = step.get("test_data", "")
                if test_data and len(test_data) > 3:
                    key = f"TD-{len(test_data_sets) + 1:03d}"
                    if key not in test_data_sets:
                        test_data_sets[key] = {
                            "id": key,
                            "description": test_data[:50],
                            "usage": tc.title[:60],
                            "data_type": "String",
                            "sample": test_data[:30],
                            "constraints": "Provided by test case definition",
                            "owner": "QA Team",
                        }
    
    # Add sample data rows
    for row_num, (key, test_data_dict) in enumerate(test_data_sets.items(), 2):
        data = [
            test_data_dict["id"],
            test_data_dict["description"],
            test_data_dict["usage"],
            test_data_dict["data_type"],
            test_data_dict["sample"],
            test_data_dict["constraints"],
            test_data_dict["owner"],
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            _format_cell(cell, value, header=False, wrap=True)
    
    column_widths = [15, 25, 30, 12, 20, 20, 12]
    for col_num, width in enumerate(column_widths, 1):
        _set_column_width(sheet, col_num, width)


# Function: _create_defect_log_sheet
def _create_defect_log_sheet(workbook: Workbook) -> None:
    """Create Defect Log template sheet."""
    sheet = workbook.create_sheet("Defect Log")
    
    headers = [
        "Defect ID",
        "Test Case ID",
        "Requirement ID",
        "Severity",
        "Priority",
        "Title",
        "Description",
        "Steps to Reproduce",
        "Expected Behavior",
        "Actual Behavior",
        "Assigned To",
        "Status",
        "Linked Requirement",
        "Found Date",
        "Resolution Date",
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        _format_cell(cell, header, header=True)
    
    sheet.freeze_panes = "A2"
    
    # Add empty template rows
    for row_num in range(2, 52):  # 50 empty rows for defect tracking
        for col_num in range(1, len(headers) + 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    column_widths = [12, 12, 12, 10, 8, 20, 30, 25, 20, 20, 12, 10, 12, 12, 12]
    for col_num, width in enumerate(column_widths, 1):
        _set_column_width(sheet, col_num, width)


# Function: _create_test_plan_summary_sheet
async def _create_test_plan_summary_sheet(
    workbook: Workbook,
    session: AsyncSession,
    project_id: str,
) -> None:
    """Create Test Plan Summary sheet with high-level info."""
    sheet = workbook.create_sheet("Test Plan Summary", 0)  # First sheet
    
    # Fetch project and test plan info
    from traceforge.db.models import Project
    project = await session.get(Project, project_id)
    
    test_plan_result = await session.execute(
        select(TestPlan)
        .where(TestPlan.project_id == project_id)
        .order_by(TestPlan.created_at.desc())
        .limit(1)
    )
    test_plan = test_plan_result.scalars().first()
    
    # Count metrics
    tc_result = await session.execute(
        select(func.count(TestCase.id), TestCase.test_type)
        .where(TestCase.project_id == project_id)
        .group_by(TestCase.test_type)
    )
    tc_by_type = {test_type: count for count, test_type in tc_result.all()}
    
    req_result = await session.execute(
        select(func.count(Requirement.id))
        .where(Requirement.project_id == project_id, Requirement.status == "APPROVED")
    )
    approved_reqs = req_result.scalar() or 0
    
    # Add title
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = f"{project.name} — Test Plan Summary"
    title_cell.font = Font(bold=True, size=14)

    # Test level breakdown
    level_result = await session.execute(
        select(TestCase.test_level, func.count(TestCase.id))
        .where(TestCase.project_id == project_id)
        .group_by(TestCase.test_level)
    )
    tc_by_level = {level: count for level, count in level_result.all()}

    automation_result = await session.execute(
        select(TestCase.gherkin)
        .where(TestCase.project_id == project_id)
    )
    import json as _json
    automation_counts: dict[str, int] = {}
    for (gherkin_raw,) in automation_result.all():
        if gherkin_raw and gherkin_raw.strip().startswith("{"):
            try:
                status = _json.loads(gherkin_raw).get("automation_status", "UNKNOWN")
                automation_counts[status] = automation_counts.get(status, 0) + 1
            except Exception:
                pass

    metadata = [
        ("Project Name", project.name),
        ("Project Key", project.key),
        ("Client", project.client_name or "N/A"),
        ("Test Plan Version", str(test_plan.version) if test_plan else "N/A"),
        ("Test Plan Status", test_plan.status if test_plan else "N/A"),
        ("APPROVED Requirements", str(approved_reqs)),
        ("Total Test Cases", str(sum(tc_by_type.values()))),
        ("", ""),
        ("Test Case Breakdown by Type", ""),
        ("  POSITIVE", str(tc_by_type.get("POSITIVE", 0))),
        ("  NEGATIVE", str(tc_by_type.get("NEGATIVE", 0))),
        ("  EDGE", str(tc_by_type.get("EDGE", 0))),
        ("  BOUNDARY", str(tc_by_type.get("BOUNDARY", 0))),
        ("  NEGATIVE_SECURITY", str(tc_by_type.get("NEGATIVE_SECURITY", 0))),
        ("  PERFORMANCE", str(tc_by_type.get("PERFORMANCE", 0))),
        ("", ""),
        ("Test Level Breakdown", ""),
        ("  INTEGRATION", str(tc_by_level.get("INTEGRATION", 0))),
        ("  API", str(tc_by_level.get("API", 0))),
        ("  UI_E2E", str(tc_by_level.get("UI_E2E", 0))),
        ("  UAT", str(tc_by_level.get("UAT", 0))),
        ("  UNIT", str(tc_by_level.get("UNIT", 0))),
        ("", ""),
        ("Automation Readiness", ""),
        ("  AUTOMATION_BLOCKED", str(automation_counts.get("AUTOMATION_BLOCKED", 0))),
        ("  MANUAL_ONLY", str(automation_counts.get("MANUAL_ONLY", 0))),
        ("  READY_FOR_API_AUTOMATION", str(automation_counts.get("READY_FOR_API_AUTOMATION", 0))),
        ("  READY_FOR_UI_AUTOMATION", str(automation_counts.get("READY_FOR_UI_AUTOMATION", 0))),
        ("  READY_FOR_HYBRID_AUTOMATION", str(automation_counts.get("READY_FOR_HYBRID_AUTOMATION", 0))),
        ("", ""),
        ("NOTE", "All agent-generated cases start as DRAFT / Pending Business Review."),
        ("NOTE", "Resolve all [EXECUTION DETAIL BLOCKED] markers before execution."),
        ("NOTE", "Automation-blocked cases require business owner to supply application metadata."),
    ]

    for row_num, (key, value) in enumerate(metadata, 3):
        sheet.cell(row=row_num, column=1).value = key
        sheet.cell(row=row_num, column=2).value = value
        if key and key.endswith(("Breakdown", "Breakdown by Type", "Readiness")):
            for cell in [sheet.cell(row=row_num, column=1), sheet.cell(row=row_num, column=2)]:
                cell.font = Font(bold=True)

    _set_column_width(sheet, 1, 35)
    _set_column_width(sheet, 2, 50)


async def format_test_case_workbook(
    session: AsyncSession,
    project_id: str,
) -> bytes:
    """Create a comprehensive test case workbook with multiple sheets and rich metadata.
    
    Returns the workbook as bytes.
    """
    workbook = Workbook()
    
    # Remove default blank sheet
    if len(workbook.sheetnames) > 0:
        workbook.remove(workbook.active)
    
    # Create all sheets
    await _create_test_plan_summary_sheet(workbook, session, project_id)
    await _create_test_cases_sheet(workbook, session, project_id)
    await _create_requirements_traceability_sheet(workbook, session, project_id)
    await _create_source_coverage_sheet(workbook, session, project_id)
    await _create_test_data_sheet(workbook, session, project_id)
    await _create_ambiguity_register_sheet(workbook, session, project_id)
    await _create_coverage_gap_sheet(workbook, session, project_id)
    await _create_roles_access_sheet(workbook, session, project_id)
    await _create_automation_readiness_sheet(workbook, session, project_id)
    await _create_coverage_metrics_sheet(workbook, session, project_id)
    await _create_risk_assessment_sheet(workbook, session, project_id)
    await _create_interface_coverage_sheet(workbook, session, project_id)
    await _create_reconciliation_matrix_sheet(workbook, session, project_id)
    _create_defect_log_sheet(workbook)
    
    # Add Lists sheet
    sheet = workbook.create_sheet("Lists")
    lists_data = [
        ("Test Type", "Description"),
        ("POSITIVE", "Happy path and successful business flow scenarios"),
        ("NEGATIVE", "Error handling, validation rejection, and business-rule blocking"),
        ("EDGE", "Retry, concurrency, interruption, and recovery scenarios"),
        ("BOUNDARY", "Minimum, maximum, and exact-boundary value testing"),
        ("NEGATIVE_SECURITY", "Authorization, role enforcement, and access-control testing"),
        ("PERFORMANCE", "Load, throughput, and response-time testing"),
        ("", ""),
        ("Test Level", "Description"),
        ("INTEGRATION", "ERP/SAP, inter-system, accounting, authorization, and master-data testing"),
        ("API", "REST/SOAP endpoints, message queues, and interface adapters"),
        ("UAT", "Complete business journeys verified with business-approved test data"),
        ("UI_E2E", "UI-navigable workflows with stable screen/selector metadata"),
        ("UNIT", "Isolated calculation, validation, or transformation logic"),
        ("", ""),
        ("Automation Status", "Description"),
        ("AUTOMATION_BLOCKED", "Cannot be automated — missing URL, auth, selectors, or test-data API"),
        ("MANUAL_ONLY", "Requires a source-mandated physical, elapsed-time, human-presence, or regulatory activity"),
        ("READY_FOR_API_AUTOMATION", "All API metadata supplied — ready for automated execution"),
        ("READY_FOR_UI_AUTOMATION", "All UI metadata, selectors, and auth supplied — ready for automation"),
        ("READY_FOR_HYBRID_AUTOMATION", "Mixed API/UI automation — all metadata supplied"),
        ("", ""),
        ("Lifecycle Status", "Description"),
        ("DRAFT", "Agent-generated — requires business owner review before execution"),
        ("IN_REVIEW", "Under review by test lead or business owner"),
        ("APPROVED", "Approved for execution after business owner review"),
        ("REJECTED", "Rejected — requires rework"),
        ("", ""),
        ("Priority", "Description"),
        ("P1", "Critical — must pass for release"),
        ("P2", "High — should pass for release"),
        ("P3", "Medium — desirable but non-blocking"),
        ("", ""),
        ("Risk Rating", "Description"),
        ("HIGH", "Business-critical process or financial/compliance impact"),
        ("MEDIUM", "Significant functional impact if failed"),
        ("LOW", "Minor functional or cosmetic impact"),
    ]

    for row_num, (key, desc) in enumerate(lists_data, 1):
        sheet.cell(row=row_num, column=1).value = key
        sheet.cell(row=row_num, column=2).value = desc
        if key in ("Test Type", "Test Level", "Automation Status", "Lifecycle Status", "Priority", "Risk Rating"):
            for col_num in [1, 2]:
                sheet.cell(row=row_num, column=col_num).font = Font(bold=True)
        if key:
            for col_num in [1, 2]:
                sheet.cell(row=row_num, column=col_num).border = _BORDER

    _set_column_width(sheet, 1, 30)
    _set_column_width(sheet, 2, 60)
    
    # Save to bytes
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


async def _case_metadata_rows(session: AsyncSession, project_id: str):
    result = await session.execute(
        select(TestCase, Requirement)
        .join(Requirement, Requirement.id == TestCase.requirement_id)
        .where(TestCase.project_id == project_id)
        .order_by(TestCase.tc_id)
    )
    return [(tc, req, _parse_tc_metadata(tc)) for tc, req in result.all()]


async def _create_coverage_metrics_sheet(
    workbook: Workbook,
    session: AsyncSession,
    project_id: str,
) -> None:
    """Summarize generated coverage by business domain and scenario dimension."""
    aggregates: dict[tuple[str, str, str], dict[str, Any]] = {}
    for test_case, requirement, metadata in await _case_metadata_rows(session, project_id):
        process_area = metadata.get("process_area") or requirement.level or "General"
        coverage_dimension = metadata.get("coverage_dimension") or "UNCLASSIFIED"
        key = (process_area, coverage_dimension, test_case.test_type)
        aggregate = aggregates.setdefault(key, {"count": 0, "requirements": set()})
        aggregate["count"] += 1
        aggregate["requirements"].add(requirement.req_id)

    rows = [
        [process_area, coverage_dimension, test_type, values["count"],
         len(values["requirements"]), "; ".join(sorted(values["requirements"]))]
        for (process_area, coverage_dimension, test_type), values in sorted(aggregates.items())
    ]
    _write_matrix_sheet(workbook, "Coverage Metrics", [
        "Process Area", "Coverage Dimension", "Test Type", "Test Case Count",
        "Requirement Count", "Requirement IDs",
    ], rows)


async def _create_risk_assessment_sheet(
    workbook: Workbook,
    session: AsyncSession,
    project_id: str,
) -> None:
    """Summarize business risk and automation readiness by process area."""
    aggregates: dict[str, dict[str, int]] = {}
    for _, requirement, metadata in await _case_metadata_rows(session, project_id):
        process_area = metadata.get("process_area") or requirement.level or "General"
        values = aggregates.setdefault(process_area, {
            "total": 0,
            "high": 0,
            "blocked": 0,
            "manual": 0,
            "ready": 0,
            "blockers": 0,
        })
        risk_rating = metadata.get("risk_rating") or "MEDIUM"
        automation_status = metadata.get("automation_status") or "AUTOMATION_BLOCKED"
        values["total"] += 1
        values["high"] += int(risk_rating == "HIGH")
        values["blocked"] += int(automation_status == "AUTOMATION_BLOCKED")
        values["manual"] += int(automation_status == "MANUAL_ONLY")
        values["ready"] += int(automation_status.startswith("READY_FOR_"))
        values["blockers"] += len(metadata.get("automation_blockers") or [])

    rows = []
    for process_area, values in sorted(aggregates.items()):
        status = "ACTION REQUIRED" if values["high"] or values["blocked"] else "MONITOR"
        rows.append([
            process_area,
            values["total"],
            values["high"],
            values["blocked"],
            values["manual"],
            values["ready"],
            values["blockers"],
            status,
        ])
    _write_matrix_sheet(workbook, "Risk Assessment", [
        "Process Area", "Total Test Cases", "High Risk", "Automation Blocked",
        "Manual Only", "Automation Ready", "Open Blockers", "Assessment",
    ], rows)


async def _create_source_coverage_sheet(workbook: Workbook, session: AsyncSession, project_id: str) -> None:
    chunks = list((await session.scalars(
        select(Chunk).where(Chunk.project_id == project_id).order_by(Chunk.ordinal)
    )).all())
    citation_rows = (await session.execute(
        select(SourceCitation.chunk_id, Requirement.req_id)
        .join(Requirement, Requirement.id == SourceCitation.requirement_id)
        .where(Requirement.project_id == project_id)
    )).all()
    requirements_by_chunk: dict[str, set[str]] = {}
    for chunk_id, req_id in citation_rows:
        requirements_by_chunk.setdefault(str(chunk_id), set()).add(req_id)
    rows = []
    for chunk in chunks:
        mapped = sorted(requirements_by_chunk.get(str(chunk.id), set()))
        rows.append([
            str(chunk.id), chunk.ordinal, json.dumps(chunk.locator or {}, ensure_ascii=False),
            "; ".join(mapped), "MAPPED" if mapped else "UNMAPPED — EXTRACTION REVIEW REQUIRED",
            chunk.text[:500],
        ])
    _write_matrix_sheet(workbook, "Source Coverage", [
        "Chunk ID", "Ordinal", "Source Locator", "Requirement IDs", "Coverage Status", "Source Excerpt",
    ], rows)


def _write_matrix_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    sheet = workbook.create_sheet(title)
    for col_num, header in enumerate(headers, 1):
        _format_cell(sheet.cell(row=1, column=col_num), header, header=True)
    for row_num, row in enumerate(rows, 2):
        for col_num, value in enumerate(row, 1):
            _format_cell(sheet.cell(row=row_num, column=col_num), value, wrap=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    for col_num in range(1, len(headers) + 1):
        _set_column_width(sheet, col_num, 24 if col_num > 2 else 16)


async def _create_roles_access_sheet(workbook: Workbook, session: AsyncSession, project_id: str) -> None:
    rows = []
    for tc, req, meta in await _case_metadata_rows(session, project_id):
        roles = meta.get("required_roles") or []
        if not roles:
            rows.append([tc.tc_id, req.req_id, "[PENDING BUSINESS CONFIRMATION]", tc.test_type, "OPEN"])
        else:
            rows.extend([tc.tc_id, req.req_id, role, tc.test_type, "SOURCE CONFIRMED"] for role in roles)
    _write_matrix_sheet(workbook, "Roles & Access", ["Test Case ID", "Requirement ID", "Required Role", "Control Type", "Status"], rows)


async def _create_automation_readiness_sheet(workbook: Workbook, session: AsyncSession, project_id: str) -> None:
    rows = []
    for tc, req, meta in await _case_metadata_rows(session, project_id):
        context = meta.get("automation_context") or {}
        blockers = meta.get("automation_blockers") or []
        rows.append([
            tc.tc_id, req.req_id, tc.test_level,
            meta.get("automation_status") or "AUTOMATION_BLOCKED",
            "YES" if context.get("base_url") else "NO",
            "YES" if context.get("auth") else "NO",
            "YES" if context.get("locators") or context.get("endpoints") else "NO",
            "YES" if context.get("assertions") or context.get("schemas") else "NO",
            "YES" if context.get("test_data_factory") else "NO",
            "YES" if context.get("cleanup") else "NO",
            "; ".join(blockers),
        ])
    _write_matrix_sheet(workbook, "Automation Readiness", [
        "Test Case ID", "Requirement ID", "Test Level", "Readiness", "Base URL", "Authentication",
        "Selectors/Endpoints", "Assertions/Schemas", "Test Data Factory", "Cleanup", "Blockers",
    ], rows)


async def _create_interface_coverage_sheet(workbook: Workbook, session: AsyncSession, project_id: str) -> None:
    rows = []
    for tc, req, meta in await _case_metadata_rows(session, project_id):
        systems = meta.get("systems_involved") or []
        context = meta.get("automation_context") or {}
        if tc.test_level in {"API", "INTEGRATION"} or len(systems) > 1:
            rows.append([
                tc.tc_id, req.req_id, "; ".join(systems) or "[PENDING]", tc.test_level,
                "; ".join((context.get("endpoints") or {}).keys()) if isinstance(context.get("endpoints"), dict) else "[PENDING]",
                "DEFINED" if context.get("schemas") else "PENDING",
            ])
    _write_matrix_sheet(workbook, "Interface Coverage", ["Test Case ID", "Requirement ID", "Systems", "Test Level", "Interfaces", "Contract Status"], rows)


async def _create_reconciliation_matrix_sheet(workbook: Workbook, session: AsyncSession, project_id: str) -> None:
    rows = []
    terms = ("reconcil", "balance", "stock", "inventory", "accounting", "quantity", "document flow")
    for tc, req, meta in await _case_metadata_rows(session, project_id):
        evidence = " ".join([tc.title, req.statement] + [str(v) for step in (tc.steps or []) for v in step.values()])
        if any(term in evidence.lower() for term in terms):
            expected = "; ".join(str(step.get("expected_result", "")) for step in (tc.steps or []))
            rows.append([tc.tc_id, req.req_id, meta.get("process_area") or req.level, expected, "SOURCE CONFIRMED" if expected else "PENDING"])
    _write_matrix_sheet(workbook, "Reconciliation Matrix", ["Test Case ID", "Requirement ID", "Process Area", "Reconciliation Assertion", "Rule Status"], rows)


async def _create_ambiguity_register_sheet(
    workbook: Workbook,
    session: AsyncSession,
    project_id: str,
) -> None:
    """Ambiguity Register — open questions that block test case approval."""
    sheet = workbook.create_sheet("Ambiguity Register")
    headers = [
        "Ambiguity ID", "Requirement ID(s)", "Test Case ID(s)", "Description",
        "Business Owner", "Decision Required", "Blocking Execution?",
        "Blocking Automation?", "Status", "Resolution",
    ]
    for col_num, header in enumerate(headers, 1):
        _format_cell(sheet.cell(row=1, column=col_num), header, header=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Collect ambiguities from test case metadata
    result = await session.execute(
        select(TestCase, Requirement)
        .join(Requirement, Requirement.id == TestCase.requirement_id)
        .where(TestCase.project_id == project_id)
        .order_by(TestCase.tc_id)
    )
    seen_ambiguities: dict[str, dict] = {}
    for test_case, requirement in result.all():
        meta = _parse_tc_metadata(test_case)
        for amb in meta.get("ambiguities") or []:
            key = amb[:100]
            if key not in seen_ambiguities:
                seen_ambiguities[key] = {
                    "req_ids": set(), "tc_ids": set(), "description": amb,
                }
            seen_ambiguities[key]["req_ids"].add(requirement.req_id)
            seen_ambiguities[key]["tc_ids"].add(test_case.tc_id)

    row_num = 2
    for i, (key, entry) in enumerate(seen_ambiguities.items(), 1):
        data = [
            f"AMB-{i:04d}",
            "; ".join(sorted(entry["req_ids"])),
            "; ".join(sorted(entry["tc_ids"])),
            entry["description"],
            "[PENDING]",
            "Business owner must supply the missing detail.",
            "YES",
            "YES",
            "OPEN",
            "",
        ]
        for col_num, value in enumerate(data, 1):
            _format_cell(sheet.cell(row=row_num, column=col_num), value, header=False, wrap=True)
        row_num += 1

    widths = [12, 20, 20, 50, 25, 40, 20, 20, 12, 30]
    for col_num, width in enumerate(widths, 1):
        _set_column_width(sheet, col_num, width)


async def _create_coverage_gap_sheet(
    workbook: Workbook,
    session: AsyncSession,
    project_id: str,
) -> None:
    """Coverage Gap Register — requirements and business areas without test coverage."""
    sheet = workbook.create_sheet("Coverage Gaps")
    headers = [
        "Gap ID", "Requirement ID", "Business Area", "Gap Description",
        "Missing Test Types", "Root Cause", "Priority", "Action Required",
    ]
    for col_num, header in enumerate(headers, 1):
        _format_cell(sheet.cell(row=1, column=col_num), header, header=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Fetch requirements and their test case coverage
    req_result = await session.execute(
        select(Requirement)
        .where(Requirement.project_id == project_id, Requirement.status == "APPROVED")
        .order_by(Requirement.req_id)
    )
    requirements = list(req_result.scalars().all())

    tc_result = await session.execute(
        select(TestCase.requirement_id, TestCase.test_type, func.count(TestCase.id))
        .where(TestCase.project_id == project_id)
        .group_by(TestCase.requirement_id, TestCase.test_type)
    )
    coverage: dict[str, dict[str, int]] = {}
    for req_id, tc_type, count in tc_result.all():
        requirement_coverage = coverage.setdefault(str(req_id), {})
        requirement_coverage[tc_type] = count
        if tc_type == "NEGATIVE_SECURITY":
            requirement_coverage["NEGATIVE"] = requirement_coverage.get("NEGATIVE", 0) + count

    row_num = 2
    for req in requirements:
        covered = coverage.get(str(req.id), {})
        targets = minimum_scenarios_for_requirement(req)
        missing = [
            f"{test_type} ({covered.get(test_type, 0)}/{minimum})"
            for test_type, minimum in targets.items()
            if covered.get(test_type, 0) < minimum
        ]
        if missing:
            data = [
                f"GAP-{row_num - 1:04d}", req.req_id, req.level,
                f"Incomplete scenario coverage for: {req.title}",
                ", ".join(missing),
                "LLM generation did not produce all required scenario types.",
                "P2",
                f"Generate additional coverage for {', '.join(missing)} for {req.req_id}.",
            ]
            for col_num, value in enumerate(data, 1):
                _format_cell(sheet.cell(row=row_num, column=col_num), value, header=False, wrap=True)
            row_num += 1

    widths = [12, 15, 20, 50, 30, 40, 8, 40]
    for col_num, width in enumerate(widths, 1):
        _set_column_width(sheet, col_num, width)
