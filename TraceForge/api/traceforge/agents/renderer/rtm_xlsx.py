# ---------------------------------------------------------------------------
# Author: Vishnuu A
# Scope: §5.1 Agent 5 — RTM (Excel, openpyxl). Every derived column is a live formula, never
# Date: 2026-04-07
# ---------------------------------------------------------------------------
"""§5.1 Agent 5 — RTM (Excel, openpyxl). Every derived column is a live formula, never
a hardcoded value — spec's own acceptance test: 'opening the RTM and deleting a row
causes all coverage percentages to recalculate correctly.'

One deviation from the spec's literal formula text, documented where it happens: the
Test Count column returns numeric 0 (not the string "GAP") when a requirement has no
test cases, so the downstream Coverage Status formula's `IF(K2=0, ...)` numeric
comparison actually fires — the spec's two example formulas don't compose correctly
taken 100% literally (a text "GAP" would make `K2=0` evaluate to FALSE in Excel,
silently skipping the "NO TESTS" branch). "GAP" as a concept is still surfaced, just
one level down in Coverage Status ("NO TESTS") rather than duplicated in Test Count.
"""
from __future__ import annotations

import uuid

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from traceforge.db.models import (
    AuditEvent, Chunk, Requirement, SourceCitation, SourceDocument, TestCase, TestScript,
)
from traceforge.agents.coverage_policy import minimum_scenarios_for_requirement, requirement_is_executable
from traceforge.agents.script_gen.playwright import _parse_tc_metadata, _verified_automation_status

_HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_MAX_ROW_REF = 5000  # formula ranges are bounded, not whole-column, for perf while still comfortably covering real projects

_BRD_SECTION_BY_LEVEL = {
    "BUSINESS": "4. Business Requirements", "FUNCTIONAL": "5. Functional Requirements",
    "NON_FUNCTIONAL": "6. Non-Functional Requirements", "ASSUMPTION": "7. Assumptions & Dependencies",
    "CONSTRAINT": "7. Assumptions & Dependencies",
}


# Function: _style_header
def _style_header(ws, columns: list[str]) -> None:
    for i, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"


# Function: _load_data
async def _load_data(session: AsyncSession, project_id: uuid.UUID):
    requirements = list((await session.execute(
        select(Requirement).where(Requirement.project_id == project_id).order_by(Requirement.req_id)
    )).scalars().all())
    test_cases = list((await session.execute(
        select(TestCase).where(TestCase.project_id == project_id).order_by(TestCase.tc_id)
    )).scalars().all())
    scripts = list((await session.execute(
        select(TestScript).where(TestScript.project_id == project_id).order_by(TestScript.ts_id)
    )).scalars().all())
    audit = list((await session.execute(
        select(AuditEvent).where(AuditEvent.project_id == project_id).order_by(AuditEvent.at)
    )).scalars().all())

    tc_by_id = {tc.id: tc for tc in test_cases}
    citations_by_req: dict[uuid.UUID, list[SourceCitation]] = {}
    for req in requirements:
        result = await session.execute(
            select(SourceCitation, Chunk, SourceDocument)
            .join(Chunk, SourceCitation.chunk_id == Chunk.id)
            .join(SourceDocument, Chunk.source_document_id == SourceDocument.id)
            .where(SourceCitation.requirement_id == req.id)
        )
        citations_by_req[req.id] = list(result.all())

    return requirements, test_cases, scripts, audit, tc_by_id, citations_by_req


# Function: _build_rtm_sheet
def _build_rtm_sheet(wb: Workbook, requirements: list[Requirement], citations_by_req: dict) -> None:
    ws = wb.active
    ws.title = "RTM"
    columns = ["REQ-ID", "Level", "Statement", "Priority", "EARS Pattern", "Ambiguity Score",
               "Source Document(s)", "Source Locator(s)", "BRD Section", "Testability", "AC Count",
               "Test Case IDs", "Test Count", "Positive Count", "Negative Count", "Edge Count",
               "Boundary Count", "Security Count", "Performance Count", "Integration Count",
               "Dedicated AC Count", "Test Design Status", "Reviewed Count", "Automation Ready Count",
               "Manual Count", "Automation Blocked Count", "Script IDs", "Script Count",
               "Automation Status", "Approval Status"]
    _style_header(ws, columns)

    for row_i, req in enumerate(requirements, start=2):
        rows = citations_by_req.get(req.id, [])
        source_docs = ", ".join(sorted({sd.filename for _, _, sd in rows})) or "(none)"
        locators = "; ".join(f"{sd.filename}: {c.quoted_span[:40]}..." for c, _, sd in rows[:3]) or "(none)"

        ws.cell(row=row_i, column=1, value=req.req_id).hyperlink = "#Requirements!A1"
        ws.cell(row=row_i, column=2, value=req.level)
        ws.cell(row=row_i, column=3, value=req.statement)
        ws.cell(row=row_i, column=4, value=req.priority)
        ws.cell(row=row_i, column=5, value=req.ears_pattern)
        ws.cell(row=row_i, column=6, value=round(req.ambiguity_score, 2))
        ws.cell(row=row_i, column=7, value=source_docs)
        ws.cell(row=row_i, column=8, value=locators)
        ws.cell(row=row_i, column=9, value=_BRD_SECTION_BY_LEVEL.get(req.level, ""))
        testable = requirement_is_executable(req)
        minima = minimum_scenarios_for_requirement(req) if testable else {}
        integration_required = int(any(word in " ".join([
            req.statement, *(req.acceptance_criteria or []),
        ]).lower() for word in (
            "reconcil", "integration", "interface", "handoff", "inter-system",
            "external system", "sync", "workflow",
        )))
        ws.cell(row=row_i, column=10, value="EXECUTABLE" if testable else "INFORMATION GAP")
        ws.cell(row=row_i, column=11, value=len(req.acceptance_criteria or []))
        ws.cell(row=row_i, column=12,
                value=f'=TEXTJOIN(", ",TRUE,IF(TestCases!$B$2:$B${_MAX_ROW_REF}=A{row_i},TestCases!$A$2:$A${_MAX_ROW_REF},""))')
        ws.cell(row=row_i, column=13, value=f'=IF(L{row_i}="",0,COUNTIF(TestCases!$B$2:$B${_MAX_ROW_REF},A{row_i}))')
        ws.cell(row=row_i, column=14, value=f'=COUNTIFS(TestCases!$B$2:$B${_MAX_ROW_REF},A{row_i},TestCases!$C$2:$C${_MAX_ROW_REF},"POSITIVE")')
        ws.cell(row=row_i, column=15, value=f'=COUNTIFS(TestCases!$B$2:$B${_MAX_ROW_REF},A{row_i},TestCases!$C$2:$C${_MAX_ROW_REF},"NEGATIVE")+COUNTIFS(TestCases!$B$2:$B${_MAX_ROW_REF},A{row_i},TestCases!$C$2:$C${_MAX_ROW_REF},"NEGATIVE_SECURITY")')
        ws.cell(row=row_i, column=16, value=f'=COUNTIFS(TestCases!$B$2:$B${_MAX_ROW_REF},A{row_i},TestCases!$C$2:$C${_MAX_ROW_REF},"EDGE")')
        ws.cell(row=row_i, column=17, value=f'=COUNTIFS(TestCases!$B$2:$B${_MAX_ROW_REF},A{row_i},TestCases!$C$2:$C${_MAX_ROW_REF},"BOUNDARY")')
        ws.cell(row=row_i, column=18, value=f'=COUNTIFS(TestCases!$B$2:$B${_MAX_ROW_REF},A{row_i},TestCases!$C$2:$C${_MAX_ROW_REF},"NEGATIVE_SECURITY")')
        ws.cell(row=row_i, column=19, value=f'=COUNTIFS(TestCases!$B$2:$B${_MAX_ROW_REF},A{row_i},TestCases!$C$2:$C${_MAX_ROW_REF},"PERFORMANCE")')
        ws.cell(row=row_i, column=20, value=(
            f'=COUNTIFS(TestCases!$B$2:$B${_MAX_ROW_REF},A{row_i},TestCases!$D$2:$D${_MAX_ROW_REF},"INTEGRATION")+'
            f'COUNTIFS(TestCases!$B$2:$B${_MAX_ROW_REF},A{row_i},TestCases!$K$2:$K${_MAX_ROW_REF},"INTEGRATION_HANDOFF")+'
            f'COUNTIFS(TestCases!$B$2:$B${_MAX_ROW_REF},A{row_i},TestCases!$K$2:$K${_MAX_ROW_REF},"RECONCILIATION")+'
            f'COUNTIFS(TestCases!$B$2:$B${_MAX_ROW_REF},A{row_i},TestCases!$K$2:$K${_MAX_ROW_REF},"END_TO_END")'
        ))
        ac_terms = [
            f'COUNTIFS(TestCases!$B$2:$B${_MAX_ROW_REF},A{row_i},TestCases!$J$2:$J${_MAX_ROW_REF},"{number}")'
            for number in range(1, len(req.acceptance_criteria or []) + 1)
        ]
        ws.cell(row=row_i, column=21, value="=" + "+".join(ac_terms) if ac_terms else "=0")
        policy_checks = [
            f'N{row_i}>={minima.get("POSITIVE", 0)}', f'O{row_i}>={minima.get("NEGATIVE", 0)}',
            f'P{row_i}>={minima.get("EDGE", 0)}', f'Q{row_i}>={minima.get("BOUNDARY", 0)}',
            f'R{row_i}>={minima.get("NEGATIVE_SECURITY", 0)}', f'S{row_i}>={minima.get("PERFORMANCE", 0)}',
            f'T{row_i}>={integration_required}',
            *[
                f'COUNTIFS(TestCases!$B$2:$B${_MAX_ROW_REF},A{row_i},TestCases!$J$2:$J${_MAX_ROW_REF},"{number}")>=1'
                for number in range(1, len(req.acceptance_criteria or []) + 1)
            ],
        ]
        ws.cell(row=row_i, column=22, value=(
            f'=IF(J{row_i}="INFORMATION GAP","INFORMATION GAP",IF(M{row_i}=0,"NO TESTS",'
            f'IF(AND({",".join(policy_checks)}),"TEST DESIGNED","POLICY GAPS")))'
        ))
        ws.cell(row=row_i, column=23, value=f'=COUNTIFS(TestCases!$B$2:$B${_MAX_ROW_REF},A{row_i},TestCases!$G$2:$G${_MAX_ROW_REF},"APPROVED")')
        ws.cell(row=row_i, column=24, value=f'=COUNTIFS(TestCases!$B$2:$B${_MAX_ROW_REF},A{row_i},TestCases!$I$2:$I${_MAX_ROW_REF},"READY_FOR_UI_AUTOMATION")')
        ws.cell(row=row_i, column=25, value=f'=COUNTIFS(TestCases!$B$2:$B${_MAX_ROW_REF},A{row_i},TestCases!$I$2:$I${_MAX_ROW_REF},"MANUAL_ONLY")')
        ws.cell(row=row_i, column=26, value=f'=MAX(0,M{row_i}-X{row_i}-Y{row_i})')
        ws.cell(row=row_i, column=27,
                value=f'=TEXTJOIN(", ",TRUE,IF(Scripts!$C$2:$C${_MAX_ROW_REF}=A{row_i},Scripts!$A$2:$A${_MAX_ROW_REF},""))')
        ws.cell(row=row_i, column=28, value=f'=COUNTIFS(Scripts!$C$2:$C${_MAX_ROW_REF},A{row_i},Scripts!$J$2:$J${_MAX_ROW_REF},"YES")')
        ws.cell(row=row_i, column=29, value=(
            f'=IF(J{row_i}="INFORMATION GAP","NOT APPLICABLE",IF(X{row_i}=0,'
            f'IF(AND(Y{row_i}>0,Y{row_i}=M{row_i}),"MANUAL ONLY","AUTOMATION BLOCKED"),'
            f'IF(AB{row_i}>=X{row_i},"SCRIPTED",IF(AB{row_i}>0,"PARTIALLY SCRIPTED","READY FOR SCRIPT"))))'
        ))
        ws.cell(row=row_i, column=30, value=req.status)

    last_row = max(2, len(requirements) + 1)
    ws.conditional_formatting.add(
        f"F2:F{last_row}", CellIsRule(operator="lessThan", formula=["0.2"], fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        f"F2:F{last_row}", CellIsRule(operator="between", formula=["0.2", "0.4"], fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        f"F2:F{last_row}", CellIsRule(operator="greaterThan", formula=["0.4"], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )
    for i, width in enumerate([12, 14, 60, 8, 16, 10, 30, 40, 28, 16, 9, 20, 10, 10, 10, 10, 10, 10, 11, 11, 12, 18, 12, 14, 10, 14, 20, 10, 20, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


# Function: _build_testcases_sheet
def _build_testcases_sheet(wb: Workbook, test_cases: list[TestCase], req_id_by_uuid: dict) -> None:
    ws = wb.create_sheet("TestCases")
    columns = ["TC-ID", "REQ-ID", "Test Type", "Test Level", "Priority", "Title", "Status", "Steps (count)",
               "Verified Automation Status", "Dedicated AC Mapping", "Coverage Dimension", "Content Hash"]
    _style_header(ws, columns)
    for row_i, tc in enumerate(test_cases, start=2):
        ws.cell(row=row_i, column=1, value=tc.tc_id)
        ws.cell(row=row_i, column=2, value=req_id_by_uuid.get(tc.requirement_id, ""))
        ws.cell(row=row_i, column=3, value=tc.test_type)
        ws.cell(row=row_i, column=4, value=tc.test_level)
        ws.cell(row=row_i, column=5, value=tc.priority)
        ws.cell(row=row_i, column=6, value=tc.title)
        ws.cell(row=row_i, column=7, value=tc.status)
        ws.cell(row=row_i, column=8, value=len(tc.steps or []))
        metadata = _parse_tc_metadata(tc)
        verified_status, _ = _verified_automation_status(tc, metadata)
        ws.cell(row=row_i, column=9, value=verified_status)
        mapping = metadata.get("acceptance_criteria_mapped") or []
        ws.cell(row=row_i, column=10, value=str(mapping[0]) if len(mapping) == 1 else "")
        ws.cell(row=row_i, column=11, value=metadata.get("coverage_dimension") or "")
        ws.cell(row=row_i, column=12, value=tc.content_hash)
    for i, width in enumerate([12, 12, 14, 12, 8, 50, 12, 14, 24, 18, 22, 20], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


# Function: _build_scripts_sheet
def _build_scripts_sheet(wb: Workbook, scripts: list[TestScript], tc_by_id: dict, req_id_by_uuid: dict) -> None:
    ws = wb.create_sheet("Scripts")
    columns = ["TS-ID", "TC-ID", "REQ-ID", "Target", "File Path", "Compiles", "Status",
               "Upstream TC Hash", "Current TC Hash", "Valid Current Script"]
    _style_header(ws, columns)
    for row_i, ts in enumerate(scripts, start=2):
        tc = tc_by_id.get(ts.test_case_id)
        if ts.compiles is True:
            compile_status = "Yes"
        elif ts.compiles is False:
            compile_status = "No"
        else:
            compile_status = "Not validated"
        ws.cell(row=row_i, column=1, value=ts.ts_id)
        ws.cell(row=row_i, column=2, value=tc.tc_id if tc else "")
        ws.cell(row=row_i, column=3, value=req_id_by_uuid.get(tc.requirement_id, "") if tc else "")
        ws.cell(row=row_i, column=4, value=ts.target)
        ws.cell(row=row_i, column=5, value=ts.file_path)
        ws.cell(row=row_i, column=6, value=compile_status)
        ws.cell(row=row_i, column=7, value=ts.status)
        ws.cell(row=row_i, column=8, value=ts.upstream_tc_hash)
        ws.cell(row=row_i, column=9, value=f'=IFERROR(XLOOKUP(B{row_i},TestCases!$A$2:$A${_MAX_ROW_REF},TestCases!$L$2:$L${_MAX_ROW_REF}),"")')
        ws.cell(row=row_i, column=10, value=f'=IF(AND(G{row_i}<>"REJECTED",G{row_i}<>"SUSPECT",H{row_i}=I{row_i}),"YES","NO")')
    for i, width in enumerate([12, 12, 12, 16, 40, 12, 12, 20, 20, 18], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


# Function: _build_coverage_summary_sheet
def _build_coverage_summary_sheet(wb: Workbook, requirements: list[Requirement]) -> None:
    ws = wb.create_sheet("Coverage Summary")
    ws.cell(row=1, column=1, value="Coverage by Level").font = Font(bold=True, size=13)
    columns = ["Level", "Baseline", "Executable", "Test Designed", "Test Design Coverage %", "Information Gaps"]
    for i, col in enumerate(columns, start=1):
        ws.cell(row=3, column=i, value=col).font = Font(bold=True)

    levels = ["BUSINESS", "FUNCTIONAL", "NON_FUNCTIONAL", "ASSUMPTION", "CONSTRAINT"]
    last_rtm_row = max(2, len(requirements) + 1)
    for row_i, level in enumerate(levels, start=4):
        ws.cell(row=row_i, column=1, value=level)
        ws.cell(row=row_i, column=2, value=f'=COUNTIF(RTM!$B$2:$B${last_rtm_row},A{row_i})')
        ws.cell(row=row_i, column=3, value=f'=COUNTIFS(RTM!$B$2:$B${last_rtm_row},A{row_i},RTM!$J$2:$J${last_rtm_row},"EXECUTABLE")')
        ws.cell(row=row_i, column=4, value=f'=COUNTIFS(RTM!$B$2:$B${last_rtm_row},A{row_i},RTM!$V$2:$V${last_rtm_row},"TEST DESIGNED")')
        ws.cell(row=row_i, column=5, value=f'=IF(C{row_i}=0,"N/A",D{row_i}/C{row_i})')
        ws.cell(row=row_i, column=5).number_format = "0%"
        ws.cell(row=row_i, column=6, value=f'=COUNTIFS(RTM!$B$2:$B${last_rtm_row},A{row_i},RTM!$J$2:$J${last_rtm_row},"INFORMATION GAP")')

    summary_row = 4 + len(levels) + 1
    ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=f'=SUM(B4:B{3 + len(levels)})')
    ws.cell(row=summary_row, column=3, value=f'=SUM(C4:C{3 + len(levels)})')
    ws.cell(row=summary_row, column=4, value=f'=SUM(D4:D{3 + len(levels)})')
    ws.cell(row=summary_row, column=5, value=f'=IF(C{summary_row}=0,"N/A",D{summary_row}/C{summary_row})')
    ws.cell(row=summary_row, column=5).number_format = "0%"
    ws.cell(row=summary_row, column=6, value=f'=SUM(F4:F{3 + len(levels)})')

    ambiguity_row = summary_row + 2
    ws.cell(row=ambiguity_row, column=1, value="Ambiguity distribution (score > 0.4)").font = Font(bold=True)
    ws.cell(row=ambiguity_row + 1, column=1, value="Blocked requirements")
    ws.cell(row=ambiguity_row + 1, column=2, value=f'=COUNTIF(RTM!$F$2:$F${last_rtm_row},">0.4")')

    for i, width in enumerate([16, 10, 12, 14, 22, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


# Function: _build_gaps_sheet
def _build_gaps_sheet(wb: Workbook, requirements: list[Requirement]) -> None:
    ws = wb.create_sheet("Gaps")
    ws.cell(row=1, column=1, value="Executable requirements with actionable test-design gaps — live view (Excel 365 FILTER)").font = Font(bold=True)
    last_rtm_row = max(2, len(requirements) + 1)
    ws.cell(row=2, column=1, value=f'=IFERROR(FILTER(RTM!A2:AD{last_rtm_row},(RTM!V2:V{last_rtm_row}="NO TESTS")+(RTM!V2:V{last_rtm_row}="POLICY GAPS")),"No actionable test-design gaps.")')
    ws.column_dimensions["A"].width = 100


# Function: _build_audit_sheet
def _build_audit_sheet(wb: Workbook, audit_events: list[AuditEvent]) -> None:
    ws = wb.create_sheet("Audit")
    columns = ["At", "Actor", "Action", "Entity Type", "Entity ID", "Rationale/Detail"]
    _style_header(ws, columns)
    for row_i, event in enumerate(audit_events, start=2):
        detail = ""
        if event.after and isinstance(event.after, dict):
            detail = str(event.after.get("rationale") or event.after)[:200]
        ws.cell(row=row_i, column=1, value=event.at.strftime("%Y-%m-%d %H:%M:%S") if event.at else "")
        ws.cell(row=row_i, column=2, value=event.actor)
        ws.cell(row=row_i, column=3, value=event.action)
        ws.cell(row=row_i, column=4, value=event.entity_type)
        ws.cell(row=row_i, column=5, value=event.entity_id)
        ws.cell(row=row_i, column=6, value=detail)
    for i, width in enumerate([20, 20, 24, 16, 14, 60], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


# Function: render_rtm_xlsx
async def render_rtm_xlsx(session: AsyncSession, project_id: uuid.UUID, output_path: str) -> None:
    requirements, test_cases, scripts, audit, tc_by_id, citations_by_req = await _load_data(session, project_id)
    req_id_by_uuid = {r.id: r.req_id for r in requirements}

    wb = Workbook()
    _build_rtm_sheet(wb, requirements, citations_by_req)
    _build_testcases_sheet(wb, test_cases, req_id_by_uuid)
    _build_scripts_sheet(wb, scripts, tc_by_id, req_id_by_uuid)
    _build_coverage_summary_sheet(wb, requirements)
    _build_gaps_sheet(wb, requirements)
    _build_audit_sheet(wb, audit)
    wb.save(output_path)
