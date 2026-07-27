# ---------------------------------------------------------------------------
# Author: Vishnuu A
# Scope: §5 Agent 3 (Test Designer) output — real this pass.
# Date: 2026-02-18
# ---------------------------------------------------------------------------
"""§5 Agent 3 (Test Designer) output — real this pass."""
from __future__ import annotations

import io
import re
import uuid

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from traceforge.auth import current_user
from traceforge.db.models import AuditEvent, Project, Requirement, TestCase, TestPlan
from traceforge.db.session import get_session
from traceforge.schemas.testcase import TestCaseOut, TestCasePatch, TestPlanOut

router = APIRouter(prefix="/api/v1", tags=["testcases"])


def _safe_filename(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "-", value).strip("-") or "traceforge"


def _download_response(content: bytes, *, media_type: str, filename: str) -> Response:
    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _plan_markdown(project: Project, plan: TestPlan) -> str:
    criteria = plan.entry_exit_criteria or {}
    schedule = plan.schedule or {}

    def bullets(values) -> str:
        return "\n".join(f"- {value}" for value in (values or [])) or "- Not specified"

    return (
        f"# {plan.title}\n\n"
        f"**Project:** {project.key} — {project.name}  \n"
        f"**Status:** {plan.status}  \n"
        f"**Version:** {plan.version}\n\n"
        f"## Scope\n\n{plan.scope}\n\n"
        f"## Strategy\n\n{plan.strategy}\n\n"
        f"## Environments\n\n{bullets(plan.environments)}\n\n"
        f"## Schedule\n\n{bullets(schedule.get('phases', []))}\n\n"
        f"## Entry Criteria\n\n{bullets(criteria.get('entry', []))}\n\n"
        f"## Exit Criteria\n\n{bullets(criteria.get('exit', []))}\n"
    )


def _format_test_case_workbook(rows: list[tuple[TestCase, Requirement]]) -> bytes:
    workbook = Workbook()
    cases_sheet = workbook.active
    cases_sheet.title = "Test Cases"
    steps_sheet = workbook.create_sheet("Test Steps")
    cases_headers = [
        "Test Case ID", "Requirement ID", "Requirement Statement", "Title", "Type",
        "Level", "Priority", "Status", "Preconditions", "Step Count", "Gherkin", "Version",
    ]
    step_headers = [
        "Test Case ID", "Requirement ID", "Step No", "Action", "Expected Result", "Test Data",
    ]
    cases_sheet.append(cases_headers)
    steps_sheet.append(step_headers)

    for test_case, requirement in rows:
        cases_sheet.append([
            test_case.tc_id,
            requirement.req_id,
            requirement.statement,
            test_case.title,
            test_case.test_type,
            test_case.test_level,
            test_case.priority,
            test_case.status,
            "\n".join(str(item) for item in (test_case.preconditions or [])),
            len(test_case.steps or []),
            test_case.gherkin or "",
            test_case.version,
        ])
        for step in test_case.steps or []:
            steps_sheet.append([
                test_case.tc_id,
                requirement.req_id,
                step.get("step_no", ""),
                step.get("action", ""),
                step.get("expected_result", ""),
                step.get("test_data", ""),
            ])

    header_fill = PatternFill("solid", fgColor="1D4ED8")
    for sheet in (cases_sheet, steps_sheet):
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column in range(1, sheet.max_column + 1):
            width = max(
                len(str(sheet.cell(row=row, column=column).value or ""))
                for row in range(1, min(sheet.max_row, 100) + 1)
            )
            sheet.column_dimensions[get_column_letter(column)].width = min(max(width + 2, 12), 60)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


# Function: list_testcases
@router.get("/projects/{project_id}/testcases", response_model=list[TestCaseOut])
async def list_testcases(
    project_id: uuid.UUID, requirement_id: uuid.UUID | None = None, test_type: str | None = None, status: str | None = None,
    session: AsyncSession = Depends(get_session), user: dict = Depends(current_user),
):
    stmt = select(TestCase).where(TestCase.project_id == project_id)
    if requirement_id:
        stmt = stmt.where(TestCase.requirement_id == requirement_id)
    if test_type:
        stmt = stmt.where(TestCase.test_type == test_type)
    if status:
        stmt = stmt.where(TestCase.status == status)
    result = await session.execute(stmt.order_by(TestCase.tc_id))
    return list(result.scalars().all())


# Function: get_test_plan
@router.get("/projects/{project_id}/test-plan", response_model=TestPlanOut | None)
async def get_test_plan(project_id: uuid.UUID, session: AsyncSession = Depends(get_session), user: dict = Depends(current_user)):
    result = await session.execute(select(TestPlan).where(TestPlan.project_id == project_id).order_by(TestPlan.created_at.desc()))
    plan = result.scalars().first()
    # An absent plan is normal before TEST_DESIGN runs. JSON null keeps that
    # expected empty state from appearing as a failed request in the browser.
    return plan


@router.get("/projects/{project_id}/test-plan/download")
async def download_test_plan(
    project_id: uuid.UUID,
    session: AsyncSession = Depends(get_session),
    user: dict = Depends(current_user),
):
    project = await session.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")
    plan = await session.scalar(
        select(TestPlan).where(TestPlan.project_id == project_id).order_by(TestPlan.created_at.desc()).limit(1)
    )
    if plan is None:
        raise HTTPException(status_code=404, detail="No Test Plan is available to download.")
    filename = f"{_safe_filename(project.key)}-test-plan-v{plan.version}.md"
    return _download_response(
        _plan_markdown(project, plan).encode("utf-8"),
        media_type="text/markdown; charset=utf-8",
        filename=filename,
    )


@router.get("/projects/{project_id}/testcases/download")
async def download_test_cases(
    project_id: uuid.UUID,
    session: AsyncSession = Depends(get_session),
    user: dict = Depends(current_user),
):
    project = await session.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")
    rows = list((await session.execute(
        select(TestCase, Requirement)
        .join(Requirement, Requirement.id == TestCase.requirement_id)
        .where(TestCase.project_id == project_id)
        .order_by(TestCase.tc_id)
    )).all())
    if not rows:
        raise HTTPException(status_code=404, detail="No Test Cases are available to download.")
    filename = f"{_safe_filename(project.key)}-test-cases.xlsx"
    return _download_response(
        _format_test_case_workbook(rows),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )


# Function: patch_testcase
@router.patch("/testcases/{tc_id}", response_model=TestCaseOut)
async def patch_testcase(tc_id: uuid.UUID, body: TestCasePatch, session: AsyncSession = Depends(get_session), user: dict = Depends(current_user)):
    tc = await session.get(TestCase, tc_id)
    if not tc:
        raise HTTPException(status_code=404, detail="Test case not found")
    before = {"status": tc.status, "title": tc.title}
    updates = body.model_dump(exclude_unset=True)
    for field, value in updates.items():
        setattr(tc, field, value)
    if "steps" in updates or "title" in updates:
        tc.created_by_agent = False
        tc.version += 1
    session.add(AuditEvent(project_id=tc.project_id, actor=user.get("username", "unknown"), action="TESTCASE_EDITED",
                            entity_type="TestCase", entity_id=str(tc.id), before=before, after=updates))
    await session.commit()
    await session.refresh(tc)
    return tc
