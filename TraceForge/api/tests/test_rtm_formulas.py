# ---------------------------------------------------------------------------
# Author: Vishnuu A
# Scope: §10 Phase 2 acceptance (spec's own test): 'opening the RTM and deleting a row causes
# Date: 2026-04-30
# ---------------------------------------------------------------------------
"""§10 Phase 2 acceptance (spec's own test): 'opening the RTM and deleting a row causes
all coverage percentages to recalculate correctly — no hardcoded derived values
anywhere in the workbook.' We can't drive real Excel recalculation headlessly here, so
this asserts the structural half of that promise: every derived cell is stored as a
live formula string (openpyxl preserves '=...' text verbatim), never a plain number."""
from __future__ import annotations

import uuid

from openpyxl import load_workbook

from traceforge.agents.renderer.rtm_xlsx import render_rtm_xlsx
from traceforge.db.models import Chunk, Requirement, SourceCitation, SourceDocument


# Function: _make_requirement
async def _make_requirement(session, project, req_id: str):
    doc = SourceDocument(
        project_id=project.id, source_type="UPLOAD", filename="t.docx", blob_uri="/tmp/t.docx",
        sha256=uuid.uuid4().hex.ljust(64, "0"), doc_class="AS_IS_DOC", status="INDEXED",
    )
    session.add(doc)
    await session.flush()
    chunk = Chunk(source_document_id=doc.id, project_id=project.id, ordinal=0, text="shall do X.", token_count=3, locator={})
    session.add(chunk)
    await session.flush()

    requirement = Requirement(
        req_id=req_id, project_id=project.id, level="FUNCTIONAL", title="t", statement="The system shall do X.",
        ears_pattern="UBIQUITOUS", ears_parts={"system_name": "System"}, acceptance_criteria=["X happens"],
        priority="SHOULD", ambiguity_score=0.1, ambiguity_flags=[], status="APPROVED",
        content_hash="h", version=1, created_by_agent=True,
    )
    session.add(requirement)
    await session.flush()
    session.add(SourceCitation(requirement_id=requirement.id, chunk_id=chunk.id, relevance=1.0, quoted_span="shall do X."))
    await session.commit()
    return requirement


# Function: test_rtm_derived_columns_are_live_formulas_not_hardcoded
async def test_rtm_derived_columns_are_live_formulas_not_hardcoded(session, project, tmp_path):
    await _make_requirement(session, project, f"REQ-{uuid.uuid4().hex[:6]}")

    output_path = tmp_path / "RTM.xlsx"
    await render_rtm_xlsx(session, project.id, str(output_path))

    wb = load_workbook(str(output_path))
    assert set(wb.sheetnames) == {"RTM", "TestCases", "Scripts", "Coverage Summary", "Gaps", "Audit"}

    rtm = wb["RTM"]
    # Requirement source fields remain static; all test-design, review, automation,
    # and script result fields are formulas so workbook edits recalculate.
    for col in ("L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB", "AC"):
        cell = rtm[f"{col}2"]
        assert isinstance(cell.value, str) and cell.value.startswith("="), f"RTM!{col}2 is not a live formula: {cell.value!r}"
    assert rtm["J2"].value == "EXECUTABLE"
    assert "TEST DESIGNED" in rtm["V2"].value
    assert "Scripts!$J$2:$J$5000" in rtm["AB2"].value

    summary = wb["Coverage Summary"]
    assert str(summary["B4"].value).startswith("=COUNTIF(")
    assert str(summary["C4"].value).startswith("=COUNTIFS(")
    assert str(summary["D4"].value).startswith("=COUNTIFS(")
    assert str(summary["E4"].value).startswith("=IF(")

    gaps = wb["Gaps"]
    assert str(gaps["A2"].value).startswith("=IFERROR(FILTER(")
    assert '="NO TESTS"' in gaps["A2"].value
    assert '="POLICY GAPS"' in gaps["A2"].value


async def test_rtm_classifies_assumptions_as_information_gaps(session, project, tmp_path):
    requirement = await _make_requirement(session, project, f"REQ-{uuid.uuid4().hex[:6]}")
    requirement.level = "ASSUMPTION"
    requirement.acceptance_criteria = []
    await session.commit()

    output_path = tmp_path / "RTM-assumption.xlsx"
    await render_rtm_xlsx(session, project.id, str(output_path))

    rtm = load_workbook(str(output_path))["RTM"]
    assert rtm["J2"].value == "INFORMATION GAP"
    assert '"INFORMATION GAP"' in rtm["V2"].value
    assert '"NOT APPLICABLE"' in rtm["AC2"].value
