# ---------------------------------------------------------------------------
# Author: Vishnuu A
# Scope: TraceForge — api/tests (test_fast_pipeline.py)
# Date: 2026-04-08
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# Scope: TraceForge — optimized project-scale generation path
# ---------------------------------------------------------------------------
from __future__ import annotations

import asyncio
import io
import json
import re
import time
import zipfile
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook
from sqlalchemy import select

from traceforge.agents.doc_author import BRD_DEFINITION, run_doc_author
from traceforge.agents.script_gen.runner import run_script_generator
from traceforge.agents.test_designer import run_test_designer
from traceforge.config import TEST_DESIGN_CONCURRENCY
from traceforge.db.models import Chunk, Requirement, SourceCitation, SourceDocument
from traceforge.db.models import TestCase as TestCaseModel
from traceforge.db.models import TestScript as TestScriptModel
from traceforge.llm.ollama import OllamaProvider
from traceforge.llm.provider import LLMResponse
from traceforge.routers.scripts import download_project_scripts, download_script
from traceforge.routers.testcases import download_test_cases, download_test_plan
from traceforge.workers.tasks import _extract_resume_offset


def test_extract_resume_starts_over_when_requirements_were_cleared():
    prior_run = SimpleNamespace(stats={"chunks_processed": 7})

    assert _extract_resume_offset(prior_run, 0, 10) == 0
    assert _extract_resume_offset(prior_run, 3, 10) == 7


# Function: test_pipeline_uses_ollama_for_test_cases_and_playwright_scripts
async def test_pipeline_uses_ollama_for_test_cases_and_playwright_scripts(session, project, monkeypatch):
    ollama_calls: list[dict] = []
    active_ollama_calls = 0
    max_concurrent_ollama_calls = 0

    async def fake_ollama_generate(
        self, system, user, *, temperature, max_tokens, json_mode=True, progress=None,
    ):
        nonlocal active_ollama_calls, max_concurrent_ollama_calls
        active_ollama_calls += 1
        max_concurrent_ollama_calls = max(max_concurrent_ollama_calls, active_ollama_calls)
        await asyncio.sleep(0.01)
        active_ollama_calls -= 1
        ollama_calls.append({"system": system, "user": user, "json_mode": json_mode})
        if json_mode:
            if "enterprise test plan" in system:
                return LLMResponse(
                    text=json.dumps({
                        "scope": "Validate the cited invoice requirements.",
                        "strategy": "Use source-grounded functional tests.",
                        "environments": ["PENDING BUSINESS CONFIRMATION"],
                        "test_levels": ["INTEGRATION"],
                        "test_types": ["Functional"],
                        "schedule": {"phases": ["PENDING BUSINESS CONFIRMATION"]},
                        "entry_criteria": ["PENDING BUSINESS CONFIRMATION"],
                        "exit_criteria": ["PENDING BUSINESS CONFIRMATION"],
                        "suspension_criteria": ["PENDING BUSINESS CONFIRMATION"],
                        "risks": ["PENDING BUSINESS CONFIRMATION"],
                    }),
                    model="ollama-test-model", prompt_tokens=100,
                    completion_tokens=100, latency_ms=1,
                )
            if "planning semantic automation scripts" in system:
                tc_ids = re.findall(r"(?m)^- (TC-\d+)", user)
                text = json.dumps({"scripts": [
                    {"tc_id": tc_id, "scenario": f"Playwright scenario for {tc_id}"}
                    for tc_id in tc_ids
                ]})
                return LLMResponse(
                    text=text, model="ollama-test-model", prompt_tokens=100,
                    completion_tokens=100, latency_ms=1,
                )
            if "compact semantic JSON only" in system:
                steps_section = user.split("STEPS:\n", 1)[-1].split(
                    "\n\nReturn the semantic step plan", 1,
                )[0]
                step_numbers = re.findall(r"(?m)^(\d+)\.", steps_section)
                text = json.dumps({"steps": [
                    {
                        "step_no": int(number),
                        "action": "Execute the reviewed action",
                        "expected": "Verify the reviewed expected result",
                        "data": {"fixture": "worker-scoped"} if index == 0 else "worker-scoped fixture",
                        # Deliberately omit scenario to cover the production failure.
                    }
                    for index, number in enumerate(step_numbers)
                ]})
                return LLMResponse(
                    text=text, model="ollama-test-model", prompt_tokens=100,
                    completion_tokens=100, latency_ms=1,
                )
            scenario_types = (
                "POSITIVE", "POSITIVE", "POSITIVE",
                "NEGATIVE", "NEGATIVE", "NEGATIVE",
                "EDGE", "EDGE",
            )
            if "compact coverage dimensions only" in system:
                invoice_number = re.search(r"Invoice type (\d+)", system).group(1)
                text = json.dumps({"scenarios": [
                    {
                        "title": f"{test_type.title()} invoice scenario {index}",
                        "test_type": test_type.lower(),
                        "objective": (
                            f"Invoice type {invoice_number} is accepted when valid"
                            if test_type == "POSITIVE" else "Invalid input is rejected"
                        ),
                        "test_data": f"Worker-scoped {test_type.lower()} invoice fixture",
                        "acceptance_criteria": [1] if index == 1 else [2] if index == 2 else [],
                        "source_quote": (
                            f"Invoice type {invoice_number} is accepted when valid."
                            if test_type == "POSITIVE"
                            else "Invalid input is rejected."
                        ),
                        "coverage_dimension": (
                            "BUSINESS_RULE" if test_type == "POSITIVE" else
                            "NEGATIVE_CONTROL" if test_type == "NEGATIVE" else "EDGE_CONDITION"
                        ),
                        "priority": "P1" if test_type == "POSITIVE" else "P2",
                    }
                    for index, test_type in enumerate(scenario_types, start=1)
                ]})
                return LLMResponse(
                    text=text, model="ollama-test-model", prompt_tokens=100,
                    completion_tokens=100, latency_ms=1,
                )
            requested = re.search(r"additional (POSITIVE|NEGATIVE|EDGE|BOUNDARY|NEGATIVE_SECURITY|PERFORMANCE)", user)
            test_type = requested.group(1) if requested else "POSITIVE"
            text = json.dumps({"test_cases": [{
                "title": f"{test_type.title()} invoice validation",
                "test_type": test_type.lower(),
                "test_level": "INTEGRATION",
                "priority": "P1" if test_type == "POSITIVE" else "P2",
                "preconditions": ["The invoice requirement is approved."],
                "steps": [
                    {
                        "step_no": step_no,
                        "action": "Exercise the cited invoice validation condition.",
                        "expected_result": "Invoice type is accepted when valid and invalid input is rejected.",
                        "test_data": f"Source-grounded {test_type.lower()} invoice data",
                    }
                    for step_no in range(1, 5)
                ],
            }]})
        else:
            steps_section = user.split("STEPS:\n", 1)[-1].split("\n\nWrite the test body", 1)[0]
            step_numbers = re.findall(r"(?m)^(\d+)\.", steps_section)
            blocks = []
            for number in step_numbers:
                blocks.append(
                    "await test.step('Reviewed step', async () => {\n"
                    "  await executeReviewedStep(page, {\n"
                    "    action: 'Execute the reviewed action',\n"
                    "    expected: 'Verify the reviewed expected result',\n"
                    "    data: 'Use the reviewed worker-scoped data',\n"
                    "    scenario: 'Ollama-authored Playwright scenario',\n"
                    "  });\n"
                    "});"
                )
            text = "\n".join(blocks)
        return LLMResponse(
            text=text, model="ollama-test-model", prompt_tokens=100,
            completion_tokens=100, latency_ms=1,
        )

    async def fake_validate_typescript(code):
        assert "@playwright/test" in code
        return True, "Validated in the focused compiler test."

    async def fake_unload(self):
        return None

    monkeypatch.setattr(OllamaProvider, "generate", fake_ollama_generate)
    monkeypatch.setattr(OllamaProvider, "unload", fake_unload)
    monkeypatch.setattr(
        "traceforge.agents.script_gen.runner.validate_typescript", fake_validate_typescript,
    )
    document = SourceDocument(
        project_id=project.id, source_type="UPLOAD", filename="performance.txt", blob_uri="/tmp/performance.txt",
        sha256="8" * 64, doc_class="AS_IS_DOC", status="INDEXED",
    )
    session.add(document)
    await session.flush()
    chunk = Chunk(
        source_document_id=document.id, project_id=project.id, ordinal=0,
        text=(
            "The platform shall validate and submit invoices. "
            + " ".join(
                f"Invoice type {number} is accepted when valid. Invalid input is rejected."
                for number in range(1, 6)
            )
        ),
        token_count=60, locator={"section": "1"},
    )
    session.add(chunk)
    await session.flush()
    for index in range(5):
        requirement = Requirement(
            req_id=f"REQ-{index + 1:04d}", project_id=project.id, level="FUNCTIONAL",
            title=f"Validate invoice {index + 1}", statement=f"The platform shall validate invoice type {index + 1}.",
            ears_pattern="UBIQUITOUS", ears_parts={"system_name": "Platform"}, rationale="Source requirement",
            acceptance_criteria=[f"Invoice type {index + 1} is accepted when valid.", "Invalid input is rejected."],
            priority="MUST", ambiguity_score=0.0, ambiguity_flags=[], status="APPROVED",
            content_hash=f"{index + 1:064x}", version=1, created_by_agent=True,
        )
        session.add(requirement)
        await session.flush()
        session.add(SourceCitation(
            requirement_id=requirement.id, chunk_id=chunk.id, relevance=1.0,
            quoted_span="The platform shall validate and submit invoices.",
        ))
    await session.commit()

    started = time.perf_counter()
    design = await run_test_designer(session, project_id=project.id, pipeline_run_id=None)
    design_seconds = time.perf_counter() - started
    assert design.test_cases_created == 31
    assert design_seconds < 5
    assert any(call["json_mode"] is True for call in ollama_calls)
    assert any("evidence-first test analyst" in call["system"] for call in ollama_calls)
    detailed_category_calls = [
        call for call in ollama_calls
        if call["json_mode"] is True and call["user"].startswith("Generate exactly 1 additional")
    ]
    assert len(detailed_category_calls) == 0
    assert len([call for call in ollama_calls if "compact coverage dimensions only" in call["system"]]) == 5
    assert TEST_DESIGN_CONCURRENCY == 1
    assert max_concurrent_ollama_calls <= TEST_DESIGN_CONCURRENCY

    test_cases = list((await session.scalars(select(TestCaseModel).where(TestCaseModel.project_id == project.id))).all())
    assert all(test_case.test_level in {"UNIT", "API", "UI_E2E", "INTEGRATION", "UAT"} for test_case in test_cases)
    assert {test_case.test_level for test_case in test_cases} == {"INTEGRATION", "UAT"}
    assert all(test_case.status == "DRAFT" for test_case in test_cases)
    for requirement_id in {test_case.requirement_id for test_case in test_cases}:
        requirement_cases = [test_case for test_case in test_cases if test_case.requirement_id == requirement_id]
        assert sum(test_case.test_type == "POSITIVE" for test_case in requirement_cases) >= 1
        assert sum(test_case.test_type == "NEGATIVE" for test_case in requirement_cases) >= 1
    for test_case in test_cases:
        test_case.status = "APPROVED"
    await session.commit()

    plan_download = await download_test_plan(
        project.id, session=session, user={"username": "tester"},
    )
    assert plan_download.media_type.startswith("text/markdown")
    assert b"# Test Project Test Plan" in plan_download.body
    assert "test-plan" in plan_download.headers["content-disposition"]

    cases_download = await download_test_cases(
        project.id, session=session, user={"username": "tester"},
    )
    assert cases_download.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    workbook = load_workbook(io.BytesIO(cases_download.body), read_only=True)
    assert {
        "Test Plan Summary", "Test Cases", "Requirements Traceability", "Source Coverage", "Test Data",
        "Ambiguity Register", "Coverage Gaps", "Roles & Access", "Automation Readiness",
        "Coverage Metrics", "Risk Assessment", "Interface Coverage", "Reconciliation Matrix",
    }.issubset(workbook.sheetnames)
    assert workbook["Test Cases"].max_row == len(test_cases) + 1
    headers = [cell.value for cell in next(workbook["Test Cases"].iter_rows(max_row=1))]
    assert {"Test Steps", "Expected Result"}.issubset(headers)
    coverage_headers = [cell.value for cell in next(workbook["Coverage Metrics"].iter_rows(max_row=1))]
    assert {"Process Area", "Coverage Dimension", "Test Case Count"}.issubset(coverage_headers)

    started = time.perf_counter()
    scripts = await run_script_generator(session, project_id=project.id, pipeline_run_id=None)
    script_seconds = time.perf_counter() - started
    assert scripts["scripts_created"] == 0
    assert script_seconds < 5
    generated_scripts = list((await session.scalars(
        select(TestScriptModel).where(TestScriptModel.project_id == project.id)
    )).all())
    assert generated_scripts == []

    suite_download = await download_project_scripts(project.id, session=session, user={"username": "tester"})
    assert suite_download.media_type == "application/zip"
    with zipfile.ZipFile(io.BytesIO(suite_download.body)) as suite:
        names = suite.namelist()
        assert "package.json" in names
        assert "playwright.config.ts" in names
        assert "traceforge-manifest.json" in names
        assert len([name for name in names if name.endswith(".spec.ts")]) == 0
        manifest = json.loads(suite.read("traceforge-manifest.json"))
        assert len(manifest) == len(test_cases)
        assert all(entry["compiles"] is None and entry["syntax_status"] == "NOT_APPLICABLE" for entry in manifest)
        assert all(entry["runnable"] is False for entry in manifest)
        assert all(entry["automation_status"] == "AUTOMATION_BLOCKED" for entry in manifest)
        assert all(entry["excluded_from_playwright"] is True for entry in manifest)

    # Regeneration updates the same logical scripts instead of appending duplicates.
    rerun = await run_script_generator(session, project_id=project.id, pipeline_run_id=None)
    assert rerun["scripts_inserted"] == 0
    assert rerun["scripts_updated"] == 0
    assert len(list((await session.scalars(
        select(TestScriptModel).where(TestScriptModel.project_id == project.id)
    )).all())) == 0

    started = time.perf_counter()
    artifact = await run_doc_author(session, project_id=project.id, definition=BRD_DEFINITION, pipeline_run_id=None)
    document_seconds = time.perf_counter() - started
    assert Path(artifact.blob_uri).exists()
    assert document_seconds < 5
    Path(artifact.blob_uri).unlink(missing_ok=True)
