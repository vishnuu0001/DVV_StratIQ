from __future__ import annotations

import json
import uuid

from openpyxl import Workbook

from traceforge.agents.test_case_excel_generator import _create_requirements_traceability_sheet
from traceforge.db.models import Requirement, TestCase


def _requirement(project_id, req_id: str, *, level: str = "FUNCTIONAL", acceptance_criteria=None):
    return Requirement(
        req_id=req_id, project_id=project_id, level=level, title="Source-backed outcome",
        statement="The system shall produce the source-backed outcome.", ears_pattern="UBIQUITOUS",
        ears_parts={}, acceptance_criteria=acceptance_criteria if acceptance_criteria is not None else ["The outcome is produced."],
        priority="MUST", ambiguity_score=0, ambiguity_flags=[], conflict_flags=[], status="APPROVED",
        content_hash=uuid.uuid4().hex.ljust(64, "0"),
    )


def _test_case(project_id, requirement, number: int, test_type: str):
    return TestCase(
        tc_id=f"TC-{number:04d}", project_id=project_id, requirement_id=requirement.id,
        title=f"Source-backed {test_type.lower()} outcome {number}", test_type=test_type,
        test_level="INTEGRATION", preconditions=[], steps=[{
            "step_no": 1, "action": "Exercise the source-backed behavior.",
            "expected_result": "The outcome is produced.",
        }],
        gherkin=json.dumps({
            "automation_status": "AUTOMATION_BLOCKED",
            "acceptance_criteria_mapped": [1] if number == 1 else [],
        }),
        priority="P2", status="DRAFT", upstream_req_hash=requirement.content_hash,
        content_hash=uuid.uuid4().hex.ljust(64, "0"), created_by_agent=True,
    )


async def test_test_case_workbook_separates_design_coverage_from_information_gaps(session, project):
    executable = _requirement(project.id, "REQ-EXEC")
    information_gap = _requirement(project.id, "REQ-INFO", level="ASSUMPTION", acceptance_criteria=[])
    session.add_all([executable, information_gap])
    await session.flush()
    session.add_all([
        *[_test_case(project.id, executable, number, "POSITIVE") for number in range(1, 4)],
        *[_test_case(project.id, executable, number, "NEGATIVE") for number in range(4, 6)],
    ])
    await session.flush()

    workbook = Workbook()
    await _create_requirements_traceability_sheet(workbook, session, str(project.id))
    sheet = workbook["Requirements Traceability"]
    rows = {sheet.cell(row=row, column=1).value: row for row in range(2, sheet.max_row + 1)}

    assert sheet.cell(row=rows["REQ-EXEC"], column=8).value == "TEST DESIGNED"
    assert sheet.cell(row=rows["REQ-INFO"], column=7).value == "INFORMATION GAP"
    assert sheet.cell(row=rows["REQ-INFO"], column=8).value == "INFORMATION GAP"