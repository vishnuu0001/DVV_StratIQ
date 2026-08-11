from openpyxl import load_workbook

wb = load_workbook(r"C:\RD\DVV_StratIQ-Aqorynth\dsvstratiq\TraceForge\data\TestData\Outbound_E2E_01_Detailed_Test_Cases.xlsx")

# List all sheets
print("Available sheets:", wb.sheetnames)

# Check the Test Cases sheet
if 'Test Cases' in wb.sheetnames:
    ws = wb['Test Cases']
    print("\n=== TEST CASES SHEET ===")
    print(f"Dimensions: {ws.dimensions}")
    print(f"Max row: {ws.max_row}")
    
    # Get headers
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
    print(f"\nHeaders ({len(headers)}): {headers}")
    
    # Show first 15 test cases
    print("\n\nFirst 20 test cases:")
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=21, values_only=True)):
        tc_id = row[0]
        scenario = row[1]
        process = row[2]
        priority = row[5] if len(row) > 5 else None
        status = row[6] if len(row) > 6 else None
        print(f"{i+1}. TC={tc_id} | Scenario: {str(scenario)[:55] if scenario else 'N/A'} | Proc: {process} | Pri: {priority} | Status: {status}")
    
    # Count by process area
    process_areas = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        process = row[2] if len(row) > 2 else None
        if process:
            process_areas[process] = process_areas.get(process, 0) + 1
    
    print("\n\nTest cases by process area:")
    total = 0
    for area in sorted(process_areas.keys()):
        count = process_areas[area]
        total += count
        print(f"  {area}: {count}")
    print(f"Total: {total}")
